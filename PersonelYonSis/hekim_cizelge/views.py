from datetime import timedelta
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
import json
import calendar
from datetime import datetime
from django.template.loader import get_template
import pdfkit
from django.conf import settings
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils.timezone import now

from hekim_cizelge.forms import BildirimForm
from hekim_cizelge.utils import hesapla_bildirim_verileri, hesapla_fazla_mesai, hesapla_icap_suresi
from .models import Bildirim, Izin, MesaiKontrol, Personel, Hizmet, Birim, PersonelBirim, Mesai, ResmiTatil, UserBirim  # Removed MesaiOnay
from PersonelYonSis.models import User
from django.db import models

# PDFKit yapılandırması
config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')

def hizmet_tanimlari(request):
    hizmetler = Hizmet.objects.all()
    return render(request, 'hekim_cizelge/hizmet_tanimlari.html', {"hizmetler": hizmetler})

def add_hizmet(request):
    if request.method == 'POST':
        try:
            hizmet_name = request.POST.get('hizmet_name')
            hizmet_tipi = request.POST.get('hizmet_tipi')
            hafta_ici_suresi = request.POST.get('hizmet_suresi_hafta_ici')
            hafta_sonu_suresi = request.POST.get('hizmet_suresi_hafta_sonu')
            max_hekim = request.POST.get('max_hekim', 1)
            nobet_ertesi_izinli = bool(request.POST.get('nobet_ertesi_izinli'))
            varsayilan_hizmetle_sunulur = bool(request.POST.get('varsayilan_hizmetle_sunulur'))

            # Süreleri dakikaya çevir (saat:dakika formatından)
            def sure_to_dakika(sure_str):
                if not sure_str:
                    return None
                saat, dakika = map(int, sure_str.split(':'))
                return saat * 60 + dakika

            hafta_ici_dakika = sure_to_dakika(hafta_ici_suresi)
            hafta_sonu_dakika = sure_to_dakika(hafta_sonu_suresi) if hafta_sonu_suresi else None

            if hafta_ici_dakika is None:
                raise ValueError("Hafta içi süresi gereklidir")

            Hizmet.objects.create(
                HizmetName=hizmet_name,
                HizmetTipi=hizmet_tipi,
                HizmetSuresiHaftaIci=hafta_ici_dakika,
                HizmetSuresiHaftaSonu=hafta_sonu_dakika,
                MaxHekimSayisi=max_hekim,
                NobetErtesiIzinli=nobet_ertesi_izinli,
                VarsayilanHizmetleSunulur=varsayilan_hizmetle_sunulur
            )

            messages.success(request, "Hizmet başarıyla eklendi.")
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Hizmet eklenirken bir hata oluştu: {str(e)}")

        return redirect('hekim_cizelge:hizmet_tanimlari')

    return redirect('hekim_cizelge:hizmet_tanimlari')

def personeller(request):
    personeller = Personel.objects.all()
    birimler = Birim.objects.all()

    # Personel ile birim bilgilerini eşleştiriyoruz
    personel_birimleri = PersonelBirim.objects.select_related('personel', 'birim')

    return render(request, 'hekim_cizelge/personeller.html', {
        'personeller': personeller,
        'birimler': birimler,
        'personel_birimleri': personel_birimleri
    })

# Personel Ekleme Formunu Geri Döndür
def personel_ekle_form(request):
    return render(request, 'hekim_cizelge/personel_ekle_form.html')
# Yeni Personel Ekleme İşlemi
def personel_ekle(request):
    if request.method == 'POST':
        # Form verilerini al
        personel_id = request.POST['PersonelID']
        personel_name = request.POST['PersonelName']
        personel_title = request.POST['PersonelTitle']
        personel_branch = request.POST['PersonelBranch']

        # Yeni personel kaydet
        personel = Personel(
            PersonelID=personel_id,
            PersonelName=personel_name,
            PersonelTitle=personel_title,
            PersonelBranch=personel_branch
        )
        personel.save()

        # Başarı mesajı ekleyebilirsiniz
        return redirect('hekim_cizelge:personeller')  # Personel listesine yönlendir
    return HttpResponse("Geçersiz istek", status=400)
# Personel Adı, Unvanı, Branşı ve Çalıştığı Birim bilgilerini güncelleme işlemi
def personel_update(request):
    if request.method == 'POST':
        personel_id = request.POST.get('id')  # Formdan gelen personel ID'si
        birim_id = request.POST.get('birim')  # Formdan gelen birim ID'si
        first_name = request.POST.get('first_name')  # Formdan gelen personel adı
        last_name = request.POST.get('last_name')  # Formdan gelen personel soyadı
        personel_title = request.POST.get('title')  # Formdan gelen personel unvanı
        personel_branch = request.POST.get('branch')  # Formdan gelen personel branşı

        # Modellerden ilgili kayıtları çekiyoruz
        personel = Personel.objects.get(PersonelID=personel_id)
        birim = Birim.objects.get(BirimID=birim_id)
        
        # Personel bilgilerini güncelle
        personel.FirstName = first_name
        personel.LastName = last_name
        personel.PersonelTitle = personel_title
        personel.PersonelBranch = personel_branch
        personel.save()
        
        # PersonelBirim kaydını güncelle veya oluştur
        personel_birim, created = PersonelBirim.objects.update_or_create(
            personel=personel,
            defaults={'birim': birim}
        )

        return JsonResponse({'status': 'success', 'birim_adi': birim.BirimAdi})
    return JsonResponse({'status': 'error', 'message': 'Hata oluştu.'})

def birim_tanimlari(request):
    birimler = Birim.objects.prefetch_related('DigerHizmetler').select_related('VarsayilanHizmet').all()
    hizmetler = Hizmet.objects.all()
    return render(request, 'hekim_cizelge/birim_tanimlari.html', {
        'birimler': birimler,
        'hizmetler': hizmetler
    })
def add_birim(request):
    if request.method == 'POST':
        birim_adi = request.POST.get('birim_adi')
        varsayilan_hizmet_id = request.POST.get('varsayilan_hizmet')
        diger_hizmetler_ids = request.POST.getlist('diger_hizmetler')

        # VarsayilanHizmet ve diger_hizmetler kontrolü
        if not varsayilan_hizmet_id:
            messages.error(request, "Varsayılan hizmet seçilmelidir.")
            return redirect('hekim_cizelge:birim_tanimlari')
        
        if not birim_adi:
            messages.error(request, "Birim adı boş bırakılamaz.")
            return redirect('hekim_cizelge:birim_tanimlari')

        # Yeni birim oluşturma
        try:
            varsayilan_hizmet = Hizmet.objects.get(HizmetID=varsayilan_hizmet_id)
            yeni_birim = Birim.objects.create(
                BirimAdi=birim_adi,
                VarsayilanHizmet=varsayilan_hizmet
            )

            # Diğer hizmetleri ManyToMany ilişkisinde ekleme
            if diger_hizmetler_ids:
                diger_hizmetler = Hizmet.objects.filter(HizmetID__in=diger_hizmetler_ids)
                yeni_birim.DigerHizmetler.set(diger_hizmetler)

            messages.success(request, "Birim başarıyla eklendi.")
        except Hizmet.DoesNotExist:
            messages.error(request, "Seçilen varsayılan hizmet bulunamadı.")
        except Exception as e:
            messages.error(request, f"Birim eklenirken bir hata oluştu: {str(e)}")

        return redirect('hekim_cizelge:birim_tanimlari')

    # GET isteği durumunda kullanıcı birim tanımlama sayfasına yönlendirilirr
    return redirect('hekim_cizelge:birim_tanimlari')
def birim_yetkileri(request, user_id):
    user = get_object_or_404(User, UserID=user_id)

    if request.method == 'GET':
        yetkili_birimler = UserBirim.objects.filter(user=user).select_related('birim')
        tum_birimler = Birim.objects.all()

        yetkili_birimler_list = [{'BirimID': b.birim.BirimID, 'BirimAdi': b.birim.BirimAdi} for b in yetkili_birimler]
        tum_birimler_list = [{'BirimID': b.BirimID, 'BirimAdi': b.BirimAdi} for b in tum_birimler]

        return JsonResponse({
            'yetkili_birimler': yetkili_birimler_list,
            'tum_birimler': tum_birimler_list
        })

    elif request.method == 'POST':
        birim_id = request.POST.get('birim_id')
        is_add = request.POST.get('is_add')

        birim = get_object_or_404(Birim, BirimID=birim_id)

        if is_add == 'true':
            UserBirim.objects.get_or_create(user=user, birim=birim)
        elif is_add == 'false':
            UserBirim.objects.filter(user=user, birim=birim).delete()

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)

def birim_duzenle_form(request, birim_id):
    birim = get_object_or_404(Birim, BirimID=birim_id)
    hizmetler = Hizmet.objects.all()
    birim_data = {
        'BirimID': birim.BirimID,
        'BirimAdi': birim.BirimAdi,
        'VarsayilanHizmetID': birim.VarsayilanHizmet.HizmetID,
        'DigerHizmetler': [hizmet.HizmetID for hizmet in birim.DigerHizmetler.all()]
    }
    return JsonResponse({'birim': birim_data, 'hizmetler': list(hizmetler.values())})

def birim_duzenle(request, birim_id):
    if request.method == 'POST':
        birim_adi = request.POST.get('birim_adi')
        varsayilan_hizmet_id = request.POST.get('varsayilan_hizmet')
        diger_hizmetler_ids = request.POST.getlist('diger_hizmetler')

        if not varsayilan_hizmet_id:
            messages.error(request, "Varsayılan hizmet seçilmelidir.")
            return redirect('hekim_cizelge:birim_duzenle_form', birim_id)
        
        if not birim_adi:
            messages.error(request, "Birim adı boş bırakılamaz.")
            return redirect('hekim_cizelge:birim_duzenle_form', birim_id)

        try:
            varsayilan_hizmet = Hizmet.objects.get(HizmetID=varsayilan_hizmet_id)
            birim = Birim.objects.get(BirimID=birim_id)
            birim.BirimAdi = birim_adi
            birim.VarsayilanHizmet = varsayilan_hizmet
            birim.save()

            if diger_hizmetler_ids:
                diger_hizmetler = Hizmet.objects.filter(HizmetID__in=diger_hizmetler_ids)
                birim.DigerHizmetler.set(diger_hizmetler)

            messages.success(request, "Birim başarıyla düzenlendi.")
        except Hizmet.DoesNotExist:
            messages.error(request, "Seçilen varsayılan hizmet bulunamadı.")
        except Exception as e:
            messages.error(request, f"Birim düzenlenirken bir hata oluştu: {str(e)}")

        return redirect('hekim_cizelge:birim_tanimlari')

def cizelge(request):
    current_year = int(request.GET.get('year', datetime.now().year))
    current_month = int(request.GET.get('month', datetime.now().month))

    days_in_month = calendar.monthrange(current_year, current_month)[1]
    days = [
        {
            'full_date': f"{current_year}-{current_month:02}-{day:02}",
            'day_num': day,
            'is_weekend': calendar.weekday(current_year, current_month, day) >= 5,
            'is_holiday': ResmiTatil.objects.filter(
                TatilTarihi=datetime(current_year, current_month, day).date()
            ).exists()
        }
        for day in range(1, days_in_month + 1)
    ]

    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim', flat=True)
    birimler = Birim.objects.filter(BirimID__in=user_birimler)

    selected_birim_id = request.GET.get('birim_id')
    if selected_birim_id:
        personeller = Personel.objects.filter(personelbirim__birim_id=selected_birim_id)
        selected_birim = Birim.objects.get(BirimID=selected_birim_id)
        hizmetler = list(selected_birim.DigerHizmetler.all()) + [selected_birim.VarsayilanHizmet]
    else:
        personeller = Personel.objects.none()
        hizmetler = []
        selected_birim = None

    mesailer = Mesai.objects.filter(
        Personel__in=personeller,
        MesaiDate__year=current_year,
        MesaiDate__month=current_month
    ).select_related('Personel', 'Izin').prefetch_related(
        'Hizmetler'
    )

    personel_dict = {p.PersonelID: p.PersonelName for p in personeller}
    for personel in personeller:
        personel.mesai_data = []
        for mesai in mesailer.filter(Personel=personel):
            hizmet_list = []
            for hizmet in mesai.Hizmetler.all():
                hizmet_list.append({
                    'name': hizmet.HizmetName,
                    'type': hizmet.HizmetTipi,
                    'is_varsayilan': hizmet.HizmetID == selected_birim.VarsayilanHizmet.HizmetID
                })
            
            mesai_data = {
                'MesaiDate': mesai.MesaiDate.strftime("%Y-%m-%d"),
                'Hizmetler': hizmet_list,
                'OnayDurumu': mesai.OnayDurumu,
                'MesaiID': mesai.MesaiID
            }

            if mesai.Izin:
                mesai_data['Izin'] = {
                    'id': mesai.Izin.IzinID,
                    'tip': mesai.Izin.IzinTipi,
                    'renk': mesai.Izin.IzinRenk
                }
            
            personel.mesai_data.append(mesai_data)

    is_approval_mode = request.GET.get('mode') == 'approval'
    
    context = {
        'personeller': personeller,
        'days': days,
        'birimler': birimler,
        'current_month': current_month,
        'current_year': current_year,
        'months': [{'value': i, 'label': calendar.month_name[i]} for i in range(1, 13)],
        'years': [year for year in range(current_year - 1, current_year + 2)],
        'hizmetler': hizmetler,
        'izinler': Izin.objects.all(),
        'selected_birim': selected_birim,
        'is_approval_mode': is_approval_mode,
    }

    return render(request, 'hekim_cizelge/cizelge.html', context)

@csrf_exempt
def cizelge_kaydet(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            changes = data.get('changes', {})

            for key, value in changes.items():
                personel_id, date = key.split('_')
                personel = Personel.objects.get(PersonelID=personel_id)
                
                if value.get('type') == 'hizmet':
                    # Hizmet kombinasyonunu kontrol ett
                    hizmetler = Hizmet.objects.filter(HizmetID__in=value.get('hizmetler', []))
                    is_valid, errors = MesaiKontrol.validate_hizmet_combination(hizmetler)
                    
                    if not is_valid:
                        return JsonResponse({
                            'status': 'error',
                            'message': '\n'.join(errors)
                        })

                mesai, created = Mesai.objects.get_or_create(
                    Personel=personel,
                    MesaiDate=date,
                    defaults={
                        'OnayDurumu': 0,
                        'Degisiklik': True,
                        'SilindiMi': False
                    }
                )

                if not created:
                    mesai.OnayDurumu = 0
                    mesai.Degisiklik = True
                    mesai.save()

                if value.get('type') == 'clear':
                    mesai.Hizmetler.clear()
                    mesai.Izin = None
                    mesai.save()
                elif value.get('type') == 'hizmet':
                    mesai.Izin = None
                    mesai.save()
                    # Buradaki hatayı düzelttim - parantezleri kaldırdımm
                    hizmetler = Hizmet.objects.filter(HizmetID__in=value.get('hizmetler', []))
                    mesai.Hizmetler.set(hizmetler)
                elif value.get('type') == 'izin':
                    mesai.Hizmetler.clear()
                    if value.get('izin'):
                        izin = Izin.objects.get(IzinID=value.get('izin'))
                        mesai.Izin = izin
                    else:
                        mesai.Izin = None
                    mesai.save()

            return JsonResponse({'status': 'success'})
        except ValueError as e:
            return JsonResponse({
                'status': 'error', 
                'message': f'Veri format hatası: {str(e)}'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Beklenmeyen bir hata oluştu: {str(e)}'
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@csrf_exempt
def mesai_onay(request, mesai_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mesai = get_object_or_404(Mesai, MesaiID=mesai_id)
            
            mesai.OnayDurumu = data.get('onay_durumu', 0)
            mesai.Onaylayan = request.user
            mesai.OnayTarihi = datetime.now()
            mesai.Degisiklik = False
            mesai.save()
            
            return JsonResponse({
                'status': 'success',
                'mesai_id': mesai_id
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

def mesai_onay_durumu(request, mesai_id):
    mesai = get_object_or_404(Mesai, MesaiID=mesai_id)
    return JsonResponse({
        'mesai_id': mesai_id,
        'onay_durumu': mesai.OnayDurumu,
        'onay_tarihi': mesai.OnayTarihi.isoformat() if mesai.OnayTarihi else None,
        'onaylayan': mesai.Onaylayan.username if mesai.Onaylayan else None
    })

def onay_bekleyen_mesailer(request):
    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim', flat=True)
    
    birimler = Birim.objects.filter(BirimID__in=user_birimler).annotate(
        bekleyen_mesai=models.Count(
            'personelbirim__personel__mesai',
            filter=models.Q(
                personelbirim__personel__mesai__OnayDurumu=0,
                personelbirim__personel__mesai__Degisiklik=True
            )
        )
    ).filter(bekleyen_mesai__gt=0)

    return render(request, 'hekim_cizelge/onay_bekleyen_mesailer.html', {
        'birimler': birimler
    })

@csrf_exempt
def toplu_onay(request, birim_id):
    if request.method == 'POST':
        try:
            mesailer = Mesai.objects.filter(
                Personel__personelbirim__birim_id=birim_id,
                OnayDurumu=0,
                Degisiklik=True
            )
            onay_sayisi = mesailer.count()
            
            mesailer.update(
                OnayDurumu=1,
                Onaylayan=request.user,
                OnayTarihi=datetime.now(),
                Degisiklik=False
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'{onay_sayisi} adet mesai onaylandı.'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

@csrf_exempt
def auto_fill_default(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year'))
            month = int(data.get('month'))
            birim_id = data.get('birim_id')

            birim = Birim.objects.get(BirimID=birim_id)
            personeller = Personel.objects.filter(personelbirim__birim=birim)
            
            days_in_month = calendar.monthrange(year, month)[1]
            dates = [
                datetime(year, month, day) 
                for day in range(1, days_in_month + 1)
                if calendar.weekday(year, month, day) < 5
            ]

            count = 0
            for personel in personeller:
                for date in dates:
                    current_date = date.date()
                    
                    prev_date = current_date - timedelta(days=1)
                    
                    has_previous_nobet = Mesai.objects.filter(
                        Personel=personel,
                        MesaiDate=prev_date,
                        Hizmetler__HizmetTipi='Nöbet',
                        Hizmetler__NobetErtesiIzinli=True,
                        SilindiMi=False
                    ).exists()
                    
                    if has_previous_nobet:
                        continue
                    
                    mesai = Mesai.objects.filter(
                        Personel=personel,
                        MesaiDate=current_date
                    ).first()

                    if not mesai:
                        mesai = Mesai.objects.create(
                            Personel=personel,
                            MesaiDate=current_date,
                            OnayDurumu=0,
                            Degisiklik=True,
                            SilindiMi=False
                        )
                        mesai.Hizmetler.add(birim.VarsayilanHizmet)
                        count += 1
                    elif mesai.SilindiMi or not mesai.Hizmetler.exists():
                        mesai.SilindiMi = False
                        mesai.OnayDurumu = 0
                        mesai.Degisiklik = True
                        mesai.save()
                        mesai.Hizmetler.clear()
                        mesai.Hizmetler.add(birim.VarsayilanHizmet)
                        count += 1

            return JsonResponse({
                'status': 'success',
                'count': count,
                'message': f'{count} adet varsayılan hizmet eklendi.'
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def bildirimler(request):
    """Mesai ve İcap bildirimlerini birleşik görüntüler"""
    current_year = int(request.GET.get('year', datetime.now().year))
    current_month = int(request.GET.get('month', datetime.now().month))

    # Yetkili olunan birimleri al
    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim', flat=True)
    birimler = Birim.objects.filter(BirimID__in=user_birimler)

    # Seçili birimi al
    selected_birim_id = request.GET.get('birim_id')
    selected_birim = None
    personeller = None
    bildirimler = None

    if selected_birim_id:
        selected_birim = get_object_or_404(Birim, BirimID=selected_birim_id)
        personeller = Personel.objects.filter(personelbirim__birim=selected_birim)
        
        donem = datetime(current_year, current_month, 1).date()
        # Hem mesai hem icap bildirimlerini al
        bildirimler = Bildirim.objects.filter(
            PersonelBirim__birim=selected_birim,
            DonemBaslangic=donem,
            SilindiMi=False
        ).select_related('PersonelBirim__personel')

    days = []
    if selected_birim_id:
        days_in_month = calendar.monthrange(current_year, current_month)[1]
        days = [
            {
                'day_num': day,
                'is_weekend': calendar.weekday(current_year, current_month, day) >= 5,
                'is_holiday': ResmiTatil.objects.filter(
                    TatilTarihi=datetime(current_year, current_month, day).date()
                ).exists()
            }
            for day in range(1, days_in_month + 1)
        ]

    context = {
        'current_year': current_year,
        'current_month': current_month,
        'birimler': birimler,
        'selected_birim': selected_birim,
        'personeller': personeller,
        'bildirimler': bildirimler,
        'months': [{'value': i, 'label': calendar.month_name[i]} for i in range(1, 13)],
        'years': range(current_year - 1, current_year + 2),
        'days': days,
    }

    return render(request, 'hekim_cizelge/bildirimler.html', context)

@csrf_exempt 
def bildirim_olustur(request):
    """Birleşik mesai/icap bildirimi oluşturur"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            personel_id = data.get('personel_id')
            birim_id = data.get('birim_id')
            donem = datetime.strptime(data.get('donem'), '%Y-%m').date()

            personel_birim = get_object_or_404(
                PersonelBirim, 
                personel_id=personel_id,
                birim_id=birim_id
            )

            # Birleşik hesaplama yap
            bildirim_verileri = hesapla_bildirim_verileri(personel_id, donem)
            
            # Tek bildirim oluştur/güncelle
            bildirim, _ = Bildirim.objects.update_or_create(
                PersonelBirim=personel_birim,
                DonemBaslangic=donem,
                defaults={
                    'OlusturanKullanici': request.user,
                    'NormalFazlaMesai': bildirim_verileri['mesai']['normal'],
                    'BayramFazlaMesai': bildirim_verileri['mesai']['bayram'],
                    'RiskliNormalFazlaMesai': bildirim_verileri['mesai']['riskli_normal'],
                    'RiskliBayramFazlaMesai': bildirim_verileri['mesai']['riskli_bayram'],
                    'NormalIcap': bildirim_verileri['icap']['normal'],
                    'BayramIcap': bildirim_verileri['icap']['bayram'],
                    'MesaiDetay': bildirim_verileri['mesai']['gunluk_detay'],
                    'IcapDetay': bildirim_verileri['icap']['gunluk_detay']
                }
            )

            response_data = {
                'status': 'success',
                'bildirim_data': {
                    'normal_mesai': float(bildirim.NormalFazlaMesai),
                    'bayram_mesai': float(bildirim.BayramFazlaMesai),
                    'riskli_normal': float(bildirim.RiskliNormalFazlaMesai),
                    'riskli_bayram': float(bildirim.RiskliBayramFazlaMesai),
                    'normal_icap': float(bildirim.NormalIcap),
                    'bayram_icap': float(bildirim.BayramIcap),
                    'toplam_mesai': float(bildirim.ToplamFazlaMesai),
                    'toplam_icap': float(bildirim.ToplamIcap),
                    'MesaiDetay': bildirim.MesaiDetay,
                    'IcapDetay': bildirim.IcapDetay,
                    'onay_durumu': bildirim.OnayDurumu,
                    'bildirim_id': bildirim.BildirimID,  # Eksik ID eklendi
                    'mutemet_kilit': bildirim.MutemetKilit # Eksik kilit durumu eklendi
                }
            }

            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def bildirim_toplu_olustur(request, birim_id):
    """Birim personellerinin tümü için bildirim oluşturur"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year'))
            month = int(data.get('month'))
            
            birim = get_object_or_404(Birim, BirimID=birim_id)
            personeller = Personel.objects.filter(personelbirim__birim=birim)
            count = 0

            donem = datetime(year, month, 1).date()
            for personel in personeller:
                personel_birim = PersonelBirim.objects.get(personel=personel, birim=birim)
                
                # Mevcut bildirim kontrolü
                bildirim = Bildirim.objects.filter(
                    PersonelBirim=personel_birim,
                    DonemBaslangic=donem,
                    SilindiMi=False
                ).first()

                if not bildirim:
                    # Birleşik hesaplama yap
                    bildirim_verileri = hesapla_bildirim_verileri(personel.PersonelID, donem)
                    
                    bildirim = Bildirim(
                        PersonelBirim=personel_birim,
                        DonemBaslangic=donem,
                        OlusturanKullanici=request.user,
                        NormalFazlaMesai=bildirim_verileri['mesai']['normal'],
                        BayramFazlaMesai=bildirim_verileri['mesai']['bayram'],
                        RiskliNormalFazlaMesai=bildirim_verileri['mesai']['riskli_normal'],
                        RiskliBayramFazlaMesai=bildirim_verileri['mesai']['riskli_bayram'],
                        NormalIcap=bildirim_verileri['icap']['normal'],
                        BayramIcap=bildirim_verileri['icap']['bayram'],
                        MesaiDetay=bildirim_verileri['mesai']['gunluk_detay'],
                        IcapDetay=bildirim_verileri['icap']['gunluk_detay']
                    )
                    bildirim.save()
                    count += 1

            return JsonResponse({
                'status': 'success',
                'count': count
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def bildirim_form(request, birim_id):
    """Bildirim formunu PDF olarak oluşturur"""
    try:
        # URL'den yıl, ay parametrelerini al
        year = request.GET.get('year')
        month = request.GET.get('month')
        if not year or not month:
            raise ValueError("Yıl ve ay parametreleri gerekli")
            
        donem = datetime(int(year), int(month), 1).date()
        
        # Birim ve dönem bilgisiyle bildirimleri sorgula
        bildirimler = Bildirim.objects.filter(
            PersonelBirim__birim_id=birim_id,
            DonemBaslangic=donem,
            SilindiMi=False
        ).select_related('PersonelBirim__personel', 'PersonelBirim__birim')
        
        if not bildirimler.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'Bu dönem için bildirim bulunamadı.'
            })
        
        # Ayın günlerini oluştur
        days_in_month = calendar.monthrange(int(year), int(month))[1]
        days = []
        for day in range(1, days_in_month + 1):
            current_date = datetime(int(year), int(month), day).date()
            days.append({
                'day_num': day,
                'full_date': current_date.strftime('%Y-%m-%d'),
                'is_weekend': calendar.weekday(int(year), int(month), day) >= 5,
                'is_holiday': ResmiTatil.objects.filter(TatilTarihi=current_date).exists()
            })
        
        # Birim bilgisini al
        birim = Birim.objects.get(BirimID=birim_id)
        
        # Bildirimleri hazırla
        bildirim_data = {
            'donem': donem,
            'birim': birim.BirimAdi,
            'days': days,
            'personeller': []
        }
        
        for bildirim in bildirimler:
            personel_data = {
                'PersonelID': bildirim.PersonelBirim.personel.PersonelID,
                'PersonelName': bildirim.PersonelBirim.personel.PersonelName,
                'normal_mesai': float(bildirim.NormalFazlaMesai),
                'bayram_mesai': float(bildirim.BayramFazlaMesai),
                'riskli_normal': float(bildirim.RiskliNormalFazlaMesai),
                'riskli_bayram': float(bildirim.RiskliBayramFazlaMesai),
                'normal_icap': float(bildirim.NormalIcap),
                'bayram_icap': float(bildirim.BayramIcap),
                'toplam_mesai': float(bildirim.ToplamFazlaMesai),
                'toplam_icap': float(bildirim.ToplamIcap),
                'onay_durumu': bildirim.OnayDurumu,
                'mesai_detay': bildirim.MesaiDetay,
                'icap_detay': bildirim.IcapDetay
            }
            bildirim_data['personeller'].append(personel_data)
        
        # PDF oluştur
        template = get_template('hekim_cizelge/bildirim_form.html')
        html = template.render({'bildirimler': [bildirim_data]})
        
        # PDF oluşturma seçenekleri
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '1.5cm',
            'margin-right': '1.5cm',
            'margin-bottom': '1.1cm',
            'margin-left': '1.5cm',
            'encoding': 'UTF-8',
            'no-outline': None,
            'enable-local-file-access': True,
            'enable-external-links': True,
            'quiet': ''
        }
        
        # PDF oluştur
        pdf = pdfkit.from_string(html, False, options=options, configuration=config)
        
        # HTTP response oluştur
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="bildirimler_{year}_{month}.pdf"'
        return response
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@csrf_exempt
def bildirim_guncelle(request, bildirim_id):
    """Bildirim detaylarını günceller"""
    if request.method == 'POST':
        try:
            bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
            
            # Onaylı bildirimleri güncelleme
            if bildirim.OnayDurumu == 1:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Onaylanmış bildirimler güncellenemez'
                })
            
            data = json.loads(request.body)
            donem = datetime.strptime(data.get('donem'), '%Y-%m').date()
            
            # Bildirim tipine göre hesaplama yap
            if bildirim.BildirimTipi == 'MESAI':
                mesai_detay = hesapla_fazla_mesai(bildirim.Personel.PersonelID, donem)
                bildirim.NormalFazlaMesai = mesai_detay['normal']  
                bildirim.BayramFazlaMesai = mesai_detay['bayram']
                bildirim.RiskliNormalFazlaMesai = mesai_detay['riskli_normal']
                bildirim.RiskliBayramFazlaMesai = mesai_detay['riskli_bayram']
                bildirim.MesaiDetay = mesai_detay['gunluk_detay']
            else:
                icap_detay = hesapla_icap_suresi(bildirim.Personel.PersonelID, donem)
                bildirim.NormalIcap = icap_detay['normal']
                bildirim.BayramIcap = icap_detay['bayram']
                bildirim.IcapDetay = icap_detay['gunluk_detay']
            
            bildirim.save()
            
            return JsonResponse({
                'status': 'success',
                'bildirim_id': bildirim.BildirimID
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def bildirim_sil(request, bildirim_id):
    """Bildirimi soft-delete yapar"""
    if request.method == 'POST':
        try:
            bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
            
            # Onaylı bildirimleri silme
            if bildirim.OnayDurumu == 1:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Onaylanmış bildirimler silinemez'
                })
            
            bildirim.SilindiMi = True
            bildirim.save()
            
            return JsonResponse({'status': 'success'})
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt 
def bildirim_toplu_onay(request, birim_id):
    """Birime ait bildirimleri toplu onaylar/onaylarını kaldırır"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year'))
            month = int(data.get('month'))
            onay_durumu = int(data.get('onay_durumu'))
            bildirim_id = data.get('bildirim_id')  # Tek bildirim için ID
            
            donem = datetime(year, month, 1).date()
            
            if bildirim_id:
                # Tekil onay işlemi
                bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
                if bildirim.MutemetKilit:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Bu kayıt kilitli olduğu için işlem yapılamaz.'
                    })
                bildirim.OnayDurumu = onay_durumu
                bildirim.OnaylayanKullanici = request.user if onay_durumu == 1 else None
                bildirim.OnayTarihi = datetime.now() if onay_durumu == 1 else None
                bildirim.save()
                return JsonResponse({
                    'status': 'success',
                    'message': 'Bildirim durumu güncellendi.'
                })
            else:
                # Toplu onay işlemi
                kilitli_kayit_var = Bildirim.objects.filter(
                    PersonelBirim__birim_id=birim_id,
                    DonemBaslangic=donem,
                    SilindiMi=False,
                    MutemetKilit=True
                ).exists()
                
                if kilitli_kayit_var:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Bazı kayıtlar kilitli olduğu için işlem yapılamaz.'
                    })
                
                # Filtre kriterlerini güncelle
                filtre = {
                    'PersonelBirim__birim_id': birim_id,
                    'DonemBaslangic': donem,
                    'SilindiMi': False,
                    'MutemetKilit': False  # Sadece kilitli olmayan kayıtlar
                }
                
                if onay_durumu == 1:
                    filtre['OnayDurumu'] = 0
                else:
                    filtre['OnayDurumu'] = 1
                
                bildirimler = Bildirim.objects.filter(**filtre)
                islem_sayisi = bildirimler.count()
                
                if islem_sayisi == 0:
                    mesaj = 'Onaylanacak bildirim bulunamadı.' if onay_durumu == 1 else 'Onayı kaldırılacak bildirim bulunamadı.'
                    return JsonResponse({
                        'status': 'error',
                        'message': mesaj
                    })
                
                # Bildirimleri güncelle
                bildirimler.update(
                    OnayDurumu=onay_durumu,
                    OnaylayanKullanici=request.user if onay_durumu == 1 else None,
                    OnayTarihi=datetime.now() if onay_durumu == 1 else None
                )
                
                mesaj = f'{islem_sayisi} adet bildirim onaylandı.' if onay_durumu == 1 else f'{islem_sayisi} adet bildirimin onayı kaldırıldı.'
                
                return JsonResponse({
                    'status': 'success',
                    'count': islem_sayisi,
                    'message': mesaj
                })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def bildirim_tekil_onay(request, bildirim_id):
    """Tekil bildirimi onaylar/onayını kaldırır"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            onay_durumu = int(data.get('onay_durumu'))
            
            bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
            
            # Kilit kontrolü
            if bildirim.MutemetKilit:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Bu kayıt kilitli olduğu için işlem yapılamaz.'
                })

            bildirim.OnayDurumu = onay_durumu
            bildirim.OnaylayanKullanici = request.user if onay_durumu == 1 else None
            bildirim.OnayTarihi = datetime.now() if onay_durumu == 1 else None
            bildirim.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Bildirim durumu güncellendi.'
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def resmi_tatiller(request):
    tatiller = ResmiTatil.objects.all().order_by('TatilTarihi')
    return render(request, 'hekim_cizelge/resmi_tatiller.html', {
        'tatiller': tatiller
    })

@csrf_exempt
def tatil_ekle(request):
    if request.method == 'POST':
        try:
            tarih = request.POST.get('tarih')
            aciklama = request.POST.get('aciklama')
            tip = request.POST.get('tip')
            bayram_adi = request.POST.get('bayram_adi', '').strip() or None
            bayram_mi = bool(request.POST.get('bayram_mi'))
            arefe_mi = bool(request.POST.get('arefe_mi'))

            ResmiTatil.objects.create(
                TatilTarihi=tarih,
                Aciklama=aciklama,
                TatilTipi=tip,
                BayramAdi=bayram_adi,
                BayramMi=bayram_mi,
                ArefeMi=arefe_mi
            )
            messages.success(request, "Resmi tatil başarıyla eklendi.")
        except Exception as e:
            messages.error(request, f"Hata oluştu: {str(e)}")
    return redirect('hekim_cizelge:resmi_tatiller')

@csrf_exempt
def tatil_duzenle(request):
    if request.method == 'POST':
        try:
            tatil_id = request.POST.get('tatil_id')
            tarih = request.POST.get('tarih')
            aciklama = request.POST.get('aciklama')
            tip = request.POST.get('tip')
            bayram_adi = request.POST.get('bayram_adi', '').strip() or None
            bayram_mi = request.POST.get('bayram_mi') in ['true', 'True', 'on', '1']
            arefe_mi = request.POST.get('arefe_mi') in ['true', 'True', 'on', '1']

            tatil = ResmiTatil.objects.get(TatilID=tatil_id)
            tatil.TatilTarihi = tarih
            tatil.Aciklama = aciklama
            tatil.TatilTipi = tip
            tatil.BayramAdi = bayram_adi
            tatil.BayramMi = bayram_mi
            tatil.ArefeMi = arefe_mi
            tatil.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@csrf_exempt
def tatil_sil(request, tatil_id):
    """Resmi tatil kaydını siler"""
    if request.method == 'POST':
        try:
            tatil = get_object_or_404(ResmiTatil, TatilID=tatil_id)
            tatil.delete()
            messages.success(request, "Resmi tatil başarıyla silindi.")
        except Exception as e:
            messages.error(request, f"Silme işlemi başarısız: {str(e)}")
            
        return redirect('hekim_cizelge:resmi_tatiller')
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@csrf_exempt
def bildirim_listele(request, yil, ay, birim_id):
    """Birime ait bildirimleri listeler"""
    try:
        donem = datetime(yil, ay, 1).date()
        bildirimler = Bildirim.objects.filter(
            PersonelBirim__birim_id=birim_id,
            DonemBaslangic=donem,
            SilindiMi=False
        ).select_related(
            'PersonelBirim__personel'
        )
        
        data = []
        for bildirim in bildirimler:
            item = {
                'personel_id': bildirim.PersonelBirim.personel.PersonelID,
                'bildirim_id': bildirim.BildirimID,  # BildirimID eklendi
                'normal_mesai': float(bildirim.NormalFazlaMesai),
                'bayram_mesai': float(bildirim.BayramFazlaMesai),
                'riskli_normal': float(bildirim.RiskliNormalFazlaMesai),
                'riskli_bayram': float(bildirim.RiskliBayramFazlaMesai),
                'normal_icap': float(bildirim.NormalIcap),
                'bayram_icap': float(bildirim.BayramIcap),
                'toplam_mesai': float(bildirim.ToplamFazlaMesai),
                'toplam_icap': float(bildirim.ToplamIcap),
                'onay_durumu': bildirim.OnayDurumu,
                'mutemet_kilit': bildirim.MutemetKilit,  # Kilit durumu eklendi
                'MesaiDetay': bildirim.MesaiDetay,
                'IcapDetay': bildirim.IcapDetay
            }
            data.append(item)
            
        return JsonResponse({'status': 'success', 'data': data})
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@csrf_exempt
def hizmet_raporu(request):
    """Hizmet raporu sayfasını gösterir"""
    if request.method == 'GET':
        hizmetler = Hizmet.objects.all()
        return render(request, 'hekim_cizelge/hizmet_raporu.html', {
            'hizmetler': hizmetler
        })
    
    elif request.method == 'POST':
        try:
            baslangic_tarihi = request.POST.get('baslangic_tarihi')
            bitis_tarihi = request.POST.get('bitis_tarihi')
            hizmet_ids = request.POST.getlist('hizmetler[]')

            if not all([baslangic_tarihi, bitis_tarihi, hizmet_ids]):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Tüm alanları doldurunuz.'
                })

            # Tarihleri datetime objesine çevir
            baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()

            # Mesaileri sorgula
            mesailer = Mesai.objects.filter(
                MesaiDate__range=[baslangic, bitis],
                Hizmetler__HizmetID__in=hizmet_ids,
                SilindiMi=False
            ).select_related('Personel').prefetch_related('Hizmetler')

            # Sonuçları hazırla
            rapor_data = []
            gunluk_ozet = {}
            
            for mesai in mesailer:
                tarih = mesai.MesaiDate.strftime('%d.%m.%Y')
                hizmet_isimleri = ', '.join([h.HizmetName for h in mesai.Hizmetler.all()])
                durum = 'Onaylı' if mesai.OnayDurumu == 1 else 'Beklemede'
                
                # Günlük özet bilgilerini güncelle
                if tarih not in gunluk_ozet:
                    gunluk_ozet[tarih] = {
                        'tarih': tarih,
                        'hekim_sayisi': 0,
                        'onayli_sayisi': 0,
                        'bekleyen_sayisi': 0,
                        'detaylar': []
                    }
                
                gunluk_ozet[tarih]['hekim_sayisi'] += 1
                if mesai.OnayDurumu == 1:
                    gunluk_ozet[tarih]['onayli_sayisi'] += 1
                else:
                    gunluk_ozet[tarih]['bekleyen_sayisi'] += 1
                
                # Detay bilgilerini ekle
                gunluk_ozet[tarih]['detaylar'].append({
                    'personel': mesai.Personel.PersonelName,
                    'hizmetler': hizmet_isimleri,
                    'durum': durum
                })

            # Günlük özetleri sıralı şekilde rapor_data'ya ekle
            for tarih in sorted(gunluk_ozet.keys()):
                ozet = gunluk_ozet[tarih]
                rapor_data.append({
                    'tarih': tarih,
                    'ozet': {
                        'hekim_sayisi': ozet['hekim_sayisi'],
                        'onayli_sayisi': ozet['onayli_sayisi'],
                        'bekleyen_sayisi': ozet['bekleyen_sayisi']
                    },
                    'detaylar': ozet['detaylar']
                })

            return JsonResponse({
                'status': 'success',
                'data': rapor_data
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

def hizmet_raporu_pdf(request):
    """Hizmet raporunu PDF olarak oluşturur"""
    try:
        baslangic_tarihi = request.GET.get('baslangic')
        bitis_tarihi = request.GET.get('bitis')
        hizmet_ids = request.GET.get('hizmetler').split(',')

        if not all([baslangic_tarihi, bitis_tarihi, hizmet_ids]):
            return JsonResponse({
                'status': 'error',
                'message': 'Eksik parametre.'
            })

        # Tarihleri datetime objesine çevir
        baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
        bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()

        # Mesaileri sorgula
        mesailer = Mesai.objects.filter(
            MesaiDate__range=[baslangic, bitis],
            Hizmetler__HizmetID__in=hizmet_ids,
            SilindiMi=False
        ).select_related('Personel').prefetch_related('Hizmetler')

        # Rapor verilerini hazırla
        rapor_data = {
            'baslangic_tarihi': baslangic.strftime('%d.%m.%Y'),
            'bitis_tarihi': bitis.strftime('%d.%m.%Y'),
            'hizmetler': Hizmet.objects.filter(HizmetID__in=hizmet_ids).values_list('HizmetName', flat=True),
            'mesailer': []
        }

        for mesai in mesailer:
            hizmet_isimleri = ', '.join([h.HizmetName for h in mesai.Hizmetler.all()])
            durum = 'Onaylı' if mesai.OnayDurumu == 1 else 'Beklemede'
            
            rapor_data['mesailer'].append({
                'tarih': mesai.MesaiDate.strftime('%d.%m.%Y'),
                'personel': mesai.Personel.PersonelName,
                'hizmetler': hizmet_isimleri,
                'durum': durum
            })

        # Tarihe göre sırala
        rapor_data['mesailer'].sort(key=lambda x: datetime.strptime(x['tarih'], '%d.%m.%Y'))

        # PDF oluştur
        template = get_template('hekim_cizelge/hizmet_raporu_pdf.html')
        html = template.render({'rapor': rapor_data})

        # PDF oluşturma seçenekleri
        options = {
            'page-size': 'A4',
            'orientation': 'Portrait',
            'margin-top': '1.5cm',
            'margin-right': '1.5cm',
            'margin-bottom': '1.1cm',
            'margin-left': '1.5cm',
            'encoding': 'UTF-8',
            'no-outline': None,
            'enable-local-file-access': True,
            'enable-external-links': True,
            'quiet': ''
        }

        # PDF oluştur
        pdf = pdfkit.from_string(html, False, options=options, configuration=config)

        # HTTP response oluştur
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="hizmet_raporu_{baslangic_tarihi}_{bitis_tarihi}.pdf"'
        return response

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

def birim_dashboard(request):
    """Birim bazlı dashboard sayfası"""
    if request.method == 'GET':
        # Kullanıcının yetkili olduğu birimleri al
        user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim', flat=True)
        birimler = Birim.objects.filter(BirimID__in=user_birimler)
        
        # Seçili birimi al
        selected_birim_id = request.GET.get('birim_id')
        dashboard_data = None
        
        if selected_birim_id:
            selected_birim = get_object_or_404(Birim, BirimID=selected_birim_id)
            
            # Birime ait personelleri al
            personeller = Personel.objects.filter(personelbirim__birim=selected_birim)
            
            # Her personel için istatistikleri hesapla
            personel_stats = []
            nobet_labels = []
            nobet_values = []
            icap_labels = []
            icap_values = []
            
            for personel in personeller:
                # Personelin ilk mesai tarihini bul
                ilk_mesai = Mesai.objects.filter(
                    Personel=personel,
                    SilindiMi=False
                ).order_by('MesaiDate').first()
                
                ilk_tarih = ilk_mesai.MesaiDate if ilk_mesai else None
                
                # Nöbet ve icap sayılarını hesapla
                nobet_sayisi = Mesai.objects.filter(
                    Personel=personel,
                    Hizmetler__HizmetTipi='Nöbet',
                    SilindiMi=False
                ).count()
                
                icap_sayisi = Mesai.objects.filter(
                    Personel=personel,
                    Hizmetler__HizmetTipi='İcap',
                    SilindiMi=False
                ).count()
                
                # Nöbet katsayısını hesapla
                nobet_katsayisi = None
                if ilk_tarih and nobet_sayisi > 0:
                    calisma_suresi = (datetime.now().date() - ilk_tarih).days
                    nobet_katsayisi = round(calisma_suresi / nobet_sayisi, 2)
                
                personel_stats.append({
                    'personel': personel,
                    'ilk_tarih': ilk_tarih,
                    'nobet_sayisi': nobet_sayisi,
                    'icap_sayisi': icap_sayisi,
                    'nobet_katsayisi': nobet_katsayisi
                })
                
                # Grafik için personel bazlı verileri ekle
                nobet_labels.append(personel.PersonelName)
                nobet_values.append(nobet_sayisi)
                icap_labels.append(personel.PersonelName)
                icap_values.append(icap_sayisi)
            
            dashboard_data = {
                'selected_birim': selected_birim,
                'personel_stats': personel_stats,
                'nobet_labels': nobet_labels,
                'nobet_values': nobet_values,
                'icap_labels': icap_labels,
                'icap_values': icap_values
            }
        
        return render(request, 'hekim_cizelge/birim_dashboard.html', {
            'birimler': birimler,
            'dashboard_data': dashboard_data
        })

import os
from io import BytesIO
from django.conf import settings
import openpyxl

def mutemetlik_islemleri(request):
    """Mutemetlik işlemleri sayfası"""
    # Mevcut yıl ve ay bilgisini al
    current_year = int(request.GET.get('year', datetime.now().year))
    current_month = int(request.GET.get('month', datetime.now().month))
    tc_kimlik = request.GET.get('tc_kimlik', '')

    # Kullanıcının yetkili olduğu birimleri al
    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim', flat=True)
    birimler = Birim.objects.filter(BirimID__in=user_birimler)

    # Sorgu parametreleri hazırla
    filtre = {
        'DonemBaslangic': datetime(current_year, current_month, 1).date(),
        'OnayDurumu': 1,  # Sadece onaylı bildirimleri getir
        'SilindiMi': False,
        'PersonelBirim__birim__in': birimler  # Sadece yetkili olunan birimlerin bildirimlerini getir
    }

    if tc_kimlik:
        filtre['PersonelBirim__personel__TCKimlikNo'] = tc_kimlik

    # Bildirimleri sorgula
    bildirimler = Bildirim.objects.filter(**filtre).select_related(
        'PersonelBirim__personel',
        'PersonelBirim__birim'
    )

    # İstatistik hesapla
    onayli_sayisi = bildirimler.count()
    onaysiz_sayisi = Bildirim.objects.filter(
        DonemBaslangic=filtre['DonemBaslangic'],
        OnayDurumu=0,
        SilindiMi=False,
        PersonelBirim__birim__in=birimler
    ).count()

    context = {
        'current_year': current_year,
        'current_month': current_month,
        'years': range(current_year - 1, current_year + 2),
        'months': [{'value': i, 'label': calendar.month_name[i]} for i in range(1, 13)],
        'bildirimler': bildirimler,
        'onayli_sayisi': onayli_sayisi,
        'onaysiz_sayisi': onaysiz_sayisi,
        'birimler': birimler  # Add birimler to context
    }

    return render(request, 'hekim_cizelge/mutemetlik_islemleri.html', context)

@csrf_exempt
def bildirim_kilit(request, bildirim_id):
    """Locks a specific record."""
    try:
        bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
        if bildirim.MutemetKilit:
            return JsonResponse({'status': 'error', 'message': 'Bildirim zaten kilitli.'})
        bildirim.MutemetKilit = True
        bildirim.MutemetKilitUser = request.user
        bildirim.MutemetKilitTime = now()
        bildirim.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def bildirim_kilit_ac(request, bildirim_id):
    """Unlocks a specific record."""
    try:
        bildirim = get_object_or_404(Bildirim, BildirimID=bildirim_id)
        if not bildirim.MutemetKilit:
            return JsonResponse({'status': 'error', 'message': 'Bildirim zaten kilitli değil.'})
        bildirim.MutemetKilit = False
        bildirim.MutemetKilitUser = None
        bildirim.MutemetKilitTime = None
        bildirim.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def toplu_kilit(request):
    """Locks or unlocks all filtered records."""
    try:
        data = json.loads(request.body)
        year = int(data.get('year'))
        month = int(data.get('month'))
        tc_kimlik = data.get('tc_kimlik', '')
        action = data.get('action')  # 'lock' or 'unlock'

        filtre = {
            'DonemBaslangic': datetime(year, month, 1).date(),
            'SilindiMi': False
        }
        if tc_kimlik:
            filtre['PersonelBirim__personel__TCKimlikNo'] = tc_kimlik

        bildirimler = Bildirim.objects.filter(**filtre)
        if action == 'lock':
            bildirimler.update(
                MutemetKilit=True,
                MutemetKilitUser=request.user,
                MutemetKilitTime=now()
            )
        elif action == 'unlock':
            bildirimler.update(
                MutemetKilit=False,
                MutemetKilitUser=None,
                MutemetKilitTime=None
            )
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def bildirim_excel(request):
    """Exports records to Excel and locks them."""
    try:
        year = int(request.GET.get('year'))
        month = int(request.GET.get('month'))
        tc_kimlik = request.GET.get('tc_kimlik', '')

        # Sorgu parametreleri hazırla
        filtre = {
            'DonemBaslangic': datetime(year, month, 1).date(),
            'OnayDurumu': 1,
            'SilindiMi': False
        }

        if tc_kimlik:
            filtre['PersonelBirim__personel__TCKimlikNo'] = tc_kimlik

        bildirimler = Bildirim.objects.filter(**filtre).select_related(
            'PersonelBirim__personel',
            'PersonelBirim__birim'
        )

        if not bildirimler.exists():
            messages.warning(request, "İndirilebilecek onaylı bildirim bulunamadı.")
            return redirect('hekim_cizelge:mutemetlik_islemleri')

        # Lock all records before exporting
        bildirimler.update(
            MutemetKilit=True,
            MutemetKilitUser=request.user,
            MutemetKilitTime=now()
        )

        # Şablon dosyasını aç
        template_path = os.path.join(settings.STATIC_ROOT, 'excels', 'BildirimSablon.xlsx')
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Verileri yaz (2. satırdan başla)
        row = 2
        for bildirim in bildirimler:
            personel = bildirim.PersonelBirim.personel
            birim = bildirim.PersonelBirim.birim
            
            ws.cell(row=row, column=1, value=personel.PersonelID)
            ws.cell(row=row, column=2, value=personel.FirstName)
            ws.cell(row=row, column=3, value=personel.LastName)
            ws.cell(row=row, column=4, value=1)
            ws.cell(row=row, column=5, value=float(bildirim.NormalFazlaMesai or 0))
            ws.cell(row=row, column=7, value=float(bildirim.BayramFazlaMesai or 0))
            ws.cell(row=row, column=8, value=float(bildirim.RiskliNormalFazlaMesai or 0))
            ws.cell(row=row, column=9, value=float(bildirim.RiskliBayramFazlaMesai or 0))
            ws.cell(row=row, column=10, value=float(bildirim.NormalIcap or 0))
            ws.cell(row=row, column=11, value=float(bildirim.BayramIcap or 0))
            ws.cell(row=row, column=12, value=birim.BirimAdi)
            row += 1
        
        # Excel dosyasını kaydet
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # HTTP response oluştur
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="bildirimler_{year}_{month}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f"Excel oluşturulurken hata: {str(e)}")
        return redirect('hekim_cizelge:mutemetlik_islemleri')
