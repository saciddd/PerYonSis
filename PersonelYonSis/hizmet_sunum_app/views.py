from django.shortcuts import redirect, render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET
from datetime import datetime, date
from .models import Birim, HizmetSunumAlani, UserBirim, HizmetSunumCalismasi, Personel, Kurum, Idare
import json
import calendar
from django.utils.timezone import now as django_now
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
import pdfkit
import openpyxl
import os
from django.conf import settings
from io import BytesIO
from django.db.models import Count, Prefetch
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.db.models.functions import TruncMonth
from django.forms import Form
from django import forms
from django.contrib import messages
from django.contrib.auth import get_user_model
from PersonelYonSis.views import get_user_permissions

from .models import UserBirim  # hizmet_sunum_app_userbirim modelinizin importu

User = get_user_model()

@login_required
def bildirim(request):
    # Kullanıcının yetkili olduğu birimleri getir 
    user_birimler_qs = UserBirim.objects.filter(user=request.user).select_related('birim', 'birim__HSAKodu')
    birimler = [ub.birim for ub in user_birimler_qs]
    kesinlestirme_yetkisi = request.user.has_permission('HSA Bildirim Kesinleştirme')
    tum_birimler_yetkisi = request.user.has_permission('HSA Tüm Birimleri Görebilir')
    kurumlar = Kurum.objects.all()
    idareler = Idare.objects.all()
    tum_birimler = Birim.objects.all().order_by('BirimAdi') if tum_birimler_yetkisi else birimler
    # Tüm HSA listesini modal için alalım
    hsa_listesi = HizmetSunumAlani.objects.all().order_by('AlanAdi')

    # Birimleri JSON'a çevir
    birimler_json = json.dumps([
        {'BirimId': b.BirimId, 'BirimAdi': b.BirimAdi} for b in birimler
    ])

    # Başlangıçta context'e seçili birim veya bildirim eklemeye gerek yok,
    # bu bilgiler AJAX ile yüklenecek.
    context = {
        'birimler': tum_birimler,
        'kurumlar': kurumlar,
        'idareler': idareler,
        'birimler_json': birimler_json,
        'hsa_listesi': hsa_listesi,
        'kesinlestirme_yetkisi': kesinlestirme_yetkisi,
        'tum_birimler_yetkisi': tum_birimler_yetkisi,
        # 'donemler' frontend JS ile oluşturuluyor, Django context'ine gerek yok.
    }
    return render(request, 'hizmet_sunum_app/bildirim.html', context)

@login_required
@require_GET
def birim_detay(request, birim_id):
    # Yetki kontrolü
    if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
        if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
            return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
            
    birim = get_object_or_404(Birim.objects.select_related('HSAKodu'), BirimId=birim_id)
    
    # Frontend'in beklediği formatta döndür
    data = {
        'BirimId': birim.BirimId,
        'BirimAdi': birim.BirimAdi,
        'KurumAdi': birim.KurumAdi,
        'IdareAdi': birim.IdareAdi,
        'HSAKodu': birim.HSAKodu.AlanKodu if birim.HSAKodu else None,
        'HSAAdi': birim.HSAKodu.AlanAdi if birim.HSAKodu else None
    }
    return JsonResponse({'status': 'success', 'data': data})

@login_required
@require_POST
def birim_ekle(request):
    try:
        birim_adi = request.POST.get('birimAdi')
        kurum_adi = request.POST.get('kurumAdi')
        idare_adi = request.POST.get("idareAdi")
        hsa_kodu_str = request.POST.get('hsaKodu')

        if not all([birim_adi, kurum_adi, hsa_kodu_str]):
             return JsonResponse({'status': 'error', 'message': 'Tüm alanlar zorunludur.'}, status=400)

        hsa_nesnesi = get_object_or_404(HizmetSunumAlani, AlanKodu=hsa_kodu_str)
        
        # Aynı isimde başka birim var mı kontrol et (opsiyonel)
        # if Birim.objects.filter(BirimAdi=birim_adi).exists():
        #    return JsonResponse({'status': 'error', 'message': 'Bu isimde bir birim zaten mevcut.'}, status=400)

        with transaction.atomic(): # Atomik işlem başlat
            birim = Birim.objects.create(
                BirimAdi=birim_adi,
                KurumAdi=kurum_adi,
                HSAKodu=hsa_nesnesi,
                IdareAdi=idare_adi
                # AlanKodu alanı modelde otomatik atanıyor olabilir, kontrol edin.
                # Eğer modelde AlanKodu'nu ayrıca set ediyorsanız, burada da yapın.
            )
            
            # Kullanıcıyı yeni birime otomatik olarak yetkilendir
            UserBirim.objects.create(user=request.user, birim=birim)
            
        return JsonResponse({'status': 'success', 'message': 'Birim başarıyla eklendi.'})
    except HizmetSunumAlani.DoesNotExist:
         return JsonResponse({'status': 'error', 'message': 'Geçersiz Hizmet Sunum Alanı kodu.'}, status=400)
    except Exception as e:
        # Hata loglama eklenebilir
        return JsonResponse({'status': 'error', 'message': f'Birim eklenirken hata oluştu: {str(e)}'}, status=500)
    
@login_required
@require_POST
def birim_guncelle(request, birim_id):
    # Yetki kontrolü
    if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
        if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
            return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
        
    birim = get_object_or_404(Birim, BirimId=birim_id)
    
    try:
        birim_adi = request.POST.get('birimAdi')
        kurum_adi = request.POST.get('kurumAdi')
        hsa_kodu_str = request.POST.get('hsaKodu')
        idare_adi = request.POST.get("idareAdi")

        if not all([birim_adi, kurum_adi, hsa_kodu_str]):
             return JsonResponse({'status': 'error', 'message': 'Tüm alanlar zorunludur.'}, status=400)

        hsa_nesnesi = get_object_or_404(HizmetSunumAlani, AlanKodu=hsa_kodu_str)

        # Aynı isimde başka birim var mı kontrol et (kendisi hariç)
        # if Birim.objects.filter(BirimAdi=birim_adi).exclude(BirimId=birim_id).exists():
        #    return JsonResponse({'status': 'error', 'message': 'Bu isimde başka bir birim zaten mevcut.'}, status=400)

        birim.BirimAdi = birim_adi
        birim.KurumAdi = kurum_adi
        birim.HSAKodu = hsa_nesnesi
        birim.IdareAdi = idare_adi
        # AlanKodu alanı modelde otomatik atanıyor olabilir, kontrol edin.
        birim.save()
        messages.success(request, "Birim bilgileri başarıyla güncellendi.")
        return JsonResponse({'status': 'success', 'message': 'Birim başarıyla güncellendi.'})
    except HizmetSunumAlani.DoesNotExist:
         return JsonResponse({'status': 'error', 'message': 'Geçersiz Hizmet Sunum Alanı kodu.'}, status=400)
    except Exception as e:
        # Hata loglama eklenebilir
        return JsonResponse({'status': 'error', 'message': f'Birim güncellenirken hata oluştu: {str(e)}'}, status=500)

@login_required
@require_POST
def birim_sil(request, birim_id):
    # Yetki kontrolü
    if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
        if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
            return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
            
    birim = get_object_or_404(Birim, BirimId=birim_id)
    
    try:
        with transaction.atomic():
            # İlişkili HizmetSunumCalismasi kayıtları varsa ne yapılacağına karar verin.
            # Modelde on_delete=models.CASCADE varsa, birim silinince çalışmalar da silinir.
            # Eğer silinmemesi gerekiyorsa veya başka bir işlem yapılacaksa burada ele alınmalı.
            # Örneğin:
            # if HizmetSunumCalismasi.objects.filter(CalisilanBirimId=birim).exists():
            #     return JsonResponse({'status': 'error', 'message': 'Bu birime ait çalışma kayıtları olduğu için silinemez.'}, status=400)
            
            birim_adi = birim.BirimAdi # Silmeden önce adı alalım
            birim.delete()
        return JsonResponse({'status': 'success', 'message': f'{birim_adi} birimi başarıyla silindi.'})
    except Exception as e:
         # Hata loglama eklenebilir
        return JsonResponse({'status': 'error', 'message': f'Birim silinirken hata oluştu: {str(e)}'}, status=500)

@login_required
@require_GET
def onceki_donem_personel(request, donem, birim_id):
    """Bir önceki döneme ait personelleri getir (Mevcut fonksiyon doğru görünüyor)"""
    # Yetki kontrolü
    if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
        if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
            return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
            
    try:
        donem_date = datetime.strptime(donem, '%Y-%m')
        if donem_date.month == 1:
            onceki_ay = datetime(donem_date.year - 1, 12, 1)
        else:
            onceki_ay = datetime(donem_date.year, donem_date.month - 1, 1)
            
        # Önceki dönem personellerini getir
        # Model alan adlarının doğru olduğunu varsayıyoruz (PersonelId__PersonelId vs.)
        personeller_qs = HizmetSunumCalismasi.objects.filter(
            CalisilanBirimId_id=birim_id,
            Donem=onceki_ay.date()
        ).select_related('PersonelId').values(
            'PersonelId__PersonelId', # Personel modelinin PK'sı
            'PersonelId__TCKimlikNo', # TC Kimlik No eklendi
            'PersonelId__PersonelAdi',
            'PersonelId__PersonelSoyadi'
        ).distinct()
        
        data = [{
            'personel_id': p['PersonelId__PersonelId'],
            'tc_kimlik': p['PersonelId__TCKimlikNo'], # TC Kimlik No eklendi
            'adi': p['PersonelId__PersonelAdi'],
            'soyadi': p['PersonelId__PersonelSoyadi']
        } for p in personeller_qs]
        
        # Başarı durumunda status eklemek iyi pratik olabilir
        return JsonResponse(data, safe=False) # Direkt listeyi döndür
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
@require_POST
def personel_kaydet(request):
    """Personel kayıtlarını oluşturur. Aynı dönem/birim için mükerrer personel eklemeye izin verir."""
    try:
        data = json.loads(request.body)
        donem_str = data.get('donem')
        birim_id = data.get('birim_id')
        personeller_data = data.get('personeller', [])

        if not donem_str or not birim_id:
            return JsonResponse({'status': 'error', 'message': 'Dönem veya Birim ID eksik.'}, status=400)
        
        # Yetki kontrolü
        if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
            if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
                return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
        
        donem = datetime.strptime(donem_str, '%Y-%m')
        baslangic = donem.replace(day=1).date()
        bitis = donem.replace(day=calendar.monthrange(donem.year, donem.month)[1]).date()
        
        birim = get_object_or_404(Birim.objects.select_related('HSAKodu'), BirimId=birim_id)
        alan_kodu = birim.HSAKodu.AlanKodu if birim.HSAKodu else None
        
        created_count = 0
        error_list = []

        with transaction.atomic():
            for p_data in personeller_data:
                tc_kimlik = p_data.get('tc_kimlik', '').strip()
                adi = p_data.get('adi', '').strip()
                soyadi = p_data.get('soyadi', '').strip()
                
                if not (tc_kimlik and adi and soyadi):
                    error_list.append(f"Eksik bilgi (TC: {tc_kimlik}, Ad: {adi}, Soyad: {soyadi})")
                    continue

                try:
                    # Personel'i al veya oluştur/güncelle
                    personel, p_created = Personel.objects.update_or_create(
                        TCKimlikNo=tc_kimlik,
                        defaults={'PersonelAdi': adi, 'PersonelSoyadi': soyadi}
                    )
                    
                    # Her zaman yeni HizmetSunumCalismasi kaydı oluştur
                    HizmetSunumCalismasi.objects.create(
                        PersonelId=personel,
                        CalisilanBirimId=birim,
                        Donem=baslangic,
                        HizmetBaslangicTarihi=baslangic,
                        HizmetBitisTarihi=bitis,
                        CreatedBy=request.user,
                        OzelAlanKodu=alan_kodu,
                        Sorumlu=False,
                        Sertifika=False,
                        Kesinlestirme=False
                    )
                    created_count += 1
                except Exception as inner_e:
                    error_list.append(f"TC: {tc_kimlik} için hata: {str(inner_e)}")

        message = f"{created_count} personel için bildirim kaydı oluşturuldu."
        if error_list:
             message += f" Hatalar: {'; '.join(error_list)}"
             status_code = 207 if created_count > 0 else 400
             return JsonResponse({'status': 'warning' if created_count > 0 else 'error', 'message': message}, status=status_code)

        return JsonResponse({'status': 'success', 'message': message})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Geçersiz JSON formatı.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Genel hata: {str(e)}'}, status=500)

@login_required
@require_GET
def bildirimler_listele(request, year, month, birim_id):
    """
    Seçili yıl, ay ve birim için bildirimleri JSON olarak döndürür.
    (Mevcut fonksiyon doğru görünüyor, yetki kontrolü eklenebilir)
    """
    # Yetki kontrolü
    if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
        if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
            return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
    # Yetkisi varsa veya UserBirim ilişkisi varsa devam et
    try:
        donem = date(year, month, 1)
        # İlgili birim nesnesini de alalım (HSA adı için)
        birim = get_object_or_404(Birim.objects.select_related('HSAKodu'), BirimId=birim_id)
        hsa_adi = birim.HSAKodu.AlanAdi if birim.HSAKodu else "Tanımsız"

        bildirimler = HizmetSunumCalismasi.objects.filter(
            CalisilanBirimId_id=birim_id,
            Donem=donem
        ).select_related('PersonelId').order_by('PersonelId__PersonelSoyadi', 'PersonelId__PersonelAdi') # Sıralama ekleyelim

        data = []
        for b in bildirimler:
            # getattr kullanımı yerine doğrudan erişim daha okunaklı olabilir
            personel = b.PersonelId
            data.append({
                'id': b.CalismaId,
                'tc_kimlik_no': personel.TCKimlikNo if personel else '',
                'ad': personel.PersonelAdi if personel else '',
                'soyad': personel.PersonelSoyadi if personel else '',
                'baslangic': b.HizmetBaslangicTarihi.strftime('%Y-%m-%d') if b.HizmetBaslangicTarihi else '',
                'bitis': b.HizmetBitisTarihi.strftime('%Y-%m-%d') if b.HizmetBitisTarihi else '',
                'ozel_alan_kodu': b.OzelAlanKodu,
                'sorumlu': b.Sorumlu,
                'sertifika': b.Sertifika,
                'kesinlesmis': b.Kesinlestirme,
            })

        # Başarı durumunda seçili birim bilgilerini de gönderelim
        return JsonResponse({
            'status': 'success', 
            'data': data,
            'birim_adi': birim.BirimAdi,
            'hsa_adi': hsa_adi,
            'alan_kodu': birim.HSAKodu.AlanKodu if birim.HSAKodu else None
        })
    except Exception as e:
        # Hata loglama eklenebilir
        return JsonResponse({'status': 'error', 'message': f'Bildirimler listelenirken hata: {str(e)}'}, status=500)

@csrf_exempt
@login_required
def bildirimler_kesinlestir(request):
    """
    Bildirimleri kesinleştirir.
    Sadece çakışmasız kayıtları kesinleştirir, çakışanları kullanıcıya bildirir.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            donem = data.get('donem')
            birim_id = data.get('birim_id')
            bildirimler_data = data.get('bildirimler')

            if not all([donem, birim_id, bildirimler_data]):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Eksik parametreler.'
                })

            donem_date = datetime.strptime(donem, '%Y-%m').date()
            birim = get_object_or_404(Birim, BirimId=birim_id)

            # Çakışma kontrolü: sadece çakışmasız kayıtlar kesinleştirilsin
            kaydedilenler = 0
            errors = []
            with transaction.atomic():
                for bildirim_data in bildirimler_data:
                    bildirim_id = bildirim_data.get('id')
                    if not bildirim_id:
                        continue
                    try:
                        bildirim = HizmetSunumCalismasi.objects.get(CalismaId=bildirim_id)
                    except HizmetSunumCalismasi.DoesNotExist:
                        errors.append({'id': bildirim_id, 'message': 'Kayıt bulunamadı.'})
                        continue

                    # Çakışma kontrolü
                    cakisma_sorgusu = Q(
                        PersonelId=bildirim.PersonelId,
                        HizmetBaslangicTarihi__lte=datetime.strptime(bildirim_data['bitis'], '%Y-%m-%d').date(),
                        HizmetBitisTarihi__gte=datetime.strptime(bildirim_data['baslangic'], '%Y-%m-%d').date()
                    ) & ~Q(CalismaId=bildirim_id)
                    cakisan_kayitlar = HizmetSunumCalismasi.objects.filter(cakisma_sorgusu)
                    if cakisan_kayitlar.exists():
                        hata_mesaji = "Kesinleştirilemedi, çakışan kayıtlar: " + "; ".join([f"{c.CalisilanBirimId.BirimAdi} {c.HizmetBaslangicTarihi.strftime('%d.%m')}-{c.HizmetBitisTarihi.strftime('%d.%m.%Y')}" for c in cakisan_kayitlar])
                        errors.append({'id': bildirim_id, 'message': hata_mesaji})
                        continue

                    # Verileri güncelle ve kesinleştir
                    bildirim.HizmetBaslangicTarihi = datetime.strptime(bildirim_data['baslangic'], '%Y-%m-%d').date()
                    bildirim.HizmetBitisTarihi = datetime.strptime(bildirim_data['bitis'], '%Y-%m-%d').date()
                    bildirim.Sorumlu = bildirim_data['sorumlu']
                    bildirim.Sertifika = bildirim_data['sertifika']
                    bildirim.Kesinlestirme = True
                    bildirim.save()
                    kaydedilenler += 1

            if errors:
                msg = f"{kaydedilenler} kayıt kesinleştirildi. {len(errors)} kayıt kesinleştirilemedi."
                return JsonResponse({
                    'status': 'partial' if kaydedilenler > 0 else 'error',
                    'message': msg,
                    'errors': errors
                }, status=207 if kaydedilenler > 0 else 400)
            return JsonResponse({
                'status': 'success',
                'message': 'Tüm kayıtlar başarıyla kesinleştirildi.'
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })

    return JsonResponse({
        'status': 'error',
        'message': 'Geçersiz istek metodu.'
    })

@csrf_exempt
@login_required
@require_POST
def bildirimler_kesinlestirmeyi_kaldir(request):
    """
    Birim ve dönem bazında tüm HizmetSunumCalismasi kayıtlarının Kesinlestirme alanını False yapar.
    Sadece 'HSA Bildirim Kesinleştirme' yetkisi olan kullanıcılar erişebilir.
    POST ile birim_id ve donem (YYYY-MM) parametreleri bekler.
    """
    try:
        data = json.loads(request.body)
        birim_id = data.get('birim_id')
        donem_str = data.get('donem')
        if not birim_id or not donem_str:
            return JsonResponse({'status': 'error', 'message': 'Birim ID ve dönem zorunludur.'}, status=400)
        if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
            return JsonResponse({'status': 'error', 'message': 'Yetkiniz yok.'}, status=403)
        try:
            donem = datetime.strptime(donem_str, '%Y-%m').date()
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'Geçersiz dönem formatı.'}, status=400)
        with transaction.atomic():
            calismalar = HizmetSunumCalismasi.objects.filter(CalisilanBirimId_id=birim_id, Donem=donem, Kesinlestirme=True)
            updated = calismalar.update(Kesinlestirme=False)
        return JsonResponse({'status': 'success', 'message': f'{updated} kayıt kesinleştirmesi kaldırıldı.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'İşlem sırasında hata oluştu: {str(e)}'}, status=500)

@login_required
@require_POST
def bildirim_sil(request, bildirim_id):
    """Bildirim kaydını siler"""
    try:
        # Her zaman JSON bekle (fetch ile gönderiliyor)
        try:
            data = json.loads(request.body)
        except Exception:
            data = {}
        birim_id = data.get('birim_id')
        if not birim_id:
            return JsonResponse({'status': 'error', 'message': 'Birim ID eksik.'}, status=400)

        print(f"Bildirim silme isteği: Bildirim ID: {bildirim_id}, Birim ID: {birim_id}")

        try:
            bildirim_id_int = int(bildirim_id)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Geçersiz Bildirim ID.'}, status=400)
        
        bildirim = get_object_or_404(HizmetSunumCalismasi, CalismaId=bildirim_id_int)
        
        # Yetki kontrolü
        if not request.user.has_permission('HSA Bildirim Kesinleştirme'):
            if not UserBirim.objects.filter(user=request.user, birim=birim_id).exists():
                return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
        
        # Kesinleşmiş kayıt silinemez kontrolü
        if bildirim.Kesinlestirme:
            return JsonResponse({'status': 'error', 'message': 'Kesinleşmiş kayıtlar silinemez.'}, status=400)

        bildirim.delete()
        return JsonResponse({'status': 'success', 'message': 'Bildirim başarıyla silindi.'})
    except HizmetSunumCalismasi.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Bildirim bulunamadı.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Bildirim silinirken hata oluştu: {str(e)}'}, status=500)


@login_required
def bildirim_yazdir(request, year=None, month=None, birim_id=None):
    """
    Seçili yıl, ay ve birim(ler) için bildirimleri PDF olarak döndürür.
    birim_id: tek birim için int veya birden fazla için '1,2,3' gibi string olabilir.
    Ayrıca GET ile birim_id'ler alınabilir: ?donem=YYYY-MM&birim_id=1,2,3
    """
    # Parametreli çağrı mı, query string mi?
    if year is not None and month is not None:
        try:
            year = int(year)
            month = int(month)
            donem = f"{year}-{month:02d}"
        except Exception:
            return HttpResponse('Geçersiz yıl/ay parametresi.', status=400)
    else:
        donem = request.GET.get('donem')
        if not donem or '-' not in donem:
            return HttpResponse('Dönem (YYYY-MM) parametresi eksik veya hatalı.', status=400)
        try:
            year, month = map(int, donem.split('-'))
        except Exception:
            return HttpResponse('Dönem (YYYY-MM) parametresi hatalı.', status=400)

    # Birim ID'leri al
    birim_ids = []
    if birim_id:
        if ',' in str(birim_id):
            birim_ids = [int(bid) for bid in str(birim_id).split(',') if bid.strip().isdigit()]
        else:
            birim_ids = [int(birim_id)]
    else:
        birim_id_param = request.GET.get('birim_id')
        if birim_id_param:
            birim_ids = [int(bid) for bid in birim_id_param.split(',') if bid.strip().isdigit()]

    if not birim_ids:
        return HttpResponse('Birim seçilmedi.', status=400)

    birim_bildirimleri = []
    for bid in birim_ids:
        # Her bir birim için bildirimleri çek
        response = bildirimler_listele(request, year, month, bid)
        if hasattr(response, 'content'):
            data = json.loads(response.content)
        else:
            data = response
        bildirimler = data.get('data', [])
        # Birim adı ve HSA adı
        birim_adi = data.get('birim_adi', f'Birim {bid}')
        hsa_adi = data.get('hsa_adi', '')
        alan_kodu = data.get('alan_kodu', '')
        birim_bildirimleri.append({
            'birim_id': bid,
            'birim_adi': birim_adi,
            'hsa_adi': hsa_adi,
            'alan_kodu': alan_kodu,
            'bildirimler': bildirimler
        })

    # PDF için template render
    template = get_template('hizmet_sunum_app/bildirim_formu.html')
    html = template.render({
        'birim_bildirimleri': birim_bildirimleri,
        'now': django_now(),
        'donem': donem
    })

    options = {
        'page-size': 'A4',
        'orientation': 'Portrait',
        'margin-top': '1.5cm',
        'margin-right': '1.5cm',
        'margin-bottom': '1.1cm',
        'margin-left': '1.5cm',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': True,
        'enable-external-links': True,
        'quiet': ''
    }
    config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
    pdf = pdfkit.from_string(html, False, options=options, configuration=config)

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="hizmet_sunum_bildirim_{year}_{month}.pdf"'
    return response

def raporlama(request):
    calismalar_by_birim = None
    donem = None
    kurum = None
    idare = None
    excel_url = None
    error_message = None
    info_message = None
    birimler_json = None
    birimler = Birim.objects.all()
    kurumlar = birimler.values_list('KurumAdi', flat=True).distinct()
    idareler = birimler.values_list('IdareAdi', flat=True).distinct()

    # Form yerine GET parametrelerini doğrudan al
    donem_str = request.GET.get('donem')
    kurum = request.GET.get('kurum')
    idare = request.GET.get('idare')
    durum = request.GET.get('durum')  # "1", "0" veya ""

    toplam_kayit = 0

    # Dönem parametresi varsa veriyi çek
    if donem_str:
        try:
            # 'YYYY-MM' formatındaki stringi ayın ilk günü olan date objesine çevir
            donem = datetime.strptime(donem_str, "%Y-%m").date()

            calismalar_query = HizmetSunumCalismasi.objects.select_related(
                'PersonelId',
                'CalisilanBirimId'
            ).filter(
                 Donem=donem
            )

            if kurum:
                calismalar_query = calismalar_query.filter(
                    CalisilanBirimId__KurumAdi=kurum
                )
            if idare:
                calismalar_query = calismalar_query.filter(
                    CalisilanBirimId__IdareAdi=idare
                )

            if durum == "1":
                calismalar_query = calismalar_query.filter(Kesinlestirme=True)
            elif durum == "0":
                calismalar_query = calismalar_query.filter(Kesinlestirme=False)

            # Birimlere göre grupla ve her birim altındaki çalışmaları Prefetch ile çek
            birim_ids_with_calisma = calismalar_query.values_list('CalisilanBirimId', flat=True).distinct()

            birimler_with_calismalar = Birim.objects.filter(BirimId__in=birim_ids_with_calisma).prefetch_related(
                Prefetch(
                    'hizmetsunumcalismasi_set',
                    queryset=calismalar_query.order_by('PersonelId__PersonelSoyadi', 'PersonelId__PersonelAdi', 'HizmetBaslangicTarihi'),
                    to_attr='calismalar'
                )
            ).order_by('BirimAdi')
                # Birimleri JSON'a çevir
            birimler_json = json.dumps([
                {'BirimId': b.BirimId, 'BirimAdi': b.BirimAdi} for b in birimler_with_calismalar
            ])
            calismalar_by_birim = []
            for birim in birimler_with_calismalar:
                if birim.calismalar:
                    kesinlesmis = sum(1 for c in birim.calismalar if c.Kesinlestirme)
                    beklemede = sum(1 for c in birim.calismalar if not c.Kesinlestirme)
                    calismalar_by_birim.append({
                        'birim': birim,
                        'calismalar': birim.calismalar,
                        'personel_sayisi': len(set([c.PersonelId.PersonelId for c in birim.calismalar])),
                        'kesinlesmis_sayisi': kesinlesmis,
                        'beklemede_sayisi': beklemede,
                    })

            toplam_kayit = sum(len(birim['calismalar']) for birim in calismalar_by_birim)

            # Excel indirme linki oluştur
            if calismalar_by_birim:
                 excel_url = reverse('hizmet_sunum_app:export_raporlama_excel')
                 excel_url += f'?donem={donem.strftime("%Y-%m")}'

                 if kurum:
                     excel_url += f'&kurum={kurum}'
                 if durum:
                     excel_url += f'&durum={durum}'
                 if idare:
                     excel_url += f'&idare={idare}'

            if toplam_kayit == 0:
                 info_message = "Seçilen dönem ve kuruma ait hizmet sunum çalışması bulunamadı."

        except ValueError:
             error_message = "Geçersiz dönem formatı seçildi."
        except Exception as e:
            # Hata yönetimi
            error_message = f"Rapor oluşturulurken bir hata oluştu: {str(e)}"
            print(f"Raporlama Hatası: {e}") # Konsola yazdır (geliştirme için)
    else:
        # İlk sayfa yüklemesi veya dönem seçilmemişse
        info_message = "Lütfen bir dönem ve isteğe bağlı kriterlerinizi seçerek raporlayın."

    user_permissions = get_user_permissions(request.user)
    context = {
        # 'form': form, # Form kaldırıldı
        'calismalar_by_birim': calismalar_by_birim,
        'birimler': birimler, # Kurum seçimi için tüm birimleri geçiyoruz (kurum listesi çekmek için)
        'birimler_json': birimler_json,
        'kurumlar': kurumlar, # Kurum seçimi için tüm kurumları geçiyoruz
        'idareler': idareler,
        'selected_donem': donem_str, # Şablona dönemin string halini gönderelim ki selectbox'ta seçili kalsın
        'selected_kurum': kurum,
        'selected_idare': idare,
        'selected_durum': durum,
        'excel_url': excel_url,
        # 'is_form_valid': is_form_valid, # Form kaldırıldı
        'error_message': error_message,
        'info_message': info_message,
        'toplam_kayit': toplam_kayit,
        'user_permissions': user_permissions,
    }
    return render(request, 'hizmet_sunum_app/raporlama.html', context)

def export_raporlama_excel(request):
    donem_str = request.GET.get('donem')
    kurum = request.GET.get('kurum')
    idare = request.GET.get('idare')

    if not donem_str:
        messages.error(request, "Lütfen bir dönem seçin.")
        return redirect('hizmet_sunum_app:raporlama')

    try:
        donem = datetime.strptime(donem_str, "%Y-%m").date()
    except ValueError:
         messages.error(request, "Geçersiz dönem formatı.")
         return redirect('hizmet_sunum_app:raporlama')

    calismalar_query = HizmetSunumCalismasi.objects.select_related(
        'PersonelId',
        'CalisilanBirimId'
    ).filter(
        Donem=donem
    )

    if kurum:
        calismalar_query = calismalar_query.filter(
            CalisilanBirimId__KurumAdi=kurum
        )
    if idare:
        calismalar_query = calismalar_query.filter(
            CalisilanBirimId__IdareAdi=idare
        )

    calismalar = calismalar_query.order_by(
        'CalisilanBirimId__BirimAdi',
        'PersonelId__PersonelSoyadi',
        'PersonelId__PersonelAdi',
        'HizmetBaslangicTarihi'
    )

    if not calismalar.exists():
        messages.warning(request, "Seçilen filtreye uygun veri bulunamadı.")
        return redirect('hizmet_sunum_app:raporlama')

    # Şablon dosyasını aç
    template_path = os.path.join(settings.STATIC_ROOT, 'excels', 'HizmetSunumSablon.xlsx')

    # Şablon dosyasının varlığını kontrol et
    if not os.path.exists(template_path):
         messages.error(request, f"Excel şablon dosyası bulunamadı: {template_path}")
         return redirect('hizmet_sunum_app:raporlama')

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Verileri yaz (2. satırdan başla)
    row = 2
    for calisma in calismalar:
        personel = calisma.PersonelId
        birim = calisma.CalisilanBirimId

        ws.cell(row=row, column=1, value=personel.TCKimlikNo)
        ws.cell(row=row, column=2, value=personel.PersonelAdi)
        ws.cell(row=row, column=3, value=personel.PersonelSoyadi)
        ws.cell(row=row, column=4, value=calisma.HizmetBaslangicTarihi)
        ws.cell(row=row, column=5, value=calisma.HizmetBitisTarihi)
        ws.cell(row=row, column=6, value=birim.BirimAdi)
        ws.cell(row=row, column=7, value=calisma.OzelAlanKodu)
        ws.cell(row=row, column=10, value=birim.KurumAdi)
        ws.cell(row=row, column=11, value=birim.IdareAdi)
        # Sütun eklenirse devam ettir
        if calisma.Sorumlu:
            sorumlu_alan = None
            if calisma.OzelAlanKodu:
                try:
                    sorumlu_alan = HizmetSunumAlani.objects.get(AlanKodu=calisma.OzelAlanKodu)
                except HizmetSunumAlani.DoesNotExist:
                    sorumlu_alan = None
            # SorumluAtanabilir True ise 1, False ise 0 yaz
            if sorumlu_alan is not None and hasattr(sorumlu_alan, "SorumluAtanabilir"):
                ws.cell(row=row, column=8, value=1 if sorumlu_alan.SorumluAtanabilir else 0)
            else:
                ws.cell(row=row, column=8, value="")
        else:
            ws.cell(row=row, column=8, value=0)
        ws.cell(row=row, column=9, value=1 if calisma.Sertifika else 0)
        row += 1

    # Excel dosyasını kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # HTTP response oluştur
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    file_name = f"hizmet_sunum_rapor_{donem_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

def birim_yetkililer(request, birim_id):
    # Birime atanmış kullanıcıları getir
    yetkiler = UserBirim.objects.filter(birim_id=birim_id).select_related('user')
    data = [
        {
            "username": y.user.Username,
            "full_name": y.user.FullName,
        }
        for y in yetkiler
    ]
    return JsonResponse({"status": "success", "data": data})

def kullanici_ara(request):
    username = request.GET.get('username', '').strip()
    try:
        user = User.objects.get(Username=username)
        data = {
            "username": user.Username,
            "full_name": user.FullName
        }
        return JsonResponse({"status": "success", "data": data})
    except User.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Kullanıcı bulunamadı."})

@csrf_exempt
@require_POST
def birim_yetki_ekle(request, birim_id):
    import json
    try:
        body = json.loads(request.body)
        username = body.get('username')
        user = User.objects.get(Username=username)
        birim = Birim.objects.get(pk=birim_id)
        obj, created = UserBirim.objects.get_or_create(user=user, birim=birim)
        if created:
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "message": "Kullanıcı zaten yetkili."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@csrf_exempt
@require_POST
def birim_yetki_sil(request, birim_id):
    import json
    try:
        body = json.loads(request.body)
        username = body.get('username')
        user = User.objects.get(Username=username)
        birim = Birim.objects.get(pk=birim_id)
        deleted, _ = UserBirim.objects.filter(user=user, birim=birim).delete()
        if deleted:
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "message": "Yetki bulunamadı."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required
def birim_yonetim(request):
    birimler = Birim.objects.select_related('HSAKodu').all()
    # Modal için değerler
    kurumlar = Kurum.objects.all()
    idareler = Idare.objects.all()
    hsa_listesi = HizmetSunumAlani.objects.all().order_by('AlanAdi')

    birim_list = []
    for birim in birimler:
        yetkiler = UserBirim.objects.filter(birim=birim).select_related('user')
        yetkili_users = [
            {
                "username": y.user.Username,
                "full_name": y.user.FullName,
            }
            for y in yetkiler
        ]
        birim_list.append({
            "id": birim.BirimId,
            "adi": birim.BirimAdi,
            "kurum": birim.KurumAdi,
            "idare": birim.IdareAdi,
            "hsa_kodu": birim.HSAKodu.AlanKodu,
            "hsa_adi": birim.HSAKodu.AlanAdi,
            "yetkili_sayisi": len(yetkili_users),
            "yetkililer": yetkili_users,
        })
    context = {
        "kurumlar": kurumlar,
        "idareler": idareler,
        "hsa_listesi": hsa_listesi,
        "birimler": birim_list
    }
    return render(request, "hizmet_sunum_app/birim_yonetim.html", context)

@csrf_exempt
@require_POST
def birim_yetki_sil(request, birim_id):
    import json
    try:
        body = json.loads(request.body)
        username = body.get('username')
        user = User.objects.get(Username=username)
        birim = Birim.objects.get(pk=birim_id)
        deleted, _ = UserBirim.objects.filter(user=user, birim=birim).delete()
        if deleted:
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "message": "Yetki bulunamadı."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required
def birim_yonetim(request):
    birimler = Birim.objects.select_related('HSAKodu').all()
    # Modal için değerler
    kurumlar = Kurum.objects.all()
    idareler = Idare.objects.all()
    hsa_listesi = HizmetSunumAlani.objects.all().order_by('AlanAdi')
    

    birim_list = []
    for birim in birimler:
        yetkiler = UserBirim.objects.filter(birim=birim).select_related('user')
        yetkili_users = [
            {
                "username": y.user.Username,
                "full_name": y.user.FullName,
            }
            for y in yetkiler
        ]
        birim_list.append({
            "id": birim.BirimId,
            "adi": birim.BirimAdi,
            "kurum": birim.KurumAdi,
            "idare": birim.IdareAdi,
            "hsa_kodu": birim.HSAKodu.AlanKodu,
            "hsa_adi": birim.HSAKodu.AlanAdi,
            "yetkili_sayisi": len(yetkili_users),
            "yetkililer": yetkili_users,
        })
    context = {
            "kurumlar": kurumlar,
            "idareler": idareler,
            "hsa_listesi": hsa_listesi,
            "birimler": birim_list
            }
    return render(request, "hizmet_sunum_app/birim_yonetim.html", context)