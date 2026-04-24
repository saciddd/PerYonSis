import json
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Sertifika
from hizmet_sunum_app.models import Personel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

@csrf_exempt
@login_required
def sertifika_guncelle(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            tc_kimlik_no = data.get('tc_kimlik_no')
            aciklama = data.get('sertifika_aciklamasi')
            baslangic = data.get('baslangic_tarihi')
            bitis = data.get('bitis_tarihi')
            alanda_kullaniliyor = data.get('alanda_kullaniliyor', False)

            if not all([tc_kimlik_no, aciklama, baslangic, bitis]):
                return JsonResponse({'status': 'error', 'message': 'Eksik veri gönderildi.'}, status=400)

            personel = Personel.objects.filter(TCKimlikNo=tc_kimlik_no).first()
            if not personel:
                return JsonResponse({'status': 'error', 'message': 'Personel bulunamadı.'}, status=404)

            sertifika, created = Sertifika.objects.update_or_create(
                personel=personel,
                defaults={
                    'sertifika_aciklamasi': aciklama,
                    'baslangic_tarihi': baslangic,
                    'bitis_tarihi': bitis,
                    'alanda_kullaniliyor': alanda_kullaniliyor
                }
            )

            return JsonResponse({'status': 'success', 'message': 'Sertifika güncellendi.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Geçersiz istek metodu.'}, status=405)

@login_required
def sertifikali_personeller_raporu(request):
    sertifikalar = list(Sertifika.objects.select_related('personel').all())
    
    today = datetime.now().date()
    kurumlar = set()
    birimler = set()
    
    for s in sertifikalar:
        if s.bitis_tarihi:
            s.kalan_gun = (s.bitis_tarihi - today).days
        else:
            s.kalan_gun = None
            
        if s.personel and s.personel.Birim:
            if s.personel.Birim.KurumAdi:
                kurumlar.add(s.personel.Birim.KurumAdi)
            if s.personel.Birim.BirimAdi:
                birimler.add(s.personel.Birim.BirimAdi)
                
    def get_sort_key(s):
        personel = s.personel
        if personel and personel.Birim:
            b = personel.Birim
            return (b.KurumAdi or "", b.IdareAdi or "", b.BirimAdi or "", personel.PersonelAdi or "")
        return ("ZZZ", "ZZZ", "ZZZ", personel.PersonelAdi if personel else "")
        
    sertifikalar.sort(key=get_sort_key)
    
    return render(request, 'ik_core/sertifika_raporu.html', {
        'sertifikalar': sertifikalar,
        'kurumlar': sorted(list(kurumlar)),
        'birimler': sorted(list(birimler)),
    })

@login_required
def sertifikali_personeller_excel_export(request):
    sertifikalar = list(Sertifika.objects.select_related('personel').all())
    
    def get_sort_key(s):
        personel = s.personel
        if personel and personel.Birim:
            b = personel.Birim
            return (b.KurumAdi or "", b.IdareAdi or "", b.BirimAdi or "", personel.PersonelAdi or "")
        return ("ZZZ", "ZZZ", "ZZZ", personel.PersonelAdi if personel else "")
        
    sertifikalar.sort(key=get_sort_key)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sertifikalı Personeller"

    # Header styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "Sıra No", "Kurum", "İdare", "Birim", "T.C. Kimlik No", "Adı Soyadı", 
        "Sertifika Açıklaması", "Başlangıç Tarihi", 
        "Bitiş Tarihi", "Alanda Kullanılıyor"
    ]
    
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border

    for index, sertifika in enumerate(sertifikalar, start=1):
        baslangic = sertifika.baslangic_tarihi.strftime("%d.%m.%Y") if sertifika.baslangic_tarihi else ""
        bitis = sertifika.bitis_tarihi.strftime("%d.%m.%Y") if sertifika.bitis_tarihi else ""
        alanda_kullanim = "Evet" if sertifika.alanda_kullaniliyor else "Hayır"
        personel = sertifika.personel
        birim = personel.Birim if personel else None
        
        row = [
            index,
            birim.KurumAdi if birim else "-",
            birim.IdareAdi if birim else "-",
            birim.BirimAdi if birim else "-",
            personel.TCKimlikNo if personel else "",
            f"{personel.PersonelAdi} {personel.PersonelSoyadi}" if personel else "",
            sertifika.sertifika_aciklamasi,
            baslangic,
            bitis,
            alanda_kullanim
        ]
        ws.append(row)
        
        # Apply borders to the row
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=index + 1, column=col)
            cell.border = thin_border
            if col in [1, 5, 8, 9, 10]:  # Center some columns
                cell.alignment = alignment_center

    # Adjust column widths
    column_widths = {
        'A': 10,  # Sıra No
        'B': 20,  # Kurum
        'C': 20,  # İdare
        'D': 25,  # Birim
        'E': 15,  # T.C. Kimlik No
        'F': 25,  # Adı Soyadı
        'G': 35,  # Sertifika Açıklaması
        'H': 15,  # Başlangıç Tarihi
        'I': 15,  # Bitiş Tarihi
        'J': 20   # Alanda Kullanılıyor
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sertifikali_personeller_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    
    return response
