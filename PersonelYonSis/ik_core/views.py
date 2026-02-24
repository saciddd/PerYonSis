from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
import re
from django.db.models.query import QuerySet
from django.db.models import Q, Value as V
from django.db.models.functions import Lower, Concat, Coalesce
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.staticfiles.storage import staticfiles_storage
from django.template.loader import render_to_string, get_template
import pdfkit
from django.urls import reverse, reverse_lazy
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib import messages # Import messages framework
from django.contrib.auth.decorators import login_required
from .models.personel import Personel, KisaUnvan, Brans, Kurum, OzelDurum, UnvanBransEslestirme, Unvan
from .models.GeciciGorev import GeciciGorev
from .models import UstBirim, Bina, Birim, PersonelBirim, Kampus
# Import value lists needed
from .models.valuelists import (
    TESKILAT_DEGERLERI, EGITIM_DEGERLERI, MAZERET_DEGERLERI,
    AYRILMA_NEDENI_DEGERLERI, ENGEL_DERECESI_DEGERLERI # Add necessary value lists
)
from .forms import PersonelForm, KurumForm, UnvanForm, BransForm, GeciciGorevForm
from django.db import models
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_GET
from django.forms import ModelForm
from .models.Teblig import TebligImzasi, TebligMetni
from django.utils import timezone
from django.db.models import OuterRef, Exists
from datetime import date

# Türkçe lower/upper normalize edici yardımcı
def lower_tr(text: str) -> str:
    """
    Türkçe küçük harf dönüşümü.
    Normal .lower() "I/İ/ı/i" harflerinde sorun çıkarabilir.
    """
    if not text:
        return text
    return (text
        .replace("I", "ı")
        .replace("İ", "i")
        .lower()
    )

@login_required
def personel_list(request):
    personeller = _get_filtered_personel_list(request.GET)
    kisa_unvanlar = KisaUnvan.objects.select_related('ust_birim').order_by('ust_birim__ad', 'ad')

    # Export için alan seçenekleri
    export_fields = [
        ('tc_kimlik_no', 'T.C. Kimlik No'),
        ('ad', 'Ad'),
        ('soyad', 'Soyad'),
        ('yas', 'Yaş'),
        ('dogum_tarihi', 'Doğum Tarihi'),
        ('cinsiyet', 'Cinsiyet'),
        ('telefon', 'Telefon'),
        ('kisa_unvan', 'Kısa Unvan'),
        ('baslangic_tarihi', 'Başlangıç Tarihi'),
        ('durum', 'Durum'),
        ('unvan', 'Unvan'),
        ('brans', 'Branş'),
        ('sicil_no', 'Sicil No'),
        ('kurum', 'Kurum'),
        ('kadro_yeri', 'Kadro Yeri'),
        ('fiili_gorev_yeri', 'Fiili Görev Yeri'),
        ('tahsil_durumu', 'Öğrenim Durumu'),
        ('memuriyet_durumu', 'Memuriyet Durumu'),
        ('ozel_durumu', 'Özel Durumu'),
        ('son_mercis657_listesi', 'Son Mercis657 Listesi'),
        ('son_hizmet_sunum_bildirimi', 'Son Hizmet Sunum Bildirimi'),
    ]

    return render(request, 'ik_core/personel_list.html', {
        'personeller': personeller,
        'personel_sayisi': len(personeller),
        'kisa_unvanlar': kisa_unvanlar,
        'export_fields': export_fields,
        'ust_birimler': UstBirim.objects.all().order_by('ad'),
        'binalar': Bina.objects.all().order_by('ad'),
        'arama': {
            'tc_kimlik_no': request.GET.get('tc_kimlik_no', ''),
            'ad_soyad': request.GET.get('ad_soyad', ''),
            'telefon': request.GET.get('telefon', ''),
            'kisa_unvan': request.GET.getlist('kisa_unvan'),
            'durum': request.GET.get('durum', ''),
            'tc_list': request.GET.get('tc_list', ''),
            'ad_soyad_list': request.GET.get('ad_soyad_list', ''),
        }
    })


def _get_filtered_personel_list(data, limit_limitless=True):
    query = Q()
    tc_kimlik_no = data.get('tc_kimlik_no', '').strip()
    ad_soyad = data.get('ad_soyad', '').strip()
    ad_soyad_norm = lower_tr(ad_soyad) if ad_soyad else ''
    telefon = data.get('telefon', '').strip()
    kisa_unvan_list = data.getlist('kisa_unvan')
    durum = data.get('durum', '')

    if tc_kimlik_no:
        query &= Q(tc_kimlik_no__icontains=tc_kimlik_no)

    # Toplu Arama (TC Listesi)
    tc_list_str = data.get('tc_list', '').strip()
    if tc_list_str:
        tcs = [t.strip() for t in re.split(r'[,\n\r]+', tc_list_str) if t.strip()]
        if tcs:
            query &= Q(tc_kimlik_no__in=tcs)
            # Toplu aramada limit olmamalı
            limit_limitless = False


    if telefon:
        query &= Q(telefon__icontains=telefon)
    if kisa_unvan_list:
        # DB'de filtreleme yapmak için: Seçilen Kısa Unvan'lara ait Unvan-Branş eşleşmelerini bul
        eslesmeler = UnvanBransEslestirme.objects.filter(kisa_unvan_id__in=kisa_unvan_list).values_list('unvan_id', 'brans_id')
        
        kisa_unvan_query = Q()
        for u_id, b_id in eslesmeler:
            if b_id:
                kisa_unvan_query |= Q(unvan_id=u_id, brans_id=b_id)
            else:
                kisa_unvan_query |= Q(unvan_id=u_id, brans__isnull=True)
        
        if kisa_unvan_query:
            query &= kisa_unvan_query
        else:
            # Eşleşme yoksa boş döndür
            query &= Q(pk__in=[])

    if limit_limitless and not any([tc_kimlik_no, ad_soyad, telefon, kisa_unvan_list, durum, tc_list_str, data.get('ad_soyad_list')]):
        return list(Personel.objects.select_related('unvan', 'kurum').order_by('-kayit_tarihi')[:10])

    # Önce diğer filtreleri uygula (performans için)
    queryset = Personel.objects.filter(query).select_related('unvan', 'kurum', 'brans', 'kadro_yeri', 'fiili_gorev_yeri')
    
    personeller = list(queryset) # QuerySet'i listeye çevir (Python tarafı filtreleme için)

    # Ad/soyad filtresini Python tarafında uygula (Türkçe karakter desteği için)
    if ad_soyad_norm:
        arama_kelimeleri = [kelime.strip() for kelime in ad_soyad_norm.split() if kelime.strip()]
        filtered_results = []
        for p in personeller:
            ad_normalized = lower_tr(p.ad or '')
            soyad_normalized = lower_tr(p.soyad or '')
            tam_ad_soyad = f"{ad_normalized} {soyad_normalized}".strip()
            eslesme = False
            
            if len(arama_kelimeleri) == 1:
                kelime = arama_kelimeleri[0]
                if kelime in ad_normalized or kelime in soyad_normalized:
                    eslesme = True
            else:
                ilk_kelime = arama_kelimeleri[0]
                son_kelime = arama_kelimeleri[-1]
                if ad_soyad_norm in ad_normalized or ad_soyad_norm in soyad_normalized or ad_soyad_norm in tam_ad_soyad:
                    eslesme = True
                elif ad_normalized.startswith(ilk_kelime) and soyad_normalized.startswith(son_kelime):
                    eslesme = True
                elif ilk_kelime in ad_normalized and son_kelime in soyad_normalized:
                    eslesme = True
                elif len(arama_kelimeleri) > 2:
                    if ad_normalized.startswith(ilk_kelime):
                        ad_kalan = ad_normalized[len(ilk_kelime):].strip()
                        ortadaki_eslesiyor = True
                        for i in range(1, len(arama_kelimeleri) - 1):
                            if arama_kelimeleri[i] not in ad_kalan and arama_kelimeleri[i] not in soyad_normalized:
                                ortadaki_eslesiyor = False
                                break
                        if ortadaki_eslesiyor and soyad_normalized.startswith(son_kelime):
                            eslesme = True
            
            if eslesme:
                filtered_results.append(p)
        personeller = filtered_results

    # Toplu İsim Arama
    ad_soyad_list_str = data.get('ad_soyad_list', '').strip()
    if ad_soyad_list_str:
        name_list = [n.strip() for n in re.split(r'[,\n\r]+', ad_soyad_list_str) if n.strip()]
        if name_list:
            filtered_results = []
            for p in personeller:
                ad_normalized = lower_tr(p.ad or '')
                soyad_normalized = lower_tr(p.soyad or '')
                tam_ad_soyad = f"{ad_normalized} {soyad_normalized}".strip()
                
                matched_any = False
                for search_name in name_list:
                    search_norm = lower_tr(search_name)
                    # Basit 'contains' kontrolü yapıyoruz toplu aramada
                    if search_norm in tam_ad_soyad:
                        matched_any = True
                        break
                
                if matched_any:
                    filtered_results.append(p)
            personeller = filtered_results


    # Durum filtresi (Python tarafında)
    if durum:
        if durum in ("Aktif", "Pasif"):
            personeller = [p for p in personeller if (p.durum or "").startswith(durum)]
        else:
            personeller = [p for p in personeller if p.durum == durum]
            
    return personeller


@login_required
def personel_export_xlsx(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    # Seçili alanları al (checkbox name="fields")
    selected_fields = request.POST.getlist('fields')
    if not selected_fields:
        # Default fields if none selected
        selected_fields = ['tc_kimlik_no', 'ad', 'soyad', 'unvan', 'durum']

    # Filtrelenmiş listeyi al (limit_limitless=False -> tüm kayıtlar)
    personeller = _get_filtered_personel_list(request.POST, limit_limitless=False)

    # Alan Tanımları
    FIELD_MAP = {
        'tc_kimlik_no': 'T.C. Kimlik No',
        'ad': 'Ad',
        'soyad': 'Soyad',
        'telefon': 'Telefon',
        'kisa_unvan': 'Kısa Unvan',
        'baslangic_tarihi': 'Başlangıç Tarihi',
        'durum': 'Durum',
        # Ekstra alanlar
        'unvan': 'Unvan',
        'brans': 'Branş',
        'sicil_no': 'Sicil No',
        'kurum': 'Kurum',
        'kadro_yeri': 'Kadro Yeri',
        'fiili_gorev_yeri': 'Fiili Görev Yeri',
        'yas': 'Yaş',
        'dogum_tarihi': 'Doğum Tarihi',
        'cinsiyet': 'Cinsiyet',
        'tahsil_durumu': 'Öğrenim Durumu',
        'memuriyet_durumu': 'Memuriyet Durumu',
        'ozel_durumu': 'Özel Durumu',
        'son_mercis657_listesi': 'Son Mercis657 Listesi',
        'son_hizmet_sunum_bildirimi': 'Son Hizmet Sunum Bildirimi',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personel Listesi"

    # Başlıklar
    headers = [FIELD_MAP.get(f, f) for f in selected_fields]
    ws.append(headers)

    # Veriler
    for p in personeller:
        row = []
        for field in selected_fields:
            val = ""
            if field == 'tc_kimlik_no': val = p.tc_kimlik_no
            elif field == 'ad': val = p.ad
            elif field == 'soyad': val = p.soyad
            elif field == 'telefon': val = p.telefon
            elif field == 'kisa_unvan': val = p.kisa_unvan
            elif field == 'baslangic_tarihi': val = p.goreve_baslama_tarihi
            elif field == 'durum': val = p.durum
            elif field == 'unvan': val = p.unvan.ad if p.unvan else ""
            elif field == 'brans': val = p.brans.ad if p.brans else ""
            elif field == 'sicil_no': val = p.sicil_no
            elif field == 'kurum': val = p.kurum.ad if p.kurum else ""
            elif field == 'kadro_yeri': val = p.kadro_yeri.ad if p.kadro_yeri else ""
            elif field == 'fiili_gorev_yeri': val = p.fiili_gorev_yeri.ad if p.fiili_gorev_yeri else ""
            elif field == 'dogum_tarihi': val = p.dogum_tarihi
            elif field == 'yas': val = p.yas
            elif field == 'cinsiyet': val = p.cinsiyet
            elif field == 'tahsil_durumu': val = p.tahsil_durumu
            elif field == 'memuriyet_durumu': val = p.memuriyet_durumu
            elif field == 'ozel_durumu':
                val = ", ".join([od.ad for od in p.ozel_durumu.all()]) if p.pk else ""
            elif field == 'son_mercis657_listesi': 
                m = p.son_mercis657_listesi
                val = f"{m['yil']}/{m['ay']} {m['birim']}" if m else "-"
            elif field == 'son_hizmet_sunum_bildirimi': 
                h = p.son_hizmet_sunum_bildirimi
                val = f"{h['yil']}/{h['ay']} {h['birim']}" if h else "-"
            row.append(val)
        ws.append(row)
    
    # Sütun Genişlikleri Ayarla
    for i, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Personel_Listesi.xlsx'
    wb.save(response)
    return response

@login_required
def personel_dashboard_export_xlsx(request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    # Tüm personelleri getir (performans için select_related)
    personeller = Personel.objects.select_related(
        'unvan', 'brans', 'kurum', 'kadro_yeri', 'fiili_gorev_yeri', 'unvan_brans_eslestirme__kisa_unvan__ust_birim'
    ).prefetch_related('gecicigorev_set', 'ozel_durumu')

    # Sheet 1: Pano
    wb = openpyxl.Workbook()
    ws_pano = wb.active
    ws_pano.title = "Pano"

    header_fill = PatternFill(start_color="0c0c2e", end_color="0c0c2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws_pano['A1'] = "Durum = 'Aktif'"
    ws_pano['A1'].font = Font(bold=True)
    ws_pano['A2'] = "Üst Birim"
    ws_pano['B2'] = "Kısa Ünvan"
    ws_pano['C2'] = "Kadrolu"
    ws_pano['D2'] = "Geçici"
    
    ws_pano['F1'] = "Durum = 'Pasif'"
    ws_pano['F1'].font = Font(bold=True)
    ws_pano['F2'] = "Üst Birim"
    ws_pano['G2'] = "Kısa Ünvan"
    ws_pano['H2'] = "Kadrolu"
    ws_pano['I2'] = "Geçici"

    for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I']:
        ws_pano[f'{col}2'].fill = header_fill
        ws_pano[f'{col}2'].font = header_font

    # Eşleştirme haritasını oluştur (Kısa Ünvan property olduğu için manuel eşleştirme yapıyoruz)
    eslestirmeler = UnvanBransEslestirme.objects.select_related('kisa_unvan__ust_birim').all()
    eslestirme_map = {(e.unvan_id, e.brans_id): e for e in eslestirmeler}

    aktif_data = {} # (ust_birim, kisa_unvan) -> {'kadrolu': X, 'gecici': Y}
    pasif_data = {}
    doktorlar_list = []
    memur_657_list = []
    surekli_isci_list = []
    
    for p in personeller:
        durum = p.durum or ""
        
        # Kısa Ünvan ve Üst Birim bilgilerini al
        eslesme = p.unvan_brans_eslestirme
        if not eslesme:
            eslesme = eslestirme_map.get((p.unvan_id, p.brans_id))
            
        kisa_unvan_ad = "Belirtilmemiş"
        ust_birim_ad = "Belirtilmemiş"
        
        if eslesme and eslesme.kisa_unvan:
            kisa_unvan_ad = eslesme.kisa_unvan.ad
            if eslesme.kisa_unvan.ust_birim:
                ust_birim_ad = eslesme.kisa_unvan.ust_birim.ad
                
        key = (ust_birim_ad, kisa_unvan_ad)
        is_kadrolu = p.kadrolu_personel is not False
        
        if "Aktif" in durum:
            if key not in aktif_data: aktif_data[key] = {'kadrolu': 0, 'gecici': 0}
            if is_kadrolu: aktif_data[key]['kadrolu'] += 1
            else: aktif_data[key]['gecici'] += 1
        elif "Pasif" in durum:
            if key not in pasif_data: pasif_data[key] = {'kadrolu': 0, 'gecici': 0}
            if is_kadrolu: pasif_data[key]['kadrolu'] += 1
            else: pasif_data[key]['gecici'] += 1
            
        is_aktif_veya_pasif = ('Aktif' in durum or 'Pasif' in durum)
        is_tabip = p.unvan and 'tabip' in p.unvan.ad.lower()
        
        # Determine memuriyet durumu from teskilat
        teskilat = getattr(p, 'teskilat', '')
        is_isci = teskilat in ["İşçi Personel 696 (Döner Sermaye)", "İşçi Personel (Genel Bütçe)"]
        is_memur = not is_isci

        if is_aktif_veya_pasif:
            if is_tabip:
                doktorlar_list.append(p)
            elif is_memur:
                memur_657_list.append(p)
                
            if is_isci:
                surekli_isci_list.append(p)
            
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    row_idx = 3
    aktif_sorted = sorted(aktif_data.items(), key=lambda x: (x[0][0], x[0][1]))
    total_aktif_kadrolu = 0
    total_aktif_gecici = 0
    for i, ((ust_birim, kisa_unvan), counts) in enumerate(aktif_sorted):
        ws_pano[f'A{row_idx}'] = ust_birim
        ws_pano[f'B{row_idx}'] = kisa_unvan
        ws_pano[f'C{row_idx}'] = counts['kadrolu']
        ws_pano[f'D{row_idx}'] = counts['gecici']
        total_aktif_kadrolu += counts['kadrolu']
        total_aktif_gecici += counts['gecici']
        if i % 2 == 1:
            for col in ['A', 'B', 'C', 'D']:
                ws_pano[f'{col}{row_idx}'].fill = alt_fill
        row_idx += 1
        
    ws_pano[f'B{row_idx}'] = "TOPLAM"
    ws_pano[f'B{row_idx}'].font = Font(bold=True)
    ws_pano[f'C{row_idx}'] = total_aktif_kadrolu
    ws_pano[f'C{row_idx}'].font = Font(bold=True)
    ws_pano[f'D{row_idx}'] = total_aktif_gecici
    ws_pano[f'D{row_idx}'].font = Font(bold=True)
    
    row_idx_pasif = 3
    pasif_sorted = sorted(pasif_data.items(), key=lambda x: (x[0][0], x[0][1]))
    total_pasif_kadrolu = 0
    total_pasif_gecici = 0
    for i, ((ust_birim, kisa_unvan), counts) in enumerate(pasif_sorted):
        ws_pano[f'F{row_idx_pasif}'] = ust_birim
        ws_pano[f'G{row_idx_pasif}'] = kisa_unvan
        ws_pano[f'H{row_idx_pasif}'] = counts['kadrolu']
        ws_pano[f'I{row_idx_pasif}'] = counts['gecici']
        total_pasif_kadrolu += counts['kadrolu']
        total_pasif_gecici += counts['gecici']
        if i % 2 == 1:
            for col in ['F', 'G', 'H', 'I']:
                ws_pano[f'{col}{row_idx_pasif}'].fill = alt_fill
        row_idx_pasif += 1
        
    ws_pano[f'G{row_idx_pasif}'] = "TOPLAM"
    ws_pano[f'G{row_idx_pasif}'].font = Font(bold=True)
    ws_pano[f'H{row_idx_pasif}'] = total_pasif_kadrolu
    ws_pano[f'H{row_idx_pasif}'].font = Font(bold=True)
    ws_pano[f'I{row_idx_pasif}'] = total_pasif_gecici
    ws_pano[f'I{row_idx_pasif}'].font = Font(bold=True)

    pasif_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    aktif_gecici_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def create_personel_sheet(wb, title, p_list, headers, include_son_birim=False):
        ws = wb.create_sheet(title=title)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            
        for row_num, dr in enumerate(p_list, 2):
            aktif_gorev = dr.aktif_gecicigorev
            gorev_birimi = aktif_gorev.gorevlendirildigi_birim if aktif_gorev else ""
            asil_kurum = aktif_gorev.asil_kurumu if aktif_gorev else ""
            bitis_tarihi = aktif_gorev.gecici_gorev_bitis if aktif_gorev else ""
            
            durum = dr.durum or ""
            is_pasif = "Pasif" in durum
            is_aktif_gecici = ("Aktif" in durum) and (dr.kadrolu_personel is False)
            
            if is_pasif:
                row_durum = "Pasif"
                asil_kurum = ""
            elif dr.kadrolu_personel is not False:
                row_durum = "Aktif (Kadrolu)"
            else:
                row_durum = "Aktif (Geçici)"
                gorev_birimi = ""

            row_data = [
                row_num - 1,
                dr.sicil_no,
                dr.tc_kimlik_no,
                dr.ad_soyad,
                dr.unvan.ad if dr.unvan else "",
                dr.brans.ad if dr.brans else ""
            ]
            
            if include_son_birim:
                row_data.append(dr.son_birim_kaydi or "")
                
            row_data.append(row_durum)

            row_data.extend([
                gorev_birimi,
                asil_kurum,
                bitis_tarihi.strftime("%d.%m.%Y") if bitis_tarihi else ""
            ])
            
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = val
                if is_pasif:
                    cell.fill = pasif_fill
                elif is_aktif_gecici:
                    cell.fill = aktif_gecici_fill
        return ws

    # Sheet 2: Doktorlar
    doktor_headers = [
        "Sıra No", "Sicil No", "T.C. Kimlik No", "Ad Soyad", "Unvan", "Branş", "Durum",
        "Görevlendirildiği Birim", "Asıl Kurumu", "Görevlendirme Bitiş Tarihi"
    ]
    ws_doktorlar = create_personel_sheet(wb, "Doktorlar", doktorlar_list, doktor_headers, include_son_birim=False)

    # Sheet 3: Memur 657
    genel_headers = [
        "Sıra No", "Sicil No", "T.C. Kimlik No", "Ad Soyad", "Unvan", "Branş", "Son Birim Kaydı", "Durum",
        "Görevlendirildiği Birim", "Asıl Kurumu", "Görevlendirme Bitiş Tarihi"
    ]
    ws_memur = create_personel_sheet(wb, "Memur 657", memur_657_list, genel_headers, include_son_birim=True)

    # Sheet 4: Sürekli İşçi
    ws_isci = create_personel_sheet(wb, "Sürekli İşçi", surekli_isci_list, genel_headers, include_son_birim=True)

    for ws in [ws_pano, ws_doktorlar, ws_memur, ws_isci]:
        for i, column_cells in enumerate(ws.columns, 1):
            if not column_cells: continue
            length = max([len(str(cell.value) or "") for cell in column_cells] + [0])
            ws.column_dimensions[get_column_letter(i)].width = length + 2

    # Dosya adı: Personel_Verileri_Bugünün tarihi
    bugun = datetime.now().strftime("%d.%m.%Y")
    dosya_adi = f"Personel_Verileri_{bugun}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={dosya_adi}'
    wb.save(response)
    return response


@login_required
def personel_kontrol(request):
    """AJAX: T.C. Kimlik No ile sistemde kayıtlı personel var mı?"""
    tc = request.GET.get('tc_kimlik_no', '').strip()
    exists = Personel.objects.filter(tc_kimlik_no=tc).exists()
    return JsonResponse({'exists': exists})

@login_required
def get_brans_by_unvan(request):
    """AJAX: Unvan ID'sine göre branşları getir"""
    unvan_id = request.GET.get('unvan_id')
    if unvan_id:
        branslar = Brans.objects.filter(unvan_id=unvan_id).values('id', 'ad')
        return JsonResponse({'branslar': list(branslar)})
    return JsonResponse({'branslar': []})

def personel_ekle(request):
    if request.method == 'POST':
        form = PersonelForm(request.POST)
        if form.is_valid():
            personel = form.save(commit=False)
            kadrolu = form.cleaned_data.get('kadrolu_personel')
            # Geçici görev alanlarını POST'tan al
            gecici_gorev_baslangic = request.POST.get('gecici_gorev_baslangic')
            gecici_gorev_bitis = request.POST.get('gecici_gorev_bitis')
            gorevlendirildigi_birim = form.cleaned_data.get('fiili_gorev_yeri').ad if form.cleaned_data.get('fiili_gorev_yeri') else ''
            asil_kurumu = form.cleaned_data.get('kadro_yeri').ad if form.cleaned_data.get('kadro_yeri') else ''
            # Kişi kaydını oluştur
            personel.save()
            form.save_m2m()
            # Eğer kadrolu değilse otomatik geçici görev kaydı oluştur
            if not kadrolu and gecici_gorev_baslangic and gorevlendirildigi_birim:
                GeciciGorev.objects.create(
                    gecici_gorev_tipi='Gelis',
                    personel=personel,
                    gecici_gorev_baslangic=gecici_gorev_baslangic,
                    gecici_gorev_bitis=gecici_gorev_bitis or None,
                    gorevlendirildigi_birim=gorevlendirildigi_birim,
                    asil_kurumu=asil_kurumu
                )
            return redirect('ik_core:personel_list')
    else:
        form = PersonelForm()
    return render(request, 'ik_core/personel_ekle.html', {'form': form})

def personel_detay(request, pk):
    personel = get_object_or_404(Personel.objects.select_related('unvan', 'brans', 'kurum', 'kadro_yeri', 'fiili_gorev_yeri').prefetch_related('ozel_durumu', 'gecicigorev_set'), pk=pk)
    unvanlar = Unvan.objects.all()
    branslar = Brans.objects.all()
    kurumlar = Kurum.objects.all()
    ozel_durum_all = OzelDurum.objects.all()
    ozel_durumu_ids = list(personel.ozel_durumu.values_list('id', flat=True))
    
    # Personelin en son çalıştığı birim bilgisini al
    son_birim_kaydi = PersonelBirim.objects.filter(personel=personel).order_by('-gecis_tarihi', '-creation_timestamp').first()
    
    # Personelin birim geçmişi
    personel_birim_gecmisi = PersonelBirim.objects.filter(personel=personel).order_by('-gecis_tarihi', '-creation_timestamp')
    
    # Modal'lar için gerekli veriler
    ust_birimler = UstBirim.objects.all().order_by('ad')
    binalar = Bina.objects.all().order_by('ad')
    teblig_imzalari = TebligImzasi.objects.all().order_by('ad')

    # Use constants from valuelists
    teskilat_choices = TESKILAT_DEGERLERI
    mazeret_durumu_choices = MAZERET_DEGERLERI
    tahsil_durumu_choices = EGITIM_DEGERLERI
    cinsiyet_choices = Personel._meta.get_field('cinsiyet').choices
    ayrilma_nedeni_choices = AYRILMA_NEDENI_DEGERLERI
    vergi_indirimi_choices = ENGEL_DERECESI_DEGERLERI

    gecici_gorev_form = GeciciGorevForm(initial={
        'asil_kurumu': personel.kurum.ad if personel.kurum else ''
    })

    # PersonelForm instance'ını oluştur
    form = PersonelForm(instance=personel)

    if request.method == 'POST':
        if 'save_main' in request.POST:
            # POST verilerini işle ve eksik alanları doldur
            post_data = request.POST.copy()
            
            # Eğer hidden input'lardan gelen değerler varsa, onları kullan
            # Aksi halde mevcut personel değerlerini kullan
            if not post_data.get('tc_kimlik_no'):
                post_data['tc_kimlik_no'] = personel.tc_kimlik_no
            if not post_data.get('ad'):
                post_data['ad'] = personel.ad
            if not post_data.get('soyad'):
                post_data['soyad'] = personel.soyad
            if not post_data.get('kurum'):
                post_data['kurum'] = personel.kurum.id if personel.kurum else ''
            
            form = PersonelForm(post_data, instance=personel)
            if form.is_valid():
                form.save()
                messages.success(request, f"{personel.ad} {personel.soyad} bilgileri başarıyla güncellendi.")
                return redirect('ik_core:personel_detay', pk=pk)
            else:
                # Form hatalarını detaylı olarak logla
                print("=== FORM HATALARI DEBUG ===")
                print(f"POST verileri: {request.POST}")
                print("İşlenmiş POST verileri: {post_data}")
                print("Form hataları:")
                for field, errors in form.errors.items():
                    print(f"  {field}: {errors}")
                
                # Form alanlarının değerlerini kontrol et
                print("Form alanları değerleri:")
                for field_name, field in form.fields.items():
                    value = form.data.get(field_name, 'BOŞ')
                    print(f"  {field_name}: {value}")
                
                # Hata mesajlarını kullanıcıya göster
                error_messages = []
                for field, errors in form.errors.items():
                    field_name = form.fields[field].label if hasattr(form.fields[field], 'label') else field
                    for error in errors:
                        error_messages.append(f"{field_name}: {error}")
                
                if error_messages:
                    messages.error(request, f"Formda hatalar var: {'; '.join(error_messages)}")
                else:
                    messages.error(request, "Formda hatalar var. Lütfen kontrol ediniz.")
        elif 'save_ayrilis' in request.POST:
            # Ayrılış bilgilerini güncelle
            personel.ayrilma_tarihi = request.POST.get('ayrilma_tarihi') or None
            personel.ayrilma_nedeni = request.POST.get('ayrilma_nedeni') or ''
            personel.ayrilma_detay = request.POST.get('ayrilma_detay') or ''
            try:
                personel.save(update_fields=['ayrilma_tarihi', 'ayrilma_nedeni', 'ayrilma_detay'])
                messages.success(request, f"{personel.ad} {personel.soyad} için ayrılış bilgileri kaydedildi.")
            except Exception as e:
                messages.error(request, f"Ayrılış bilgileri kaydedilirken hata oluştu: {e}")
            return redirect('ik_core:personel_detay', pk=pk)

    personel_ismi_degistirme_yetkisi = request.user.has_permission('İK Modülü Personel İsmi Değiştirme')

    context = {
        'personel': personel,
        'form': form,  # PersonelForm instance'ını context'e ekle
        'unvanlar': unvanlar,
        'branslar': branslar,
        'kurumlar': kurumlar,
        'gecici_gorev_form': gecici_gorev_form,
        'teskilat_choices': teskilat_choices,
        'mazeret_durumu_choices': mazeret_durumu_choices,
        'tahsil_durumu_choices': tahsil_durumu_choices,
        'cinsiyet_choices': cinsiyet_choices,
        'ozel_durumu_choices': ozel_durum_all,
        'ozel_durumu_ids': ozel_durumu_ids,
        'ayrilma_nedeni_choices': ayrilma_nedeni_choices,
        'vergi_indirimi_choices': vergi_indirimi_choices,
        'son_birim_kaydi': son_birim_kaydi,
        'personel_birim_gecmisi': personel_birim_gecmisi,
        'ust_birimler': ust_birimler,
        'binalar': binalar,
        'teblig_imzalari': teblig_imzalari,
        'personel_ismi_degistirme_yetkisi': personel_ismi_degistirme_yetkisi,
    }
    return render(request, 'ik_core/personel_detay.html', context)

@require_POST
@login_required
def personel_isim_degistir(request, personel_id):
    if not request.user.has_permission('İK Modülü Personel İsmi Değiştirme'):
        messages.error(request, "Bu işlem için yetkiniz bulunmamaktadır.")
        return redirect('ik_core:personel_detay', pk=personel_id)
    
    personel = get_object_or_404(Personel, pk=personel_id)
    yeni_ad = request.POST.get('ad', '').strip()
    yeni_soyad = request.POST.get('soyad', '').strip()
    
    if not yeni_ad or not yeni_soyad:
        messages.error(request, "Ad ve soyad alanları boş bırakılamaz.")
        return redirect('ik_core:personel_detay', pk=personel_id)
    
    eski_ad = personel.ad
    eski_soyad = personel.soyad
    
    personel.ad = yeni_ad
    personel.soyad = yeni_soyad
    personel.save(update_fields=['ad', 'soyad'])
    
    messages.success(request, f"Personel ismi '{eski_ad} {eski_soyad}' den '{yeni_ad} {yeni_soyad}' olarak güncellendi.")
    return redirect('ik_core:personel_detay', pk=personel_id)

@require_POST
def gecici_gorev_ekle(request, personel_id):
    personel = get_object_or_404(Personel, pk=personel_id)
    form = GeciciGorevForm(request.POST)
    if form.is_valid():
        gecici_gorev = form.save(commit=False)
        gecici_gorev.personel = personel
        gecici_gorev.save()
        messages.success(request, f"{personel.ad} {personel.soyad} için geçici görev kaydı başarıyla eklendi.")
    return JsonResponse({
        'success': True,
    })

@require_POST
def gecici_gorev_sil(request, personel_id, gorev_id):
    personel = get_object_or_404(Personel, pk=personel_id)
    gorev = get_object_or_404(personel.gecicigorev_set, pk=gorev_id)
    gorev.delete()
    return redirect(reverse('ik_core:personel_detay', args=[personel_id]))

def personel_duzenle(request, pk):
    personel = get_object_or_404(Personel, pk=pk)
    if request.method == 'POST':
        form = PersonelForm(request.POST, instance=personel)
        if form.is_valid():
            form.save()
            return redirect('ik_core:personel_detay', pk=pk)
    else:
        form = PersonelForm(instance=personel)
    return render(request, 'ik_core/personel_duzenle.html', {'form': form, 'personel': personel})

def kurum_tanimlari(request):
    kurumlar = Kurum.objects.all()
    if request.method == 'POST':
        form = KurumForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ik_core:kurum_tanimlari')
    else:
        form = KurumForm()
    return render(request, 'ik_core/kurum_tanimlari.html', {'kurumlar': kurumlar, 'form': form})

from django.db.models import Count
from .models.personel import KisaUnvan, UnvanBransEslestirme

def unvan_branstanimlari(request):
    unvanlar = Unvan.objects.annotate(brans_sayisi=Count('brans')).prefetch_related('unvanbranseslestirme_set__kisa_unvan__ust_birim')
    
    # Unvanlar için genel eşleştirmeyi (branşsız) bulup objeye ekleyelim
    for u in unvanlar:
        # DB'den gelen seti Python tarafında filtreliyoruz
        u.genel_eslestirme = next((e for e in u.unvanbranseslestirme_set.all() if e.brans is None), None)

    selected_unvan_id = request.GET.get('unvan_id')
    selected_unvan = None
    branslar = Brans.objects.none()
    
    if selected_unvan_id:
        selected_unvan = get_object_or_404(Unvan, id=selected_unvan_id)
        branslar = Brans.objects.filter(unvan=selected_unvan).select_related('unvan').prefetch_related('unvanbranseslestirme_set__kisa_unvan__ust_birim')
        for b in branslar:
            b.eslestirme = next((e for e in b.unvanbranseslestirme_set.all()), None)

    if request.method == 'POST':
        if 'unvan_ekle' in request.POST:
            unvan_form = UnvanForm(request.POST, prefix='unvan')
            if unvan_form.is_valid():
                unvan_form.save()
                messages.success(request, f"{unvan_form.cleaned_data['ad']} unvanı tanımlandı")
                return redirect('ik_core:unvan_branstanimlari')
        elif 'brans_ekle' in request.POST:
            brans_form = BransForm(request.POST, prefix='brans')
            if brans_form.is_valid():
                brans_form.save()
                messages.success(request, f"{brans_form.cleaned_data['ad']} branşı tanımlandı")
                return redirect(f"{request.path}?unvan_id={brans_form.cleaned_data['unvan'].id}")
    else:
        unvan_form = UnvanForm(prefix='unvan')
        brans_form = BransForm(prefix='brans')
        
    # Tüm KisaUnvan ve UstBirim listeleri (datalist ve dropdown için)
    tum_kisa_unvanlar = KisaUnvan.objects.select_related('ust_birim').all()
    tum_ust_birimler = UstBirim.objects.all().order_by('ad')

    return render(request, 'ik_core/unvan_branstanimlari.html', {
        'unvanlar': unvanlar,
        'selected_unvan': selected_unvan,
        'branslar': branslar,
        'unvan_form': unvan_form,
        'brans_form': brans_form,
        'tum_kisa_unvanlar': tum_kisa_unvanlar,
        'tum_ust_birimler': tum_ust_birimler,
    })

@login_required
@require_POST
def unvan_eslestirme_kaydet(request):
    """
    AJAX: Unvan veya Branş için KisaUnvan ve UstBirim eşleştirmesini kaydeder.
    """
    try:
        import json
        data = json.loads(request.body)
        
        unvan_id = data.get('unvan_id')
        brans_id = data.get('brans_id')
        kisa_unvan_ad = data.get('kisa_unvan_ad', '').strip()
        ust_birim_id = data.get('ust_birim_id')
        
        if not unvan_id or not kisa_unvan_ad:
            return JsonResponse({'success': False, 'message': 'Eksik bilgi.'})

        # 1. KisaUnvan Bul veya Oluştur
        kisa_unvan, created = KisaUnvan.objects.get_or_create(ad=kisa_unvan_ad)
        
        # Eğer yeni oluşturulduysa veya ust_birim güncellenmek (istenirse burası opsiyonel yapılabilir)
        # Kullanıcının seçimi varsa güncelleyelim.
        if ust_birim_id:
            ust_birim = get_object_or_404(UstBirim, pk=ust_birim_id)
            if kisa_unvan.ust_birim != ust_birim:
                kisa_unvan.ust_birim = ust_birim
                kisa_unvan.save()
        
        # 2. Eşleştirmeyi Yap
        unvan = get_object_or_404(Unvan, pk=unvan_id)
        brans = get_object_or_404(Brans, pk=brans_id) if brans_id else None
        
        eslestirme, _ = UnvanBransEslestirme.objects.update_or_create(
            unvan=unvan,
            brans=brans,
            defaults={'kisa_unvan': kisa_unvan}
        )
        
        return JsonResponse({
            'success': True,
            'kisa_unvan_ad': kisa_unvan.ad,
            'ust_birim_id': kisa_unvan.ust_birim.id if kisa_unvan.ust_birim else None,
            'ust_birim_ad': kisa_unvan.ust_birim.ad if kisa_unvan.ust_birim else ''
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def tanimlamalar(request):

    return render(request, 'ik_core/tanimlamalar.html')

# =====================
# Tebliğ: Modeller için basit formlar
# =====================

class TebligImzasiForm(ModelForm):
    class Meta:
        model = TebligImzasi
        fields = ['ad', 'metin']


class TebligMetniForm(ModelForm):
    class Meta:
        model = TebligMetni
        fields = ['ad', 'metin']


# =====================
# Tebliğ: Tanım sayfaları ve CRUD
# =====================

@login_required
def imza_tanimlari(request):
    imzalar = TebligImzasi.objects.order_by('ad')
    form = TebligImzasiForm()
    return render(request, 'ik_core/imza_tanimlari.html', {
        'imzalar': imzalar,
        'form': form,
    })


@login_required
def teblig_tanimlari(request):
    tebligler = TebligMetni.objects.order_by('ad')
    form = TebligMetniForm()
    return render(request, 'ik_core/teblig_tanimlari.html', {
        'tebligler': tebligler,
        'form': form,
    })

@login_required
@require_POST
def teblig_imzasi_ekle(request):
    pk = request.POST.get('id')
    if pk:
        instance = get_object_or_404(TebligImzasi, pk=pk)
        form = TebligImzasiForm(request.POST, instance=instance)
    else:
        form = TebligImzasiForm(request.POST)
    if form.is_valid():
        imza = form.save()
        return JsonResponse({'success': True, 'id': imza.id, 'ad': imza.ad, 'metin': imza.metin})
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
@require_POST
def teblig_imzasi_sil(request, pk):
    imza = get_object_or_404(TebligImzasi, pk=pk)
    imza.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def teblig_metni_ekle(request):
    pk = request.POST.get('id')
    if pk:
        instance = get_object_or_404(TebligMetni, pk=pk)
        form = TebligMetniForm(request.POST, instance=instance)
    else:
        form = TebligMetniForm(request.POST)
    if form.is_valid():
        tm = form.save()
        return JsonResponse({'success': True, 'id': tm.id, 'ad': tm.ad, 'metin': tm.metin})
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
@require_POST
def teblig_metni_sil(request, pk):
    tm = get_object_or_404(TebligMetni, pk=pk)
    tm.delete()
    return JsonResponse({'success': True})


# =====================
# Tebliğ: İşlemleri Sayfası
# =====================

@login_required
def teblig_islemleri(request, personel_id: int):
    personel = get_object_or_404(Personel, pk=personel_id)
    imzalar = TebligImzasi.objects.order_by('ad')
    tebligler = TebligMetni.objects.order_by('ad')
    return render(request, 'ik_core/teblig_islemleri.html', {
        'personel': personel,
        'imzalar': imzalar,
        'tebligler': tebligler,
    })


@login_required
@require_POST
def teblig_metni_guncelle(request, pk: int):
    tm = get_object_or_404(TebligMetni, pk=pk)
    metin = request.POST.get('metin', '')
    tm.metin = metin
    tm.save(update_fields=['metin'])
    return JsonResponse({'success': True})


@login_required
@require_GET
def teblig_metni_get(request, pk: int):
    tm = get_object_or_404(TebligMetni, pk=pk)
    return JsonResponse({'id': tm.id, 'ad': tm.ad, 'metin': tm.metin})

# Placeholder view for Ilisik Kesme Formu
def ilisik_kesme_formu(request, pk):
    # Prepare PDF-specific context using the supplied PDF template (ilisik_kesme_formu.html)
    try:
        file_url = f"file:///{staticfiles_storage.path('logo/kdh_logo.png')}"
    except Exception:
        file_url = None
    
    personel = get_object_or_404(Personel, pk=pk)
    
    # prepare context matching the PDF template
    context_pdf = {
        'dokuman_kodu': 'KU.FR.31',
        'form_adi': 'İlişik Kesme Formu',
        'yayin_tarihi': 'Mayıs 2023',
        'kurum': 'KAYSERİ DEVLET HASTANESİ',
        'revizyon_tarihi': '-',
        'revizyon_no': '0',
        'sayfa_no': '1',
        'pdf_logo': file_url,
        'personel': personel,
    }

    # Render PDF template and generate PDF (landscape)
    try:
        template = get_template('ik_core/pdf/ilisik_kesme_formu.html')
        html = template.render({**context_pdf})
    except Exception:
        html = render_to_string('ik_core/pdf/ilisik_kesme_formu.html', context_pdf)

    config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
    options = {
        'page-size': 'A4',
        'orientation': 'Portrait',
        'margin-top': '1.5cm',
        'margin-right': '1.5cm',
        'margin-bottom': '1.1cm',
        'margin-left': '1.5cm',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': '',
        'enable-external-links': True,
        'quiet': ''
    }

    pdf = pdfkit.from_string(html, False, options=options, configuration=config)
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"ilisik_kesme_formu_{personel.ad_soyad}.pdf"
    filename = filename.replace(' ', '_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

@require_http_methods(["POST"])
def ayrilis_kaydet(request, pk):
    """Ayrılış bilgilerini kaydet"""
    if request.method == 'POST':
        try:
            personel = get_object_or_404(Personel, pk=pk)
            
            # Ayrılış bilgilerini al
            ayrilma_tarihi = request.POST.get('ayrilma_tarihi')
            ayrilma_nedeni = request.POST.get('ayrilma_nedeni')
            ayrilma_detay = request.POST.get('ayrilma_detay', '')
            
            # Validasyon
            if not ayrilma_tarihi:
                return JsonResponse({
                    'success': False,
                    'message': 'Ayrılma tarihi zorunludur.'
                })
            
            if not ayrilma_nedeni:
                return JsonResponse({
                    'success': False,
                    'message': 'Ayrılma nedeni seçilmelidir.'
                })
            
            # Personel bilgilerini güncelle
            personel.ayrilma_tarihi = ayrilma_tarihi
            personel.ayrilma_nedeni = ayrilma_nedeni
            personel.ayrilma_detay = ayrilma_detay
            personel.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Ayrılış bilgileri başarıyla kaydedildi.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Hata oluştu: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Geçersiz istek.'
    })

@require_http_methods(["POST"])
def personeli_aktiflestir(request, pk):
    """Personeli aktifleştir - ayrılış bilgilerini temizle"""
    if request.method == 'POST':
        try:
            personel = get_object_or_404(Personel, pk=pk)
            
            # Ayrılış bilgilerini temizle
            personel.ayrilma_tarihi = None
            personel.ayrilma_nedeni = None
            personel.ayrilma_detay = None
            personel.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Personel başarıyla aktifleştirildi.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Hata oluştu: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Geçersiz istek.'
    })

@require_http_methods(["POST"])
def mazeret_sil(request, pk):
    """İlgili personel için mazeret kaydını sil"""
    if request.method == 'POST':
        try:
            personel = get_object_or_404(Personel, pk=pk)
            personel.mazeret_durumu = None
            personel.mazeret_baslangic = None
            personel.mazeret_bitis = None
            personel.save()
            return JsonResponse({
                'success': True,
                'message': 'Mazeret kaydı silindi.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Hata oluştu: {str(e)}'
            })
    return JsonResponse({
        'success': False,
        'message': 'Geçersiz istek.'
    })    

# =====================
# Geçici Görevler Sayfası
# =====================

@login_required
def gecici_gorevler(request):
    """
    RFC-001-GeciciGorevSayfasi gereksinimlerine uygun liste ve modal sayfası
    - Tarih filtresi ile o günü kapsayan kayıtlar
    - Sağ üstte modal açan buton
    """
    tarih_str = request.GET.get('tarih')
    kayitlar = []
    from .models.GeciciGorev import GeciciGorev
    if tarih_str:
        tarih = parse_date(tarih_str)
        if tarih:
            kayitlar = (GeciciGorev.objects
                        .filter(gecici_gorev_baslangic__lte=tarih)
                        .filter(models.Q(gecici_gorev_bitis__isnull=True) | models.Q(gecici_gorev_bitis__gte=tarih))
                        .select_related('personel'))
    kurumlar = Kurum.objects.order_by('ad').all()
    return render(request, 'ik_core/gecici_gorevler.html', {
        'kayitlar': kayitlar,
        'tarih': tarih_str or '',
        'kurumlar': kurumlar,
    })


@login_required
@require_POST
def gecici_gorev_bulk_kaydet(request):
    """
    Excel'den yapıştırılan satırlar için toplu kayıt oluşturma.
    Beklenen JSON body:
    {
      "kurum": "KAYSERİ DEVLET HASTANESİ",
      "satirlar": [
        {
          "tc": "123...",
          "sicil": "...",
          "ad": "...",
          "soyad": "...",
          "unvan": "...",
          "brans": "...",
          "kadro_birim": "...",
          "aktif_birim": "...",
          "baslangic": "YYYY-MM-DD",
          "bitis": "YYYY-MM-DD|null"
        }
      ]
    }
    Tip belirleme:
      - kurum == kadro_birim -> Gidis
      - kurum == aktif_birim -> Gelis
    """
    import json
    import re
    from datetime import datetime
    from .models.GeciciGorev import GeciciGorev

    def normalize_text(value: str) -> str:
        if value is None:
            return ''
        v = str(value).strip()
        # Baş/son tırnakları kaldır
        if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
            v = v[1:-1]
        # CR/LF -> boşluk, birden fazla boşluğu tek boşluk yap
        v = re.sub(r"\s+", " ", v.replace("\r", " ").replace("\n", " ")).strip()
        return v

    def upper_norm(value: str) -> str:
        return normalize_text(value).upper()

    def parse_any_date(value: str):
        if not value:
            return None
        s = normalize_text(value)
        # Doğrudan ISO (yyyy-mm-dd)
        d = parse_date(s)
        if d:
            return d
        # dd.mm.yyyy | d.mm.yyyy | dd/mm/yyyy | dd-mm-yyyy
        s2 = s.replace('/', '.').replace('-', '.')
        parts = s2.split('.')
        if len(parts) == 3 and parts[2].isdigit():
            try:
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                return datetime.strptime(f"{day}.{month}.{year}", "%d.%m.%Y").date()
            except Exception:
                pass
        # yyyy.mm.dd
        if len(parts) == 3 and parts[0].isdigit() and len(parts[0]) == 4:
            try:
                year = parts[0]
                month = parts[1].zfill(2)
                day = parts[2].zfill(2)
                return datetime.strptime(f"{year}.{month}.{day}", "%Y.%m.%d").date()
            except Exception:
                pass
        return None
    body = json.loads(request.body or '{}')
    kurum_adi = body.get('kurum')
    satirlar = body.get('satirlar', [])

    sayac_gelis = 0
    sayac_gidis = 0
    sayac_atlanan = 0

    for s in satirlar:
        tc = normalize_text(s.get('tc') or '')
        kadro_birim = normalize_text(s.get('kadro_birim') or '')
        aktif_birim = normalize_text(s.get('aktif_birim') or '')
        baslangic = parse_any_date(s.get('baslangic') or '')
        bitis = parse_any_date(s.get('bitis') or '') if s.get('bitis') else None
        if not tc or not baslangic:
            sayac_atlanan += 1
            continue
        try:
            personel = Personel.objects.get(tc_kimlik_no=tc)
        except Personel.DoesNotExist:
            sayac_atlanan += 1
            continue

        gorev_tipi = None
        if kurum_adi and upper_norm(kurum_adi) == upper_norm(kadro_birim):
            gorev_tipi = 'Gidis'
        if kurum_adi and upper_norm(kurum_adi) == upper_norm(aktif_birim):
            # aktif eşleşme varsa geliş öncelik kazansın
            gorev_tipi = 'Gelis'
        if gorev_tipi is None:
            sayac_atlanan += 1
            continue

        GeciciGorev.objects.create(
            personel=personel,
            gecici_gorev_tipi=gorev_tipi,
            gecici_gorev_baslangic=baslangic,
            gecici_gorev_bitis=bitis,
            asil_kurumu=kadro_birim,
            gorevlendirildigi_birim=aktif_birim,
        )
        if gorev_tipi == 'Gelis':
            sayac_gelis += 1
        else:
            sayac_gidis += 1

    return JsonResponse({
        'success': True,
        'message': (
            f"Geçici Görev kayıtları tamamlandı: {sayac_gelis} geliş, {sayac_gidis} gidiş. "
            f"{sayac_atlanan} satır atlandı."
        )
    })


@login_required
@require_POST
def gecici_gorev_bulk_update_mevcut(request):
    """
    Mevcut GeciciGorev kayıtlarını toplu (inline) günceller.
    """
    import json
    from datetime import datetime
    try:
        from .models.GeciciGorev import GeciciGorev
    except ImportError:
        from ik_core.models.GeciciGorev import GeciciGorev

    try:
        body = json.loads(request.body)
        updates = body.get('updates', [])
        
        updated_count = 0
        errors = []

        for item in updates:
            gorev_id = item.get('id')
            if not gorev_id:
                continue

            try:
                gorev = GeciciGorev.objects.get(pk=gorev_id)
            except GeciciGorev.DoesNotExist:
                errors.append(f"ID {gorev_id} bulunamadı.")
                continue

            # Update fields
            if 'asil_kurumu' in item:
                gorev.asil_kurumu = item['asil_kurumu'].strip()
            
            if 'gorevlendirildigi_birim' in item:
                gorev.gorevlendirildigi_birim = item['gorevlendirildigi_birim'].strip()

            if 'gecici_gorev_baslangic' in item:
                baslangic_str = item['gecici_gorev_baslangic']
                if baslangic_str:
                    try:
                        gorev.gecici_gorev_baslangic = datetime.strptime(baslangic_str, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"ID {gorev_id}: Geçersiz başlangıç tarihi ({baslangic_str})")

            if 'gecici_gorev_bitis' in item:
                bitis_str = item['gecici_gorev_bitis']
                if bitis_str:
                    try:
                        gorev.gecici_gorev_bitis = datetime.strptime(bitis_str, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"ID {gorev_id}: Geçersiz bitiş tarihi ({bitis_str})")
                else:
                    gorev.gecici_gorev_bitis = None
            
            if 'gecici_gorev_tipi' in item:
                tip = item['gecici_gorev_tipi']
                if tip in ['Gelis', 'Gidis']:
                    gorev.gecici_gorev_tipi = tip

            gorev.save()
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} kayıt güncellendi.',
            'errors': errors
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Geçersiz JSON formatı.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata: {str(e)}'}, status=500)

@login_required
def birim_yonetimi(request):
    """Birim yönetimi sayfası - unvan_branstanimlari.html benzeri yapı"""
    binalar = Bina.objects.all().order_by('ad')
    ust_birimler = UstBirim.objects.all().order_by('ad')
    selected_bina_id = request.GET.get('bina_id')
    selected_bina = None
    birimler = []
    
    if selected_bina_id:
        selected_bina = get_object_or_404(Bina, id=selected_bina_id)
        birimler = Birim.objects.filter(bina=selected_bina).order_by('ust_birim__ad', 'ad')
    
    context = {
        'binalar': binalar,
        'ust_birimler': ust_birimler,
        'selected_bina': selected_bina,
        'birimler': birimler,
        'kampusler': Kampus.objects.all().order_by('ad'), # Kampüsleri context'e ekledik
    }
    return render(request, 'ik_core/birim_yonetimi.html', context)

@login_required
@require_POST
def bina_ekle_duzenle(request):
    """Bina ekleme ve düzenleme endpoint'i"""
    bina_id = request.POST.get('bina_id')
    ad = request.POST.get('ad', '').strip()
    kampus_id = request.POST.get('kampus_id')
    aciklama = request.POST.get('aciklama', '').strip()

    if not ad:
        return JsonResponse({'success': False, 'message': 'Bina adı gereklidir.'})

    try:
        if bina_id: # Düzenleme
            bina = get_object_or_404(Bina, id=bina_id)
            if Bina.objects.filter(ad=ad).exclude(id=bina_id).exists():
                return JsonResponse({'success': False, 'message': 'Bu bina adı başka bir kayıt tarafından kullanılıyor.'})
            bina.ad = ad
            message = 'Bina başarıyla güncellendi.'
        else: # Ekleme
            if Bina.objects.filter(ad=ad).exists():
                return JsonResponse({'success': False, 'message': 'Bu bina adı zaten mevcut.'})
            bina = Bina(ad=ad)
            message = 'Bina başarıyla eklendi.'
        
        # Kampüs atama
        if kampus_id:
            bina.kampus_id = kampus_id
        else:
            bina.kampus = None
            
        bina.aciklama = aciklama
        bina.save()

        return JsonResponse({
            'success': True, 
            'message': message,
            'bina': {
                'id': bina.id, 
                'ad': bina.ad, 
                'kampus_id': bina.kampus_id,
                'aciklama': bina.aciklama,
                'birim_sayisi': bina.birim_sayisi
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata oluştu: {str(e)}'})

@login_required
@require_GET
def get_bina_detay(request):
    """AJAX ile bina detaylarını getir"""
    bina_id = request.GET.get('bina_id')
    if not bina_id:
        return JsonResponse({'success': False})
    
    bina = get_object_or_404(Bina, id=bina_id)
    return JsonResponse({
        'success': True,
        'bina': {
            'id': bina.id,
            'ad': bina.ad,
            'kampus_id': bina.kampus_id,
            'aciklama': bina.aciklama or ''
        }
    })

@login_required
@require_POST
def ust_birim_ekle(request):
    """Üst birim ekleme endpoint'i"""
    ad = request.POST.get('ad', '').strip()
    if not ad:
        return JsonResponse({'success': False, 'message': 'Üst birim adı gereklidir.'})
    
    if UstBirim.objects.filter(ad=ad).exists():
        return JsonResponse({'success': False, 'message': 'Bu üst birim adı zaten mevcut.'})
    
    ust_birim = UstBirim.objects.create(ad=ad)
    return JsonResponse({
        'success': True, 
        'message': 'Üst birim başarıyla eklendi.',
        'ust_birim': {'id': ust_birim.id, 'ad': ust_birim.ad}
    })

@login_required
@require_POST
@login_required
@require_POST
def birim_ekle(request):
    """Birim ekleme ve güncelleme endpoint'i"""
    birim_id = request.POST.get('birim_id')
    bina_id = request.POST.get('bina_id')
    ust_birim_id = request.POST.get('ust_birim_id')
    ad = request.POST.get('ad', '').strip()
    aciklama = request.POST.get('aciklama', '').strip()
    
    if not all([bina_id, ust_birim_id, ad]):
        return JsonResponse({'success': False, 'message': 'Tüm zorunlu alanlar gereklidir.'})
    
    try:
        bina = Bina.objects.get(id=bina_id)
        ust_birim = UstBirim.objects.get(id=ust_birim_id)
    except (Bina.DoesNotExist, UstBirim.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Geçersiz bina veya üst birim.'})
    
    if birim_id:
        # Güncelleme
        try:
            birim = Birim.objects.get(id=birim_id)
            # İsim çakışması kontrolü (kendisi hariç)
            if Birim.objects.filter(bina=bina, ad=ad).exclude(id=birim_id).exists():
                return JsonResponse({'success': False, 'message': 'Bu birim adı bu binada zaten mevcut.'})
            
            birim.bina = bina
            birim.ust_birim = ust_birim
            birim.ad = ad
            birim.aciklama = aciklama
            birim.save()
            message = 'Birim başarıyla güncellendi.'
        except Birim.DoesNotExist:
             return JsonResponse({'success': False, 'message': 'Güncellenecek birim bulunamadı.'})
    else:
        # Ekleme
        if Birim.objects.filter(bina=bina, ad=ad).exists():
            return JsonResponse({'success': False, 'message': 'Bu birim adı bu binada zaten mevcut.'})
        
        birim = Birim.objects.create(bina=bina, ust_birim=ust_birim, ad=ad, aciklama=aciklama)
        message = 'Birim başarıyla eklendi.'

    return JsonResponse({
        'success': True, 
        'message': message,
        'birim': {
            'id': birim.id, 
            'ad': birim.ad, 
            'bina_ad': birim.bina.ad,
            'ust_birim_ad': birim.ust_birim.ad,
            'aciklama': birim.aciklama
        }
    })

@login_required
@require_POST
def birim_sil(request, pk):
    """Birim kaydını sil"""
    try:
        from .models.BirimYonetimi import Birim
    except ImportError:
        from ik_core.models.BirimYonetimi import Birim
        
    try:
        birim = Birim.objects.get(pk=pk)
        # Check if there are personnel assigned to this unit
        if birim.personelbirim_set.exists():
             return JsonResponse({'success': False, 'message': 'Bu birime bağlı personel geçmişi kayıtları var. Önce onları silmelisiniz.'})
        
        birim.delete()
        return JsonResponse({'success': True, 'message': 'Birim başarıyla silindi.'})
    except Birim.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Birim bulunamadı.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata oluştu: {str(e)}'})

@login_required
@require_POST
def bina_sil(request, pk):
    """Bina kaydını sil"""
    try:
        from .models.BirimYonetimi import Bina
    except ImportError:
        from ik_core.models.BirimYonetimi import Bina
        
    try:
        bina = Bina.objects.get(pk=pk)
        # Check if there are units in this building
        if bina.birim_set.exists():
            return JsonResponse({'success': False, 'message': 'Bu binaya bağlı birimler var. Önce birimleri silmelisiniz.'})
            
        bina.delete()
        return JsonResponse({'success': True, 'message': 'Bina başarıyla silindi.'})
    except Bina.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Bina bulunamadı.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata oluştu: {str(e)}'})

@login_required
def get_birimler_by_bina(request):
    """Bina seçimine göre birimleri getiren AJAX endpoint'i"""
    bina_id = request.GET.get('bina_id')
    ust_birim_id = request.GET.get('ust_birim_id')
    
    if not bina_id:
        return JsonResponse({'birimler': []})
    
    birimler = Birim.objects.filter(bina_id=bina_id)
    if ust_birim_id:
        birimler = birimler.filter(ust_birim_id=ust_birim_id)
    
    birimler = birimler.order_by('ust_birim__ad', 'ad')
    
    return JsonResponse({
        'birimler': [
            {
                'id': birim.id, 
                'ad': birim.ad, 
                'ust_birim_ad': birim.ust_birim.ad, 
                'ust_birim_id': birim.ust_birim.id,
                'aciklama': birim.aciklama or ''
            }
            for birim in birimler
        ]
    })

@login_required
@require_POST
def personel_birim_ekle(request):
    """Personel birim ekleme endpoint'i"""
    personel_id = request.POST.get('personel_id')
    birim_id = request.POST.get('birim_id')
    gecis_tarihi = request.POST.get('gecis_tarihi')
    sorumlu = request.POST.get('sorumlu') == 'on'
    not_text = request.POST.get('not', '').strip()
    
    if not all([personel_id, birim_id, gecis_tarihi]):
        return JsonResponse({'success': False, 'message': 'Tüm zorunlu alanlar doldurulmalıdır.'})
    
    try:
        personel = Personel.objects.get(id=personel_id)
        birim = Birim.objects.get(id=birim_id)
    except (Personel.DoesNotExist, Birim.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Geçersiz personel veya birim.'})
    
    # PersonelBirim kaydı oluştur
    personel_birim = PersonelBirim.objects.create(
        personel=personel,
        birim=birim,
        gecis_tarihi=gecis_tarihi,
        sorumlu=sorumlu,
        not_text=not_text,
        created_by=request.user
    )
    
    return JsonResponse({
        'success': True, 
        'message': 'Birim ataması başarıyla kaydedildi.',
        'personel_birim': {
            'id': personel_birim.id,
            'birim_ad': personel_birim.birim.ad,
            'bina_ad': personel_birim.birim.bina.ad,
            'ust_birim_ad': personel_birim.birim.ust_birim.ad,
            'gecis_tarihi': (
                personel_birim.gecis_tarihi.strftime('%d.%m.%Y')
                if hasattr(personel_birim.gecis_tarihi, 'strftime')
                else str(personel_birim.gecis_tarihi)
            ),
            'sorumlu': personel_birim.sorumlu
        }
    })

@login_required
@require_POST
def personel_birim_sil(request, pk):
    """Personel birim kaydını sil"""
    try:
        from .models.BirimYonetimi import PersonelBirim
    except ImportError:
        from ik_core.models.BirimYonetimi import PersonelBirim
        
    try:
        personel_birim = PersonelBirim.objects.get(pk=pk)
        personel_birim.delete()
        return JsonResponse({'success': True, 'message': 'Kayıt başarıyla silindi.'})
    except PersonelBirim.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Kayıt bulunamadı.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata oluştu: {str(e)}'})

@login_required
def gorevlendirme_yazisi(request, personel_birim_id):
    """Görevlendirme yazısı PDF oluşturma"""
    try:
        personel_birim = get_object_or_404(PersonelBirim.objects.select_related('personel', 'birim__bina', 'birim__ust_birim', 'personel__unvan'), id=personel_birim_id)
        imza_id = request.GET.get('imza_id')
        
        if not imza_id:
            return JsonResponse({'success': False, 'message': 'İmza seçilmelidir.'})
        
        teblig_imzasi = get_object_or_404(TebligImzasi, id=imza_id)
        
        # PDF template'ini render et
        html = render_to_string('ik_core/pdf/gorevlendirme_yazisi.html', {
            'personel_birim': personel_birim,
            'teblig_imzasi': teblig_imzasi,
        })
        
        # PDF konfigürasyonu
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {
            'page-size': 'A4',
            'orientation': 'Portrait',
            'margin-top': '7cm',
            'margin-right': '1.5cm',
            'margin-bottom': '1.1cm',
            'margin-left': '1.5cm',
            'encoding': 'UTF-8',
            'no-outline': None,
            'enable-local-file-access': '',
            'enable-external-links': True,
            'quiet': ''
        }
        
        pdf = pdfkit.from_string(html, False, options=options, configuration=config)
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"Gorevlendirme_{personel_birim.personel.ad_soyad}.pdf"
        filename = filename.replace(' ', '_')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'PDF oluşturulurken hata: {str(e)}'})

@login_required
def gelen_giden_personel_list(request):
    filter_type = request.GET.get('filter_type', 'gelen')
    
    # Property üzerinden filtreleme istendiği için tüm kayıtları çekip Python'da işliyoruz.
    qs = Personel.objects.select_related('unvan', 'brans').prefetch_related(
        'personelbirim_set',
        'personelbirim_set__birim',
        'personelbirim_set__birim__bina'
    )
    
    personeller = []
    for p in qs:
        p_durum = p.durum
        if filter_type == 'gelen':
            if p_durum == 'Aktif':
                personeller.append(p)
        else: # giden
            if p_durum in ['Pasif', 'Kurumdan Ayrıldı']:
                personeller.append(p)
    
    # Sıralama
    if filter_type == 'gelen':
        personeller.sort(key=lambda x: x.goreve_baslama_tarihi or date.min, reverse=True)
    else:
        personeller.sort(key=lambda x: x.ayrilma_tarihi or date.min, reverse=True)

    context = {
        'personeller': personeller,
        'filter_type': filter_type,
    }
    return render(request, 'ik_core/gelen_giden_personel_list.html', context)
