from io import BytesIO
import os
from django.db import IntegrityError
import pandas as pd
from openpyxl import load_workbook, Workbook
import calendar
from datetime import datetime, timedelta, date
import json
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from PersonelYonSis import settings
from .models import Birim, Mesai, Mesai_Tanimlari, Personel, PersonelListesi, PersonelListesiKayit, UserBirim, Kurum, UstBirim, Idareci, Izin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import locale
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db import transaction
from django.contrib.auth import get_user_model
from dateutil.relativedelta import relativedelta
from .valuelists import CKYS_BTF_VALUES
User = get_user_model()


@login_required
def onceki_donem_personel(request, year, month, birim_id):
    # year: int, month: int, birim_id: int
    # Bir önceki ayı bul
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    try:
        birim = Birim.objects.get(BirimID=birim_id)
    except Birim.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Birim bulunamadı.'})
    try:
        liste = PersonelListesi.objects.get(birim=birim, yil=prev_year, ay=prev_month)
    except PersonelListesi.DoesNotExist:
        # Önceki dönem yoksa boş liste dön
        return JsonResponse({'status': 'success', 'data': []})
    kayitlar = PersonelListesiKayit.objects.filter(liste=liste).select_related('personel')
    data = []
    for kayit in kayitlar:
        p = kayit.personel
        data.append({
            "PersonelTCKN": p.PersonelTCKN,
            "PersonelName": p.PersonelName,
            "PersonelTitle": p.PersonelTitle,
        })
    return JsonResponse({'status': 'success', 'data': data})

@csrf_exempt
@require_POST
@login_required
def personel_kaydet(request):
    import json
    try:
        data = json.loads(request.body)
        donem = data.get('donem')
        birim_id = data.get('birim_id')
        personeller = data.get('personeller', [])
        if not (donem and birim_id and personeller):
            return JsonResponse({'status': 'error', 'message': 'Eksik veri.'})
        year, month = map(int, donem.replace('-', '/').split('/'))
        birim = Birim.objects.get(BirimID=birim_id)
        with transaction.atomic():
            liste, _ = PersonelListesi.objects.get_or_create(birim=birim, yil=year, ay=month)
            eklenenler = []
            for p in personeller:
                pid = p.get('PersonelTCKN')
                pname = p.get('PersonelName')
                ptitle = p.get('PersonelTitle')
                personel, _ = Personel.objects.get_or_create(
                    PersonelTCKN=pid,
                    defaults={'PersonelName': pname, 'PersonelTitle': ptitle}
                )
                # Eğer personel varsa, ad/ünvan güncelle
                if personel.PersonelName != pname or personel.PersonelTitle != ptitle:
                    personel.PersonelName = pname
                    personel.PersonelTitle = ptitle
                    personel.save()
                # Listeye ekle
                PersonelListesiKayit.objects.get_or_create(liste=liste, personel=personel)
                eklenenler.append(pid)
        return JsonResponse({'status': 'success', 'message': f'{len(eklenenler)} personel eklendi.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
try:
    # Windows için
    locale.setlocale(locale.LC_ALL, 'turkish')
except locale.Error:
    try:
        # Linux/Unix için
        locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')
    except locale.Error:
        # Hiçbiri çalışmazsa varsayılan locale kullan
        locale.setlocale(locale.LC_ALL, '')
from django.conf import settings

def excel_export(request):
    current_year = int(request.GET.get('year', datetime.now().year))
    current_month = int(request.GET.get('month', datetime.now().month))

    df = pd.DataFrame(columns=["Personel Adı", "Personel Unvanı"])

    personeller = Personel.objects.all()
    mesailer = Mesai.objects.filter(MesaiDate__year=current_year, MesaiDate__month=current_month)

    rows = []
    for personel in personeller:
        personel.mesai_data = mesailer.filter(Personel=personel)
        row = {
            "Personel Adı": personel.PersonelName,
            "Personel Unvanı": personel.PersonelTitle
        }
        for mesai in personel.mesai_data:
            row[mesai.MesaiDate.strftime("%Y-%m-%d")] = mesai.MesaiData
        rows.append(row)

    df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)

    template_path = os.path.join(settings.STATIC_ROOT, 'excels', 'cizelgeSablon1.xlsx')
    
    with BytesIO() as buffer:
        try:
            if os.path.exists(template_path):
                # Şablon varsa yükle
                book = load_workbook(template_path)
            else:
                # Şablon yoksa yeni bir Workbook oluştur
                book = Workbook()

            # Çalışma sayfası olup olmadığını kontrol et
            if not book.worksheets:
                book.create_sheet(title="Mesai Verileri")

            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}

                # Veriyi A2 hücresinden itibaren yazdır
                df.to_excel(writer, index=False, startrow=1, sheet_name='Mesai Verileri')

                # Başlık ekle
                sheet = writer.sheets['Mesai Verileri']
                sheet['A1'] = f"{current_year} - {current_month} dönemi Mesai verileri"

            # Yazdırma işlemi
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{current_year}_{current_month}_mesai_verileri.xlsx"'
            return response

        except Exception as e:
            return HttpResponse(f"Hata oluştu: {str(e)}", status=500)


@login_required
def cizelge(request):
    user = request.user
    user_birimler = UserBirim.objects.filter(user=user).select_related('birim')
    print(f"Kullanıcı: {user.Username}, Birimler: {[b.birim.BirimAdi for b in user_birimler]}")
    izinler = Izin.objects.all()
    kurumlar = Kurum.objects.all()
    print(f"Kurumlar: {kurumlar}")
    ust_birimler = UstBirim.objects.all()
    idareciler = Idareci.objects.all()
    # Dönemler: mevcut aydan 6 ay önce ile 2 ay sonrası arası
    today = date.today().replace(day=1)
    donemler = []
    for i in range(-6, 3):
        d = today + relativedelta(months=i)
        value = f"{d.year}/{d.month:02d}"
        label = value
        donemler.append({'value': value, 'label': label})

    selected_birim_id = request.GET.get('birim_id') or ""
    selected_donem = request.GET.get('donem') or ""
    pastcontext = {
            "user_birimler": user_birimler,
            "selected_birim_id": selected_birim_id,
            "donemler": donemler,
            "selected_donem": selected_donem,
            "mesai_options": Mesai_Tanimlari.objects.all(),
            "kurumlar": kurumlar,
            "ust_birimler": ust_birimler,
            "idareciler": idareciler,
            "izinler": izinler,
        }
    # Eğer GET ile birim ve dönem seçilmemişse, context'e sadece seçim listelerini gönder
    if not selected_birim_id or not selected_donem:
        return render(request, 'mercis657/cizelge.html', pastcontext)
    # Dönem bilgisini parse et ("YYYY/MM" formatı)
    try:
        current_year, current_month = map(int, selected_donem.split('/'))
    except Exception:
        messages.warning(request, "Geçersiz dönem formatı.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    birim_id = selected_birim_id

    if not birim_id:
        messages.warning(request, "Lütfen bir birim seçiniz.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    birim = get_object_or_404(Birim, BirimID=birim_id)

    if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
        return HttpResponseForbidden("Bu birim için yetkiniz yok.")

    # Liste varsa getir
    try:
        liste = PersonelListesi.objects.get(birim=birim, yil=current_year, ay=current_month)
    except PersonelListesi.DoesNotExist:
        messages.warning(request, f"{current_month}/{current_year} için personel listesi oluşturulmamış.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    personeller = Personel.objects.filter(personellistesikayit__liste=liste).distinct()
    mesai_tanimlari = Mesai_Tanimlari.objects.all()
    mesailer = Mesai.objects.filter(
        Personel__in=personeller,
        MesaiDate__year=current_year,
        MesaiDate__month=current_month
    ).select_related('MesaiTanim')

    # Gün listesi
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    days = [
        {
            'full_date': f"{current_year}-{current_month:02}-{day:02}",
            'day_num': day,
            'is_weekend': calendar.weekday(current_year, current_month, day) >= 5
        }
        for day in range(1, days_in_month + 1)
    ]

    # Personel mesai eşleme
    mesai_map = {}
    for mesai in mesailer:
        key = f"{mesai.Personel.PersonelID}_{mesai.MesaiDate.strftime('%Y-%m-%d')}"
        mesai_map[key] = mesai.MesaiTanim.Saat if mesai.MesaiTanim else ""

    # Personel nesnesine mesai bilgisi ekleyelim
    for p in personeller:
        p.mesai_data = [
            {
                "MesaiDate": day['full_date'],
                "Saat": mesai_map.get(f"{p.PersonelID}_{day['full_date']}", "")
            } for day in days
        ]

    context = {
        "personeller": personeller,
        "mesai_options": mesai_tanimlari,
        "izinler": izinler,
        "days": days,
        "birim": birim,
        "current_year": current_year,
        "current_month": current_month,
        "months": [{'value': i, 'label': calendar.month_name[i]} for i in range(1, 13)],
        "years": [year for year in range(2023, 2027)],
        "user_birimler": user_birimler,
        "selected_birim_id": selected_birim_id,
        "donemler": donemler,
        "selected_donem": selected_donem,
        "kurumlar": kurumlar,
        "ust_birimler": ust_birimler,
        "idareciler": idareciler,
    }
    return render(request, 'mercis657/cizelge.html', context)

def personeller(request):
    personeller = Personel.objects.all()
    return render(request, 'personeller.html', {"personeller": personeller})
# Personel Ekleme Formunu Geri Döndür
def personel_ekle_form(request):
    return render(request, 'personel_ekle_form.html')

# Yeni Personel Ekleme İşlemi
def personel_ekle(request):
    if request.method == 'POST':
        # Form verilerini al
        personel_tc = request.POST['PersonelTCKN']
        personel_name = request.POST['PersonelName']
        personel_title = request.POST['PersonelTitle']
        
        # Yeni personel kaydet
        personel = Personel(
            PersonelTCKN=personel_tc,
            PersonelName=personel_name,
            PersonelTitle=personel_title
        )
        personel.save()

        # Başarı mesajı ekleyebilirsiniz
        return redirect('mercis657:personeller')  # Personel listesine yönlendir
    return HttpResponse("Geçersiz istek", status=400)
def personel_update(request):
    if request.method == 'POST':
        personel_tc = request.POST.get('tckn')
        personel_name = request.POST.get('name')
        personel_title = request.POST.get('title')

        # Personeli bul
        personel = get_object_or_404(Personel, PersonelTCKN=personel_tc)
        
        # Güncellenen alanlar
        personel.PersonelName = personel_name
        personel.PersonelTitle = personel_title
        personel.save()  # Değişiklikleri kaydet
        
        return JsonResponse({'status': 'success'})  # Başarı mesajı

    return JsonResponse({'status': 'error'})  # Hatalı durum
def mesai_tanimlari(request):
    mesai_tanimlari = Mesai_Tanimlari.objects.all()
    return render(request, 'mercis657/mesai_tanimlari.html', {"mesai_tanimlari": mesai_tanimlari})
# Yeni Mesai Tanımı Ekleme Fonksiyonu
def add_mesai_tanim(request):
    if request.method == 'POST':
        # Formdan gelen verileri alıyoruz
        saat = request.POST.get('Saat')
        gunduz_mesaisi = request.POST.get('GunduzMesaisi') == 'on'
        aksam_mesaisi = request.POST.get('AksamMesaisi') == 'on'
        gece_mesaisi = request.POST.get('GeceMesaisi') == 'on'
        ise_geldi = request.POST.get('IseGeldi') == 'on'
        sonraki_gune_sarkiyor = request.POST.get('SonrakiGuneSarkiyor') == 'on'
        ara_dinlenme = request.POST.get('AraDinlenme')
        gecerli_mesai = request.POST.get('GecerliMesai') == 'on'
        ckys_btf_karsiligi = request.POST.get('CKYS_BTF_Karsiligi')

        # Ara dinlenme süresini timedelta nesnesine dönüştürme
        if ara_dinlenme:
            hours, minutes = map(int, ara_dinlenme.split(':'))
            ara_dinlenme_td = timedelta(hours=hours, minutes=minutes)
        else:
            ara_dinlenme_td = None

        # Yeni mesai tanımı oluşturma
        yeni_mesai = Mesai_Tanimlari(
            Saat=saat,
            GunduzMesaisi=gunduz_mesaisi,
            AksamMesaisi=aksam_mesaisi,
            GeceMesaisi=gece_mesaisi,
            IseGeldi=ise_geldi,
            SonrakiGuneSarkiyor=sonraki_gune_sarkiyor,
            AraDinlenme=ara_dinlenme_td,
            GecerliMesai=gecerli_mesai,
            CKYS_BTF_Karsiligi=ckys_btf_karsiligi
        )
        
        # Mesai süresini hesaplayarak kaydet
        yeni_mesai.calculate_sure()
        yeni_mesai.save()

        return redirect('mercis657:mesai_tanimlari')  # İşlem tamamlandığında liste sayfasına yönlendirme

    return render(request, 'mesai_tanimlari.html')

def mesai_tanim_update(request):
    mesai_id = request.POST.get('mesai_id')
    mesai = get_object_or_404(Mesai_Tanimlari, id=mesai_id)
    
    if request.method == 'POST':
        try:
            # Formdan gelen güncellenmiş verileri alıyoruz
            mesai.Saat = request.POST.get('saat')
            mesai.GunduzMesaisi = request.POST.get('GunduzMesaisi') == 'on'
            mesai.AksamMesaisi = request.POST.get('AksamMesaisi') == 'on'
            mesai.GeceMesaisi = request.POST.get('GeceMesaisi') == 'on'
            mesai.IseGeldi = request.POST.get('IseGeldi') == 'on'
            mesai.SonrakiGuneSarkiyor = request.POST.get('SonrakiGuneSarkiyor') == 'on'
            ara_dinlenme = request.POST.get('AraDinlenme')
            mesai.AraDinlenme = ara_dinlenme if ara_dinlenme else None
            mesai.GecerliMesai = request.POST.get('GecerliMesai') == 'on'
            mesai.CKYS_BTF_Karsiligi = request.POST.get('CKYS_BTF_Karsiligi')

            mesai.save()  # Güncellenmiş verileri kaydet
            return JsonResponse({'status': 'success'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Geçersiz istek.'})
def delete_mesai_tanim(request):
    if request.method == 'POST':
        mesai_id = request.POST.get('mesai_id')
        try:
            mesai = get_object_or_404(Mesai_Tanimlari, id=mesai_id)
            mesai.delete()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Geçersiz istek.'})


@login_required
def personel_listeleri(request):
    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim_id', flat=True)
    birimler = Birim.objects.filter(id__in=user_birimler)
    listeler = PersonelListesi.objects.filter(birim__in=birimler).order_by('-yil', '-ay')
    return render(request, 'mercis657/personel_listeleri.html', {
        'birimler': birimler,
        'listeler': listeler
    })


@login_required
def personel_listesi_olustur(request):
    if request.method == 'POST':
        birim_id = request.POST.get('birim_id')
        yil = int(request.POST.get('yil'))
        ay = int(request.POST.get('ay'))

        birim = get_object_or_404(Birim, id=birim_id)

        if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
            return HttpResponseForbidden('Bu birim için yetkiniz yok.')

        try:
            liste, created = PersonelListesi.objects.get_or_create(birim=birim, yil=yil, ay=ay)
            if created:
                messages.success(request, 'Personel listesi oluşturuldu.')
            else:
                messages.warning(request, 'Bu ay ve birim için liste zaten var.')
        except IntegrityError:
            messages.error(request, 'Liste oluşturulurken bir hata oluştu.')

        return redirect('mercis657:personel_listeleri')


@login_required
def personel_listesi_detay(request, liste_id):
    liste = get_object_or_404(PersonelListesi, id=liste_id)

    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        return HttpResponseForbidden('Bu listeye erişim yetkiniz yok.')

    mevcut_personeller = PersonelListesiKayit.objects.filter(personel_listesi=liste).select_related('personel')
    tum_personeller = Personel.objects.all().order_by('FirstName', 'LastName')

    return render(request, 'mercis657/personel_listesi_detay.html', {
        'liste': liste,
        'mevcut_personeller': mevcut_personeller,
        'tum_personeller': tum_personeller
    })


@login_required
def personel_ekle_listeye(request, liste_id):
    if request.method == 'POST':
        liste = get_object_or_404(PersonelListesi, id=liste_id)

        if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
            return HttpResponseForbidden('Bu listeye erişim yetkiniz yok.')

        personel_id = request.POST.get('personel_id')
        personel = get_object_or_404(Personel, PersonelID=personel_id)

        kayit, created = PersonelListesiKayit.objects.get_or_create(
            personel_listesi=liste,
            personel=personel
        )

        if created:
            messages.success(request, 'Personel listeye eklendi.')
        else:
            messages.warning(request, 'Bu personel zaten listede.')

        return redirect('mercis657:personel_listesi_detay', liste_id=liste.id)


@login_required
def personel_kaldir_liste(request, liste_id, kayit_id):
    liste = get_object_or_404(PersonelListesi, id=liste_id)
    kayit = get_object_or_404(PersonelListesiKayit, id=kayit_id, personel_listesi=liste)

    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        return HttpResponseForbidden('Bu listeye erişim yetkiniz yok.')

    kayit.delete()
    messages.success(request, 'Personel listeden kaldırıldı.')
    return redirect('mercis657:personel_listesi_detay', liste_id=liste.id)

@login_required
def personel_cikar(request, liste_id, personel_id):
    liste = get_object_or_404(PersonelListesi, id=liste_id)
    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        messages.warning(request, 'Yetkisiz işlem.')
        return redirect('mercis657:personel_listeleri')

    kayit = get_object_or_404(PersonelListesiKayit, liste=liste, personel_id=personel_id)
    kayit.delete()
    messages.success(request, 'Personel listeden çıkarıldı.')
    return redirect('mercis657:personel_listesi_detay', liste_id=liste.id)

def birim_yonetim(request):
    birimler = Birim.objects.select_related('Kurum', 'UstBirim', 'Idareci').all()
    birim_list = []
    for birim in birimler:
        yetkiler = UserBirim.objects.filter(birim=birim).select_related('user')
        yetkili_users = [
            {
                "username": y.user.Username,
                "full_name": y.user.FullName
            }
            for y in yetkiler
        ]
        birim_list.append({
            "id": birim.BirimID,
            "adi": birim.BirimAdi,
            "kurum": birim.Kurum.ad if birim.Kurum else "",
            "ust_birim": birim.UstBirim.ad if birim.UstBirim else "",
            "idareci": birim.Idareci.ad if birim.Idareci else "",
            "yetkili_sayisi": len(yetkili_users),
            "yetkililer": yetkili_users,
        })
    kurumlar = Kurum.objects.all()
    ust_birimler = UstBirim.objects.all()
    idareciler = Idareci.objects.all()
    return render(request, "mercis657/birim_yonetim.html", {
        "birimler": birim_list,
        "kurumlar": kurumlar,
        "ust_birimler": ust_birimler,
        "idareciler": idareciler,
    })

@csrf_exempt
def birim_ekle(request):
    if request.method == 'POST':
        ad = request.POST.get('BirimAdi')
        kurum_id = request.POST.get('Kurum') or None
        ust_id = request.POST.get('UstBirim') or None
        mudur_id = request.POST.get('idareci') or None
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Birim adı zorunlu.'})
        try:
            birim = Birim.objects.create(
                BirimAdi=ad,
                Kurum_id=kurum_id if kurum_id else None,
                UstBirim_id=ust_id if ust_id else None,
                Idareci_id=mudur_id if mudur_id else None
            )
            # Yeni eklenen birime mevcut kullanıcıyı yetkilendir
            UserBirim.objects.create(user=request.user, birim=birim)
            return JsonResponse({'status': 'success', 'birim_id': birim.BirimID})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Geçersiz istek.'})

@login_required
def birim_detay(request, birim_id):
    try:
        birim = get_object_or_404(Birim, BirimID=birim_id)
        # Kullanıcının bu birim için yetkisi var mı kontrol et
        if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
             return JsonResponse({'status': 'error', 'message': 'Bu birim için yetkiniz yok.'}, status=403)

        data = {
            'BirimID': birim.BirimID,
            'BirimAdi': birim.BirimAdi,
            'Kurum': birim.Kurum.pk if birim.Kurum else None,
            'UstBirim': birim.UstBirim.pk if birim.UstBirim else None,
            'idareci': birim.idareci.pk if birim.idareci else None,
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Birim.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Birim bulunamadı.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt # CSRF korumasını geçici olarak devre dışı bırakıyoruz, uygun token yönetimi eklenmeli
@require_POST
@login_required
def birim_guncelle(request, birim_id):
    try:
        birim = get_object_or_404(Birim, BirimID=birim_id)
        # Kullanıcının bu birim için yetkisi var mı kontrol et
        if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
             return JsonResponse({'status': 'error', 'message': 'Bu birim için yetkiniz yok.'}, status=403)

        ad = request.POST.get('birimAdi')
        kurum_id = request.POST.get('Kurum') or None
        ust_id = request.POST.get('UstBirim') or None
        mudur_id = request.POST.get('idareci') or None

        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Birim adı zorunlu.'})

        birim.BirimAdi = ad
        birim.Kurum_id = kurum_id
        birim.UstBirim_id = ust_id
        birim.idareci_id = mudur_id
        birim.save()

        return JsonResponse({'status': 'success', 'message': 'Birim başarıyla güncellendi.'})
    except Birim.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Birim bulunamadı.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt # CSRF korumasını geçici olarak devre dışı bırakıyoruz, uygun token yönetimi eklenmeli
@require_POST
@login_required
def birim_sil(request, birim_id):
    try:
        birim = get_object_or_404(Birim, BirimID=birim_id)
         # Kullanıcının bu birim için yetkisi var mı kontrol et (isteğe bağlı: silme yetkisi farklı olabilir)
        if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
             return JsonResponse({'status': 'error', 'message': 'Bu birim için yetkiniz yok.'}, status=403)

        birim.delete()
        return JsonResponse({'status': 'success', 'message': 'Birim başarıyla silindi.'})
    except Birim.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Birim bulunamadı.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def kurum_ekle(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Kurum adı zorunlu.'})
        if Kurum.objects.filter(ad=ad).exists():
            return JsonResponse({'status': 'error', 'message': 'Bu ad ile kurum zaten var.'})
        Kurum.objects.create(ad=ad)
        messages.success(request, f'{ad} isimli Kurum başarıyla eklendi.')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def kurum_guncelle(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Kurum adı zorunlu.'})
        kurum = Kurum.objects.get(pk=pk)
        kurum.ad = ad
        kurum.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def kurum_sil(request, pk):
    if request.method == 'POST':
        try:
            kurum = Kurum.objects.get(pk=pk)
            kurum.delete()
            messages.success(request, f'{kurum.ad} isimli Kurum başarıyla silindi.')
            return JsonResponse({'status': 'success'})
        except Kurum.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Kurum bulunamadı.'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def ust_birim_ekle(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Üst birim adı zorunlu.'})
        if UstBirim.objects.filter(ad=ad).exists():
            return JsonResponse({'status': 'error', 'message': 'Bu ad ile üst birim zaten var.'})
        UstBirim.objects.create(ad=ad)
        messages.success(request, f'{ad} isimli Üst Birim başarıyla eklendi.')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def ust_birim_guncelle(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'Üst birim adı zorunlu.'})
        ust = UstBirim.objects.get(pk=pk)
        ust.ad = ad
        ust.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def ust_birim_sil(request, pk):
    if request.method == 'POST':
        try:
            ust = UstBirim.objects.get(pk=pk)
            ust.delete()
            messages.success(request, f'{ust.ad} isimli Üst Birim başarıyla silindi.')
            return JsonResponse({'status': 'success'})
        except UstBirim.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Üst birim bulunamadı.'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def idareci_ekle(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'İdareci adı zorunlu.'})
        if Idareci.objects.filter(ad=ad).exists():
            return JsonResponse({'status': 'error', 'message': 'Bu ad ile idareci zaten var.'})
        Idareci.objects.create(ad=ad)
        messages.success(request, f'{ad} isimli İdareci başarıyla eklendi.')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def idareci_guncelle(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        ad = data.get('ad', '').strip()
        if not ad:
            return JsonResponse({'status': 'error', 'message': 'İdareci adı zorunlu.'})
        idareci = Idareci.objects.get(pk=pk)
        idareci.ad = ad
        idareci.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

def kullanici_ara(request):
    username = request.GET.get('username', '').strip()
    try:
        user = User.objects.get(Username=username)
        data = {
            "username": user.Username,
            "full_name": user.FullName
        }
        return JsonResponse({"status": "success", "data": data})
    except User.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Kullanıcı bulunamadı."})

@csrf_exempt
def birim_yetki_ekle(request, birim_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            user = User.objects.get(Username=username)
            birim = Birim.objects.get(pk=birim_id)
            obj, created = UserBirim.objects.get_or_create(user=user, birim=birim)
            if created:
                return JsonResponse({"status": "success"})
            else:
                return JsonResponse({"status": "error", "message": "Kullanıcı zaten yetkili."})
        except User.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Kullanıcı bulunamadı."})
        except Birim.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Birim bulunamadı."})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
    return JsonResponse({"status": "error", "message": "Geçersiz istek."})

@login_required
def tanimlamalar(request):
    kurumlar = Kurum.objects.all()
    ust_birimler = UstBirim.objects.all()
    idareciler = Idareci.objects.all()
    mesai_tanimlari = Mesai_Tanimlari.objects.all()
    return render(request, "mercis657/tanimlamalar.html", {
        "kurumlar": kurumlar,
        "ust_birimler": ust_birimler,
        "idareciler": idareciler,
        "mesai_tanimlari": mesai_tanimlari,
    })

@csrf_exempt
def kurum_toggle_aktif(request, pk):
    if request.method == 'POST':
        try:
            kurum = Kurum.objects.get(pk=pk)
            kurum.aktif = not kurum.aktif
            kurum.save()
            messages.success(request, f"{kurum.ad} kurumunun durumu {'aktif' if kurum.aktif else 'pasif'} olarak güncellendi.")
            return JsonResponse({'status': 'success', 'aktif': kurum.aktif})
        except Kurum.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Kurum bulunamadı.'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def ust_birim_toggle_aktif(request, pk):
    if request.method == 'POST':
        try:
            ust = UstBirim.objects.get(pk=pk)
            ust.aktif = not ust.aktif
            ust.save()
            messages.success(request, f"{ust.ad} üst biriminin durumu {'aktif' if ust.aktif else 'pasif'} olarak güncellendi.")
            return JsonResponse({'status': 'success', 'aktif': ust.aktif})
        except UstBirim.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'İdare bulunamadı.'})
    return JsonResponse({'status': 'error'})

@csrf_exempt
def idareci_toggle_aktif(request, pk):
    if request.method == 'POST':
        try:
            idareci = Idareci.objects.get(pk=pk)
            idareci.aktif = not idareci.aktif
            idareci.save()
            messages.success(request, f"{idareci.ad} idarecisinin durumu {'aktif' if idareci.aktif else 'pasif'} olarak güncellendi.")
            return JsonResponse({'status': 'success', 'aktif': idareci.aktif})
        except Idareci.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'İdareci bulunamadı.'})
    return JsonResponse({'status': 'error'})