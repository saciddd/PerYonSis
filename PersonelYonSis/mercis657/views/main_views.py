from io import BytesIO
import os
from django.db import IntegrityError
import pandas as pd
from openpyxl import load_workbook, Workbook
import calendar
from datetime import datetime, timedelta, date
import json
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from PersonelYonSis import settings
from ..models import Birim, Mesai, Mesai_Tanimlari, Personel, PersonelListesi, PersonelListesiKayit, UserBirim, Kurum, UstBirim, Idareci, Izin, ResmiTatil, IlkListe, SabitMesai
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import locale
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.contrib.auth import get_user_model
from dateutil.relativedelta import relativedelta
from ..valuelists import CKYS_BTF_VALUES
from ..forms import MesaiTanimForm, ResmiTatilForm, YarimZamanliCalismaForm
User = get_user_model()
try:
    # Windows için
    locale.setlocale(locale.LC_ALL, 'turkish')
except locale.Error:
    try:
        # Linux/Unix için
        locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')
    except locale.Error:
        # Hiçbiri çalışmazsa varsayılan locale kullan
        locale.setlocale(locale.LC_ALL, '')
from django.conf import settings

def excel_export(request):
    current_year = int(request.GET.get('year', datetime.now().year))
    current_month = int(request.GET.get('month', datetime.now().month))

    df = pd.DataFrame(columns=["Personel Adı", "Personel Unvanı"])

    personeller = Personel.objects.all()
    mesailer = Mesai.objects.filter(MesaiDate__year=current_year, MesaiDate__month=current_month)

    rows = []
    for personel in personeller:
        personel.mesai_data = mesailer.filter(Personel=personel)
        row = {
            "Personel Adı": f"{personel.PersonelName} {personel.PersonelSurname}",
            "Personel Unvanı": personel.PersonelTitle
        }
        for mesai in personel.mesai_data:
            row[mesai.MesaiDate.strftime("%Y-%m-%d")] = mesai.MesaiData
        rows.append(row)

    df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)

    template_path = os.path.join(settings.STATIC_ROOT, 'excels', 'cizelgeSablon1.xlsx')
    
    with BytesIO() as buffer:
        try:
            if os.path.exists(template_path):
                # Şablon varsa yükle
                book = load_workbook(template_path)
            else:
                # Şablon yoksa yeni bir Workbook oluştur
                book = Workbook()

            # Çalışma sayfası olup olmadığını kontrol et
            if not book.worksheets:
                book.create_sheet(title="Mesai Verileri")

            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}

                # Veriyi A2 hücresinden itibaren yazdır
                df.to_excel(writer, index=False, startrow=1, sheet_name='Mesai Verileri')

                # Başlık ekle
                sheet = writer.sheets['Mesai Verileri']
                sheet['A1'] = f"{current_year} - {current_month} dönemi Mesai verileri"

            # Yazdırma işlemi
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'inline; filename="{current_year}_{current_month}_mesai_verileri.xlsx"'
            return response

        except Exception as e:
            return HttpResponse(f"Hata oluştu: {str(e)}", status=500)

@login_required
def cizelge(request):
    user = request.user
    user_birimler = UserBirim.objects.filter(user=user).select_related('birim')
    izinler = Izin.objects.all()
    kurumlar = Kurum.objects.all()
    ust_birimler = UstBirim.objects.all()
    idareciler = Idareci.objects.all()
    # Dönemler: mevcut aydan 6 ay önce ile 2 ay sonrası arası
    today = date.today().replace(day=1)
    donemler = []
    for i in range(-6, 3):
        d = today + relativedelta(months=i)
        value = f"{d.year}/{d.month:02d}"
        label = value
        donemler.append({'value': value, 'label': label})

    selected_birim_id = request.GET.get('birim_id') or ""
    selected_donem = request.GET.get('donem') or ""
    
    # Güvenli sabit_mesailer listesi oluştur
    sabit_mesailer = []
    try:
        for sm in SabitMesai.objects.all():
            try:
                # ara_dinlenme değerini kontrol et
                if sm.ara_dinlenme is not None:
                    float(sm.ara_dinlenme)
                sabit_mesailer.append(sm)
            except (ValueError, TypeError):
                # Problemli kayıtları atla
                continue
    except Exception:
        sabit_mesailer = []
    
    pastcontext = {
            "user_birimler": user_birimler,
            "selected_birim_id": selected_birim_id,
            "donemler": donemler,
            "selected_donem": selected_donem,
            "mesai_options": Mesai_Tanimlari.objects.all(),
            "sabit_mesailer": sabit_mesailer,  # Modal için eklendi
            "kurumlar": kurumlar,
            "ust_birimler": ust_birimler,
            "idareciler": idareciler,
            "izinler": izinler,
        }
    # Eğer GET ile birim ve dönem seçilmemişse, context'e sadece seçim listelerini gönder
    if not selected_birim_id or not selected_donem:
        return render(request, 'mercis657/cizelge.html', pastcontext)
    # Dönem bilgisini parse et ("YYYY/MM" formatı)
    try:
        current_year, current_month = map(int, selected_donem.split('/'))
    except Exception:
        messages.warning(request, "Geçersiz dönem formatı.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    birim_id = selected_birim_id

    if not birim_id:
        messages.warning(request, "Lütfen bir birim seçiniz.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    birim = get_object_or_404(Birim, BirimID=birim_id)

    if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
        return HttpResponseForbidden("Bu birim için yetkiniz yok.")

    # Liste varsa getir
    try:
        liste = PersonelListesi.objects.get(birim=birim, yil=current_year, ay=current_month)
    except PersonelListesi.DoesNotExist:
        messages.warning(request, f"{current_month}/{current_year} için personel listesi oluşturulmamış.")
        return render(request, 'mercis657/cizelge.html', pastcontext)

    # Sıralı personel listesi
    kayitlar = PersonelListesiKayit.objects.filter(liste=liste).select_related('personel').order_by('sira_no', 'personel__PersonelName', 'personel__PersonelSurname')
    personeller = [k.personel for k in kayitlar]
    mesai_tanimlari = Mesai_Tanimlari.objects.all()
    mesailer = Mesai.objects.filter(
        Personel__in=personeller,
        MesaiDate__year=current_year,
        MesaiDate__month=current_month
    ).select_related('MesaiTanim', 'Izin').prefetch_related('yedekler', 'mercis657_stoplar')

    # Resmi tatilleri al
    resmi_tatiller = ResmiTatil.objects.filter(
        TatilTarihi__year=current_year,
        TatilTarihi__month=current_month
    ).values_list('TatilTarihi', flat=True)
    
    # Gün listesi
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    days = [
        {
            'full_date': f"{current_year}-{current_month:02}-{day:02}",
            'day_num': day,
            'is_weekend': calendar.weekday(current_year, current_month, day) >= 5,
            'is_resmi_tatil': f"{current_year}-{current_month:02}-{day:02}" in [t.strftime('%Y-%m-%d') for t in resmi_tatiller]
        }
        for day in range(1, days_in_month + 1)
    ]

    # Personel mesai eşleme
    mesai_map = {}
    for mesai in mesailer:
        key = f"{mesai.Personel.PersonelID}_{mesai.MesaiDate.strftime('%Y-%m-%d')}"
        last_backup = mesai.yedekler.order_by('-created_at').first()
        # get latest stop if any
        last_stop = None
        try:
            last_stop = mesai.mercis657_stoplar.order_by('-created_at').first()
        except Exception:
            last_stop = None

        mesai_map[key] = {
            "MesaiID": mesai.MesaiID,
            "MesaiTanimID": mesai.MesaiTanim.id if mesai.MesaiTanim else None,
            "MesaiTanimRenk": mesai.MesaiTanim.Renk if mesai.MesaiTanim else None,
            "IzinID": mesai.Izin.id if mesai.Izin else None,
            "SistemdekiIzin": mesai.SistemdekiIzin,
            "MesaiNotu": mesai.MesaiNotu,
            "Saat": mesai.MesaiTanim.Saat if mesai.MesaiTanim else "",
            "IzinAd": mesai.Izin.ad if mesai.Izin else "",
            "Degisiklik": mesai.Degisiklik,
            "PrevSaat": (last_backup.MesaiTanim.Saat if (last_backup and last_backup.MesaiTanim) else ""),
            "PrevIzinAd": (last_backup.Izin.ad if (last_backup and last_backup.Izin) else "")
        }

        if last_stop:
            # format datetimes for display
            try:
                sb = last_stop.StopBaslangic.strftime('%H:%M')
            except Exception:
                sb = str(last_stop.StopBaslangic)
            try:
                se = last_stop.StopBitis.strftime('%H:%M')
            except Exception:
                se = str(last_stop.StopBitis)

            mesai_map[key]['StopKaydi'] = {
                'StopBaslangic': sb,
                'StopBitis': se,
                'Sure': last_stop.Sure,
                'created_by': str(last_stop.created_by) if last_stop.created_by else '' ,
                'id': last_stop.id
            }
        else:
            mesai_map[key]['StopKaydi'] = None

    # Personel nesnesine mesai bilgisi ekleyelim
    for idx, p in enumerate(personeller):
        p.mesai_data = []
        for day in days:
            key = f"{p.PersonelID}_{day['full_date']}"
            mesai_info = mesai_map.get(key, {
                "MesaiID": None,
                "MesaiTanimID": None,
                "IzinID": None,
                "Saat": "",
                "IzinAd": "",
                "Degisiklik": False,
                "PrevSaat": "",
                "PrevIzinAd": ""
            })
            mesai_info["MesaiDate"] = day['full_date']
            mesai_info["is_weekend"] = day['is_weekend']
            mesai_info["is_resmi_tatil"] = day['is_resmi_tatil']
            p.mesai_data.append(mesai_info)

    # Güvenli sabit_mesailer listesi oluştur
    sabit_mesailer = []
    try:
        for sm in SabitMesai.objects.all():
            try:
                # ara_dinlenme değerini kontrol et
                if sm.ara_dinlenme is not None:
                    float(sm.ara_dinlenme)
                sabit_mesailer.append(sm)
            except (ValueError, TypeError):
                # Problemli kayıtları atla
                continue
    except Exception:
        sabit_mesailer = []

    context = {
        "personeller": personeller,
        "mesai_options": mesai_tanimlari,
        "sabit_mesailer": sabit_mesailer,  # Modal için eklendi
        "izinler": izinler,
        "days": days,
        "birim": birim,
        "current_year": current_year,
        "current_month": current_month,
        "months": [{'value': i, 'label': calendar.month_name[i]} for i in range(1, 13)],
        "years": [year for year in range(2023, 2027)],
        "user_birimler": user_birimler,
        "selected_birim_id": selected_birim_id,
        "liste": liste.id if liste else 0,
        "donemler": donemler,
        "selected_donem": selected_donem,
        "kurumlar": kurumlar,
        "ust_birimler": ust_birimler,
        "idareciler": idareciler,
        "aciklama": liste.aciklama if liste else "",
        "mevcut_personeller": kayitlar,  # Modal için ekledik
        "ilk_liste": IlkListe.objects.filter(PersonelListesi=liste).first() if liste else None,
    }
    return render(request, 'mercis657/cizelge.html', context)

@login_required
@require_POST
def yarim_zamanli_calisma_kaydet(request, personel_id):
    personel = get_object_or_404(Personel, pk=personel_id)
    form = YarimZamanliCalismaForm(request.POST)

    if form.is_valid():
        yz = form.save(commit=False)
        yz.personel = personel
        yz.haftalik_plan = json.loads(request.POST.get("haftalik_plan", "{}"))
        yz.save()
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error", "errors": form.errors})


@login_required
def personel_listeleri(request):
    user_birimler = UserBirim.objects.filter(user=request.user).values_list('birim_id', flat=True)
    birimler = Birim.objects.filter(id__in=user_birimler)
    listeler = PersonelListesi.objects.filter(birim__in=birimler).order_by('-yil', '-ay')
    return render(request, 'mercis657/personel_listeleri.html', {
        'birimler': birimler,
        'listeler': listeler
    })

@login_required
def personel_listesi_olustur(request):
    if request.method == 'POST':
        birim_id = request.POST.get('birim_id')
        yil = int(request.POST.get('yil'))
        ay = int(request.POST.get('ay'))

        birim = get_object_or_404(Birim, id=birim_id)

        if not UserBirim.objects.filter(user=request.user, birim=birim).exists():
            return HttpResponseForbidden('Bu birim için yetkiniz yok.')

        try:
            liste, created = PersonelListesi.objects.get_or_create(birim=birim, yil=yil, ay=ay)
            if created:
                messages.success(request, 'Personel listesi oluşturuldu.')
            else:
                messages.warning(request, 'Bu ay ve birim için liste zaten var.')
        except IntegrityError:
            messages.error(request, 'Liste oluşturulurken bir hata oluştu.')

        return redirect('mercis657:personel_listeleri')

@login_required
def personel_listesi_detay(request, liste_id):
    liste = get_object_or_404(PersonelListesi, id=liste_id)
    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        return HttpResponseForbidden('Bu listeye erişim yetkiniz yok.')

    mevcut_personeller = PersonelListesiKayit.objects.filter(personel_listesi=liste).select_related('personel').order_by('sira_no', 'personel__FirstName', 'personel__LastName')
    tum_personeller = Personel.objects.all().order_by('FirstName', 'LastName')

    return render(request, 'mercis657/personel_listesi_detay.html', {
        'liste': liste,
        'mevcut_personeller': mevcut_personeller,
        'tum_personeller': tum_personeller
    })

@login_required
def tanimlamalar(request):
    kurumlar = Kurum.objects.all()
    ust_birimler = UstBirim.objects.all()
    idareciler = Idareci.objects.all()
    izinler = Izin.objects.all()
    mesai_tanimlari = Mesai_Tanimlari.objects.all().order_by('Saat')
    mesai_form = MesaiTanimForm()
    resmi_tatil_form = ResmiTatilForm()
    tatiller = ResmiTatil.objects.all().order_by('TatilTarihi')
    return render(request, "mercis657/tanimlamalar.html", {
        "kurumlar": kurumlar,
        "ust_birimler": ust_birimler,
        "idareciler": idareciler,
        "izinler": izinler,
        "mesai_tanimlari": mesai_tanimlari,
        "form": mesai_form,
        "rt_form": resmi_tatil_form,
        "tatiller": tatiller
    })

@require_POST
@login_required
def personel_cikar(request, liste_id, personel_id):
    liste = get_object_or_404(PersonelListesi, id=liste_id)
    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        return JsonResponse({'status': 'error', 'message': 'Yetkisiz işlem.'}, status=403)

    try:
        kayit = get_object_or_404(PersonelListesiKayit, liste=liste, personel_id=personel_id)
        kayit.delete()
        return JsonResponse({'status': 'success', 'message': 'Personel listeden çıkarıldı.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def onceki_donem_personel(request, donem, birim_id):
    """
    Bir önceki döneme ait personelleri getir (PersonelListesi ve PersonelListesiKayit üzerinden)
    """
    # Yetki kontrolü
    if not UserBirim.objects.filter(user=request.user, birim__BirimID=birim_id).exists():
        return JsonResponse({'status': 'error', 'message': 'Yetkisiz erişim'}, status=403)
    try:
        # donem: "YYYY-MM" veya "YYYY/MM"
        if '-' in donem:
            year, month = map(int, donem.split('-'))
        elif '/' in donem:
            year, month = map(int, donem.split('/'))
        else:
            raise Exception("Dönem formatı hatalı")
        if month == 1:
            prev_year = year - 1
            prev_month = 12
        else:
            prev_year = year
            prev_month = month - 1

        # Önceki dönem personel listesi
        liste = PersonelListesi.objects.filter(birim__BirimID=birim_id, yil=prev_year, ay=prev_month).first()
        if not liste:
            return JsonResponse([], safe=False)

        kayitlar = PersonelListesiKayit.objects.filter(liste=liste).select_related('personel')
        data = [{
            'personel_id': k.personel.PersonelID,
            'tc_kimlik': getattr(k.personel, 'PersonelTCKN', ''),
            'adi': getattr(k.personel, 'PersonelName', ''),
            'soyadi': getattr(k.personel, 'PersonelSurname', ''),
            'unvan': getattr(k.personel, 'PersonelTitle', '')
        } for k in kayitlar]

        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@require_POST
@login_required
def liste_aciklama_kaydet(request):
    try:
        data = json.loads(request.body)
        donem = data.get('donem')
        birim_id = data.get('birim_id')
        aciklama = data.get('aciklama', '')
        if not (donem and birim_id):
            return JsonResponse({'status': 'error', 'message': 'Eksik veri.'})
        year, month = map(int, donem.replace('-', '/').split('/'))
        liste = PersonelListesi.objects.filter(birim__BirimID=birim_id, yil=year, ay=month).first()
        if not liste:
            return JsonResponse({'status': 'error', 'message': 'Liste bulunamadı.'})
        liste.aciklama = aciklama
        liste.save(update_fields=['aciklama'])
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_POST
@login_required
def personel_listesi_sira_kaydet(request, liste_id):
    """
    PersonelListesiKayit sıralamasını kaydeder.
    Test için örnek:
    curl -X POST -H "Content-Type: application/json" -H "X-CSRFToken: <token>" \
      -d '{"order":[{"id":123,"sira_no":1},{"id":456,"sira_no":2}]}' \
      http://localhost:8000/mercis657/personel-listesi/1/sira-kaydet/
    """
    liste = get_object_or_404(PersonelListesi, id=liste_id)
    if not UserBirim.objects.filter(user=request.user, birim=liste.birim).exists():
        return HttpResponseForbidden('Yetkisiz işlem.')

    try:
        data = json.loads(request.body)
        order = data.get('order', [])
        if not isinstance(order, list):
            return JsonResponse({'status': 'error', 'message': 'Geçersiz veri.'}, status=400)
        # id'leri int'e çevir
        id_list = []
        id_to_sira = {}
        for item in order:
            if not isinstance(item, dict) or 'id' not in item or 'sira_no' not in item:
                return JsonResponse({'status': 'error', 'message': 'Her eleman id ve sira_no içermeli.'}, status=400)
            try:
                int_id = int(item['id'])
            except Exception:
                return JsonResponse({'status': 'error', 'message': 'ID değeri sayı olmalı.'}, status=400)
            id_list.append(int_id)
            id_to_sira[int_id] = int(item['sira_no'])

        kayit_objs = list(PersonelListesiKayit.objects.filter(id__in=id_list, liste=liste))
        if len(kayit_objs) != len(order):
            return JsonResponse({'status': 'error', 'message': 'Bazı kayıtlar bulunamadı veya yetkisiz.'}, status=400)
        # id->obj map
        id_to_obj = {k.id: k for k in kayit_objs}
        # Sıra numaralarını güncelle
        for int_id in id_list:
            obj = id_to_obj.get(int_id)
            if obj:
                obj.sira_no = id_to_sira[int_id]
        # Bulk update
        with transaction.atomic():
            PersonelListesiKayit.objects.bulk_update(kayit_objs, ['sira_no'])
            # Normalize: Sıra numaralarını 1..N olarak düzelt (opsiyonel)
            all_kayitlar = list(PersonelListesiKayit.objects.filter(liste=liste).order_by('sira_no', 'id'))
            for idx, k in enumerate(all_kayitlar, start=1):
                k.sira_no = idx
            PersonelListesiKayit.objects.bulk_update(all_kayitlar, ['sira_no'])
        return JsonResponse({'status': 'success', 'updated': len(kayit_objs)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

