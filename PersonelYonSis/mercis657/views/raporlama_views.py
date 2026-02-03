from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from PersonelYonSis.views import get_user_permissions
from ..models import Bildirim, Kurum, PersonelListesi, UserBirim, Birim, Personel, PersonelListesiKayit, Mesai, ResmiTatil, Mesai_Tanimlari, Izin, UstBirim
from PersonelYonSis.models import User
import calendar # calendar modülü eklendi
import json
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction
from django.db.models import Count, Prefetch
from decimal import Decimal
from mercis657.utils import hesapla_fazla_mesai
import os
import openpyxl
from io import BytesIO
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt

User = get_user_model()

def raporlama(request):
    bildirimler_by_birim = None
    donem = None
    kurum = None
    idare = None
    excel_url = None
    error_message = None
    info_message = None
    birimler_json = None
    birimler = Birim.objects.all()
    kurumlar = Kurum.objects.all()
    idareler = UstBirim.objects.all()

    # Form yerine GET parametrelerini doğrudan al
    donem_str = request.GET.get('donem')
    kurum = request.GET.get('kurum')
    idare = request.GET.get('idare')
    durum = request.GET.get('durum')  # "1", "0" veya ""

    toplam_kayit = 0

    # Dönem parametresi varsa veriyi çek
    if donem_str:
        try:
            # 'YYYY-MM' formatındaki stringi ayın ilk günü olan date objesine çevir
            donem = datetime.strptime(donem_str, "%Y-%m").date()

            bildirimler_query = Bildirim.objects.select_related(
                'Personel',
                'PersonelListesi__birim',
            ).filter(
                 DonemBaslangic=donem
            )

            if kurum:
                bildirimler_query = bildirimler_query.filter(
                    PersonelListesi__birim__Kurum__ad=kurum
                )
            if idare:
                bildirimler_query = bildirimler_query.filter(
                    PersonelListesi__birim__UstBirim__ad=idare
                )

            if durum == "1":
                bildirimler_query = bildirimler_query.filter(OnayDurumu=1)
            elif durum == "0":
                bildirimler_query = bildirimler_query.filter(OnayDurumu=0)

            # Bildirimleri çek ve PersonelListesi__birim'e göre grupla (Birim modelinde doğrudan bildirim_set yok)
            birim_ids_with_calisma = list(bildirimler_query.values_list('PersonelListesi__birim', flat=True).distinct())

            birimler_with_bildirimler = Birim.objects.filter(BirimID__in=birim_ids_with_calisma).order_by('BirimAdi')

            # Bildirimleri sıralı çek ve Python tarafında birim id'lerine göre grupla
            bildirimler_list = list(bildirimler_query.order_by('Personel__PersonelName').select_related('Personel', 'PersonelListesi'))
            bildirimler_map = {}
            for b in bildirimler_list:
                try:
                    birim_obj = getattr(b.PersonelListesi, 'birim', None)
                    birim_id = birim_obj.BirimID if birim_obj is not None else None
                except Exception:
                    birim_id = None
                if birim_id is None:
                    continue
                bildirimler_map.setdefault(birim_id, []).append(b)

            # Birimleri JSON'a çevir
            birimler_json = json.dumps([
                {'BirimID': b.BirimID, 'BirimAdi': b.BirimAdi} for b in birimler_with_bildirimler
            ])

            bildirimler_by_birim = []
            for birim in birimler_with_bildirimler:
                items = bildirimler_map.get(birim.BirimID, [])
                if items:
                    onaylanmis = sum(1 for c in items if c.OnayDurumu == 1)
                    beklemede = sum(1 for c in items if c.OnayDurumu == 0)
                    kilitli = sum(1 for c in items if c.MutemetKilit == True)
                    bildirimler_by_birim.append({
                        'birim': birim,
                        'bildirimler': items,
                        'personel_sayisi': len(set([getattr(c.Personel, 'PersonelID', None) for c in items])),
                        'onaylanmis_sayisi': onaylanmis,
                        'beklemede_sayisi': beklemede,
                        'kilitli_sayisi': kilitli,
                        })

            toplam_kayit = sum(len(birim['bildirimler']) for birim in bildirimler_by_birim)

            # Excel indirme linki oluştur
            if bildirimler_by_birim:
                 excel_url = reverse('mercis657:export_raporlama_excel')
                 excel_url += f'?donem={donem.strftime("%Y-%m")}'

                 if kurum:
                     excel_url += f'&kurum={kurum}'
                 if durum:
                     excel_url += f'&durum={durum}'
                 if idare:
                     excel_url += f'&idare={idare}'

            if toplam_kayit == 0:
                 info_message = "Seçilen dönem ve kuruma ait hizmet sunum çalışması bulunamadı."

        except ValueError:
             error_message = "Geçersiz dönem formatı seçildi."
        except Exception as e:
            # Hata yönetimi
            error_message = f"Rapor oluşturulurken bir hata oluştu: {str(e)}"
            print(f"Raporlama Hatası: {e}") # Konsola yazdır (geliştirme için)
    else:
        # İlk sayfa yüklemesi veya dönem seçilmemişse
        info_message = "Lütfen bir dönem ve isteğe bağlı kriterlerinizi seçerek raporlayın."

    kod_duzenleme_yetkisi = request.user.has_permission('ÇS 657 Birim Kodlarını Düzenleyebilir')
    bildirim_sayfasi_yetkisi = request.user.has_permission('ÇS 657 Bildirim İşlemleri')
    context = {
        # 'form': form, # Form kaldırıldı
        'bildirimler_by_birim': bildirimler_by_birim,
        'birimler': birimler, # Kurum seçimi için tüm birimleri geçiyoruz (kurum listesi çekmek için)
        'birimler_json': birimler_json,
        'kurumlar': kurumlar, # Kurum seçimi için tüm kurumları geçiyoruz
        'idareler': idareler,
        'selected_donem': donem_str, # Şablona dönemin string halini gönderelim ki selectbox'ta seçili kalsın
        'selected_kurum': kurum,
        'selected_idare': idare,
        'selected_durum': durum,
        'excel_url': excel_url,
        # 'is_form_valid': is_form_valid, # Form kaldırıldı
        'error_message': error_message,
        'info_message': info_message,
        'toplam_kayit': toplam_kayit,
        'kod_duzenleme_yetkisi': kod_duzenleme_yetkisi,
        'bildirim_sayfasi_yetkisi': bildirim_sayfasi_yetkisi,
    }
    return render(request, 'mercis657/raporlama.html', context)

def export_raporlama_excel(request):
    """
    Bildirimler bildirim koduna göre gruplandırılır ve her fazla mesai türü için ayrı satır oluşturulur.
    """
    donem_str = request.GET.get('donem')
    kurum = request.GET.get('kurum')
    idare = request.GET.get('idare')
    durum = request.GET.get('durum')

    if not donem_str:
        messages.error(request, "Lütfen bir dönem seçin.")
        return redirect('mercis657:raporlama')

    try:
        donem = datetime.strptime(donem_str, "%Y-%m").date()
    except ValueError:
        messages.error(request, "Geçersiz dönem formatı.")
        return redirect('mercis657:raporlama')

    # Fetch Bildirim queryset for the period
    bildirimler_qs = Bildirim.objects.select_related('Personel', 'PersonelListesi__birim').filter(DonemBaslangic=donem)
    if kurum:
        bildirimler_qs = bildirimler_qs.filter(PersonelListesi__birim__Kurum__ad=kurum)
    if idare:
        bildirimler_qs = bildirimler_qs.filter(PersonelListesi__birim__UstBirim__ad=idare)
    if durum == "1":
        bildirimler_qs = bildirimler_qs.filter(OnayDurumu=1)
    elif durum == "0":
        bildirimler_qs = bildirimler_qs.filter(OnayDurumu=0)

    if not bildirimler_qs.exists():
        messages.warning(request, "Seçilen filtreye uygun veri bulunamadı.")
        return redirect('mercis657:raporlama')

    # Şablon dosyasını yükle
    template_path = os.path.join(settings.STATIC_ROOT, 'excels', 'FazlaMesaiSablon.xlsx')
    if not os.path.exists(template_path):
        template_path = os.path.join(settings.BASE_DIR, 'static', 'excels', 'FazlaMesaiSablon.xlsx')
    
    try:
        workbook = openpyxl.load_workbook(template_path)
        worksheet = workbook.active
    except FileNotFoundError:
        messages.error(request, f"Şablon dosyası bulunamadı: {template_path}")
        return redirect('mercis657:raporlama')

    # Stil tanımlamaları
    from openpyxl.styles import Border, Side, PatternFill
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    current_row = 2  # Veri yazmaya 2. satırdan başla (başlık satırını atla)

    # Bildirimleri personel bazında grupla
    for bildirim in bildirimler_qs.order_by('PersonelListesi__birim__BirimAdi', 'Personel__PersonelName'):
        try:
            # Personel ve birim bilgilerini al
            personel = bildirim.Personel
            birim = bildirim.PersonelListesi.birim
            
            personel_tckn = getattr(personel, 'PersonelTCKN', '')
            personel_ad = getattr(personel, 'PersonelName', '')
            personel_soyisim = getattr(personel, 'PersonelSurname', '')
            birim_adi = getattr(birim, 'BirimAdi', '')
            kadro_durumu = getattr(personel, 'KadroDurumu', '')
            
            # Birim kodlarını al
            normal_nobet_kodu = getattr(birim, 'NormalNobetKodu', '1')
            bayram_nobet_kodu = getattr(birim, 'BayramNobetKodu', '')
            riskli_normal_nobet_kodu = getattr(birim, 'RiskliNormalNobetKodu', '')
            riskli_bayram_nobet_kodu = getattr(birim, 'RiskliBayramNobetKodu', '')
            
            # Gece birim kodlarını al
            normal_gece_nobet_kodu = getattr(birim, 'NormalGeceNobetKodu', '')
            bayram_gece_nobet_kodu = getattr(birim, 'BayramGeceNobetKodu', '')
            riskli_normal_gece_nobet_kodu = getattr(birim, 'RiskliNormalGeceNobetKodu', '')
            riskli_bayram_gece_nobet_kodu = getattr(birim, 'RiskliBayramGeceNobetKodu', '')
            
            # Her bir fazla mesai/icap türü için kontrol yap ve ayrı satır oluştur
            fazla_mesai_list = [
                ('NormalFazlaMesai', normal_nobet_kodu, bildirim.NormalFazlaMesai),
                ('BayramFazlaMesai', bayram_nobet_kodu, bildirim.BayramFazlaMesai),
                ('RiskliNormalFazlaMesai', riskli_normal_nobet_kodu, bildirim.RiskliNormalFazlaMesai),
                ('RiskliBayramFazlaMesai', riskli_bayram_nobet_kodu, bildirim.RiskliBayramFazlaMesai),
                # Gece
                ('GeceNormalFazlaMesai', normal_gece_nobet_kodu, bildirim.GeceNormalFazlaMesai),
                ('GeceBayramFazlaMesai', bayram_gece_nobet_kodu, bildirim.GeceBayramFazlaMesai),
                ('GeceRiskliNormalFazlaMesai', riskli_normal_gece_nobet_kodu, bildirim.GeceRiskliNormalFazlaMesai),
                ('GeceRiskliBayramFazlaMesai', riskli_bayram_gece_nobet_kodu, bildirim.GeceRiskliBayramFazlaMesai),
                # İcap
                ('NormalIcap', '17', bildirim.NormalIcap),
                ('BayramIcap', '18', bildirim.BayramIcap),
            ]

            for fm_type, birim_kodu, value in fazla_mesai_list:
                # Değerin boş veya 0'dan büyük olup olmadığını kontrol et
                try:
                    numeric_value = float(value) if value is not None else 0
                except (TypeError, ValueError):
                    numeric_value = 0
                
                # Sadece geçerli değerler ve 0'dan büyük değerler için satır oluştur
                if value is not None and value != '' and numeric_value > 0 and birim_kodu:
                    # Yeni satır için hücrelere verileri yaz
                    worksheet.cell(row=current_row, column=1, value=personel_tckn)
                    worksheet.cell(row=current_row, column=2, value=personel_ad)
                    worksheet.cell(row=current_row, column=3, value=personel_soyisim)
                    worksheet.cell(row=current_row, column=4, value=birim_kodu)  # İlgili birim kodunu 5. sütuna yaz
                    worksheet.cell(row=current_row, column=5, value=numeric_value)  # Değeri 6. sütuna yaz
                    worksheet.cell(row=current_row, column=7, value=birim_adi)  # Birim adını 7. sütuna yaz
                    worksheet.cell(row=current_row, column=8, value=fm_type)  # Fazla mesai türünü 8. sütuna yaz

                    # Sınır ve dolgu uygula
                    for col in range(1, 9):
                        worksheet.cell(row=current_row, column=col).border = thin_border
                        if kadro_durumu == "Geçici Gelen":
                            worksheet.cell(row=current_row, column=col).fill = green_fill

                    current_row += 1  # Bir sonraki satıra geç

        except Exception as e:
            print(f"Bildirim işlenirken hata: {e}")
            continue

    # Excel dosyasını kaydet
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = f"FazlaMesaiBildirimleri_{donem_str}.xlsx"
    response['Content-Disposition'] = f'inline; filename="{file_name}"'
    return response


@require_POST
@login_required
def update_birim_kodlari_toplu(request):
    """Endpoint: Toplu birim kodlarını günceller.
    Beklenen payload: {'changes': [{'birim_id': id, 'NormalNobetKodu': val, 'BayramNobetKodu': val,...}, ...]}
    """
    try:
        data = json.loads(request.body.decode('utf-8'))
        changes = data.get('changes', [])
        
        if not changes:
            return JsonResponse({'status':'error','message':'Güncellenecek veri bulunamadı'}, status=400)

        # Permission check: kullanıcı birim bilgilerini düzenleyebilmeli
        if not request.user.has_permission('ÇS 657 Birim Kodlarını Düzenleyebilir'):
            return JsonResponse({'status':'error','message':'Yetkiniz yok'}, status=403)

        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for change in changes:
                try:
                    birim_id = change.get('birim_id')
                    if not birim_id:
                        errors.append(f"Birim ID bulunamadı: {change}")
                        continue
                        
                    birim = Birim.objects.get(BirimID=birim_id)
                    
                    # Update allowed fields
                    for field in ['NormalNobetKodu','BayramNobetKodu','RiskliNormalNobetKodu','RiskliBayramNobetKodu','NormalGeceNobetKodu','BayramGeceNobetKodu','RiskliNormalGeceNobetKodu','RiskliBayramGeceNobetKodu']:
                        if field in change:
                            val = change[field]
                            setattr(birim, field, val)
                    
                    birim.save()
                    updated_count += 1
                    
                except Birim.DoesNotExist:
                    errors.append(f"Birim bulunamadı: {birim_id}")
                except Exception as e:
                    errors.append(f"Birim {birim_id} güncellenirken hata: {str(e)}")
        
        if errors:
            message = f"{updated_count} birim güncellendi. Hatalar: {'; '.join(errors)}"
            return JsonResponse({'status':'partial','message':message, 'errors':errors})
        else:
            return JsonResponse({'status':'success','message':f'{updated_count} birim başarıyla güncellendi'})
            
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)}, status=500)


@require_POST
@login_required
def kilit_tekil(request):
    """Endpoint: Verilen birim için seçilen dönemdeki Bildirimlerin MutemetKilit alanını toggle eder.
    Payload: {'birim_id': id, 'donem': 'YYYY-MM'}
    """
    try:
        data = json.loads(request.body.decode('utf-8'))
        birim_id = data.get('birim_id')
        donem_str = data.get('donem')
        donem = datetime.strptime(donem_str, "%Y-%m").date()

        bildirimler = Bildirim.objects.filter(PersonelListesi__birim__BirimID=birim_id, DonemBaslangic=donem)
        if not bildirimler.exists():
            return JsonResponse({'status':'error','message':'Bildirim bulunamadı'}, status=404)

        with transaction.atomic():
            for b in bildirimler:
                if b.MutemetKilit == True:
                    b.MutemetKilit = False
                    b.MutemetKilitUser = None
                    b.MutemetKilitTime = None
                else:
                    b.MutemetKilit = True
                    b.MutemetKilitUser = request.user
                    b.MutemetKilitTime = timezone.now()
                b.save()

        return JsonResponse({'status':'success','message':'Kilit durumu güncellendi'})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)}, status=500)


@require_POST
@login_required
def kilit_toplu(request):
    """Endpoint: Tüm bildirimleri döneme göre kilitle veya aç.
    Payload: {'donem': 'YYYY-MM', 'action': 'lock'|'unlock'}
    """
    try:
        data = json.loads(request.body.decode('utf-8'))
        donem_str = data.get('donem')
        kurum = data.get('kurum')
        idare = data.get('idare')
        durum = data.get('durum')
        action = data.get('action')
        donem = datetime.strptime(donem_str, "%Y-%m").date()

        bildirimler = Bildirim.objects.filter(DonemBaslangic=donem)
        
        if kurum:
            bildirimler = bildirimler.filter(PersonelListesi__birim__Kurum__ad=kurum)
        if idare:
            bildirimler = bildirimler.filter(PersonelListesi__birim__UstBirim__ad=idare)
        
        # Durum kontrolü (hem string hem int gelebilir)
        if str(durum) == "1":
            bildirimler = bildirimler.filter(OnayDurumu=1)
        elif str(durum) == "0":
            bildirimler = bildirimler.filter(OnayDurumu=0)
        if not bildirimler.exists():
            return JsonResponse({'status':'error','message':'Bildirim bulunamadı'}, status=404)

        with transaction.atomic():
            for b in bildirimler:
                if action == 'unlock':
                    b.MutemetKilit = False
                    b.MutemetKilitUser = None
                    b.MutemetKilitTime = None
                else:
                    b.MutemetKilit = True
                    b.MutemetKilitUser = request.user
                    b.MutemetKilitTime = timezone.now()
                b.save()

        return JsonResponse({'status':'success','message':'Toplu işlem tamamlandı'})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)}, status=500)
